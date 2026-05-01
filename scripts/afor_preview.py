#!/usr/bin/env python3
"""
afor_preview.py — AFOR Import Translator / Pre-processor (Guideline deliverable)

Usage:
    python scripts/afor_preview.py path/to/afor.xlsx

Outputs a human-readable Markdown preview of all extracted AFOR fields.
Flags any field that is empty/N/A with ⚠️ EMPTY.

Dependencies: openpyxl only (zero web dependencies, stdlib only).
Security: local files only — remote URLs are rejected.
"""

from __future__ import annotations

import os
import sys

# ── Security: reject remote URLs ─────────────────────────────────────────────
def _validate_input_path(path: str) -> str:
    if path.startswith("http://") or path.startswith("https://") or path.startswith("ftp://"):
        print("ERROR: Remote URLs are not accepted. Provide a local file path only.", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(path):
        print(f"ERROR: File not found: {path}", file=sys.stderr)
        sys.exit(1)
    if not os.access(path, os.R_OK):
        print(f"ERROR: File not readable: {path}", file=sys.stderr)
        sys.exit(1)
    return os.path.abspath(path)


# ── Label map (mirrors TypeScript FIELD_LABELS in afor-utils.ts) ──────────────
FIELD_LABELS: dict[str, str] = {
    "responder_type": "Type of Responder",
    "fire_station_name": "Responding Fire Station",
    "notification_date": "Date of Notification",
    "notification_time": "Time of Notification",
    "region": "Region",
    "province": "Province",
    "city": "City / Municipality",
    "address": "Complete Address",
    "landmark": "Nearest Landmark",
    "caller_info": "Caller / Reporter",
    "receiver": "Personnel Who Received Call",
    "engine": "Engine Number",
    "time_dispatched": "Time Dispatched",
    "time_arrived": "Time Arrived",
    "response_time": "Total Response Time (minutes)",
    "distance_km": "Distance from Station (km)",
    "alarm_level": "Highest Alarm Level",
    "time_returned": "Time Returned",
    "gas_liters": "Total Gas Consumed (liters)",
    "classification": "Classification",
    "category": "Category / Type",
    "owner": "Owner / Establishment Name",
    "description": "Property Description",
    "origin": "Area of Origin",
    "stage": "Stage of Fire Upon Arrival",
    "extent": "Extent of Damage",
    "struct_aff": "No. of Structures Affected",
    "house_aff": "No. of Households Affected",
    "fam_aff": "No. of Families Affected",
    "indiv_aff": "No. of Individuals Affected",
    "vehic_aff": "No. of Vehicles Affected",
    "res_bfp_truck": "BFP Fire Trucks",
    "res_lgu_truck": "LGU Fire Trucks",
    "res_vol_truck": "Volunteer Fire Trucks",
    "res_bfp_amb": "BFP Ambulance",
    "res_non_amb": "Non-BFP Ambulance",
    "res_bfp_resc": "BFP Rescue Unit",
    "res_non_resc": "Non-BFP Rescue Unit",
    "icp_present": "ICP Present",
    "icp_location": "ICP Location",
    "pod_commander": "Engine Commander",
    "pod_shift": "Shift-in-Charge",
    "pod_nozzleman": "Nozzleman",
    "pod_lineman": "Lineman",
    "pod_crew": "Engine Crew",
    "pod_dpo": "Driver / Pump Operator",
    "pod_safety": "Safety Officer",
    "pod_fire_arson_inv": "Fire and Arson Investigator",
    "narrative": "Narrative Report",
    "problems": "Problems Encountered",
    "recommendations": "Recommendations",
    "disposition": "Disposition",
    "prepared_by": "Prepared By",
    "noted_by": "Noted By",
}

TIMELINE_LABELS: dict[str, str] = {
    "foua": "1st Alarm – FOUA (First On Upon Arrival)",
    "alarm_1st": "1st Alarm",
    "alarm_2nd": "2nd Alarm",
    "alarm_3rd": "3rd Alarm",
    "alarm_4th": "4th Alarm",
    "alarm_5th": "5th Alarm",
    "tf_alpha": "Task Force Alpha",
    "tf_bravo": "Task Force Bravo",
    "tf_charlie": "Task Force Charlie",
    "tf_delta": "Task Force Delta",
    "general": "General Alarm",
    "fuc": "Fire Under Control (FUC)",
    "fo": "Fire Out (FO)",
}


def _is_empty(val: object) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    return s in ("", "N/A", "NA", "n/a", "None", "none", "-", "—")


def _fmt(val: object, flag_empty: bool = True) -> str:
    if _is_empty(val):
        return "⚠️ EMPTY" if flag_empty else "—"
    return str(val).strip()


def _print_section(title: str, fields: list[tuple[str, object]]) -> None:
    print(f"\n## {title}\n")
    for label, value in fields:
        marker = _fmt(value)
        print(f"- **{label}**: {marker}")


# ── Inline parser (mirrors BfpXlsxParser from regional.py) ───────────────────
import re
import io
from datetime import datetime, timedelta
from typing import Any

_COORD_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _col_to_idx(letters: str) -> int:
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _cell_str(ws: Any, coord: str) -> str:
    try:
        v = ws[coord].value
    except Exception:
        return ""
    if v is None:
        return ""
    return str(v).strip()


def _find_structural_markers(ws: Any) -> tuple[int | None, int | None]:
    title_row = None
    section_row = None
    for row in range(1, 90):
        a_val = _cell_str(ws, f"A{row}").upper()
        b_val = _cell_str(ws, f"B{row}").upper()
        combined = f"{a_val} {b_val}".strip()
        if title_row is None and "AFTER FIRE OPERATIONS REPORT" in combined:
            title_row = row
        if section_row is None and "A. RESPONSE DETAILS" in combined:
            section_row = row
        if title_row is not None and section_row is not None:
            break
    return title_row, section_row


def _infer_offset(ws: Any) -> int:
    title_row, section_row = _find_structural_markers(ws)
    if title_row is None:
        return 0
    offset = title_row - 14
    if section_row is not None and (section_row - 18) != offset:
        return 0
    return offset


def _get(ws: Any, coord: str, offset: int) -> Any:
    m = _COORD_RE.match(coord.upper())
    if m and offset != 0:
        col, row = m.groups()
        coord = f"{col}{max(1, int(row) + offset)}"
    try:
        v = ws[coord].value
    except Exception:
        return None
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    return v


def _is_marked(ws: Any, coord: str, offset: int) -> bool:
    val = str(_get(ws, coord, offset) or "").strip().lower()
    return val in ("x", "1", "true", "v", "✓", "✔", "/")


def parse_afor_worksheet(ws: Any) -> dict[str, Any]:
    offset = _infer_offset(ws)

    def get(c: str) -> Any:
        return _get(ws, c, offset)

    def marked(c: str) -> bool:
        return _is_marked(ws, c, offset)

    # Responder type
    responder_type = (
        "First Responder"
        if marked("B20")
        else ("Augmenting Team" if marked("B21") else "First Responder")
    )

    # Classification
    classification = "Structural"
    cat_val = get("D48")
    if marked("B49"):
        classification = "Non-Structural"
        cat_val = get("D49")
    elif marked("B50"):
        classification = "Transportation"
        cat_val = get("D50")

    # Stage of fire
    stage = get("D54") or get("B54")
    if stage and "pick from dropdown" in str(stage).lower():
        stage = None

    # Extent of damage
    extent = "None / Minor"
    for coord, label in [
        ("B57", "Confined to Object"),
        ("B58", "Confined to Room"),
        ("B59", "Confined to Structure"),
        ("B60", "Total Loss"),
        ("B61", "Extended Beyond Structure"),
    ]:
        if marked(coord):
            extent = label
            break

    # Problems
    prob_map = {
        "B195": "Inaccurate address",
        "B196": "Geographically challenged",
        "B197": "Road conditions",
        "B198": "Road under construction",
        "B199": "Traffic congestion",
        "B200": "Road accidents",
        "B201": "Vehicles failure to yield",
        "B202": "Natural Disasters",
        "B203": "Civil Disturbance",
        "B204": "Uncooperative or panicked residents",
        "B205": "Safety and security threats",
        "B206": "Property security or owner delays",
        "B207": "Engine failure",
        "B208": "Uncooperative fire auxiliary",
        "B209": "Poor water supply access",
        "B210": "Intense heat and smoke",
        "B211": "Structural hazards",
        "B212": "Equipment malfunction",
        "B213": "Poor inter-agency coordination",
        "B214": "Radio communication breakdown",
        "B215": "HazMat risks",
        "B216": "Physical exhaustion and injuries",
        "B217": "Emotional and psychological effects",
        "B218": "Community complaints",
        "B219": "Others",
    }
    problems = [label for coord, label in prob_map.items() if marked(coord)]

    icp_present = marked("B102")

    # Narrative
    narrative_lines = []
    for r in range(160, 191):
        line = get(f"B{r}")
        if line:
            narrative_lines.append(str(line))

    # Other personnel
    others = []
    for r in range(124, 133):
        name = get(f"B{r}")
        rem = get(f"E{r}")
        name_str = str(name).strip() if name else ""
        if name_str and name_str.upper() not in ("N/A", "NA", "-", "—", "N.A.", ""):
            others.append({"name": name_str, "designation": str(rem).strip() if rem else ""})

    # Timeline
    def tl_entry(d_coord: str, e_coord: str, f_coord: str) -> dict:
        return {
            "time": get(d_coord),
            "date": get(e_coord),
            "commander": get(f_coord),
        }

    timeline = {
        "foua": tl_entry("D88", "E88", "F88"),
        "alarm_1st": tl_entry("D89", "E89", "F89"),
        "alarm_2nd": tl_entry("D90", "E90", "F90"),
        "alarm_3rd": tl_entry("D91", "E91", "F91"),
        "alarm_4th": tl_entry("D92", "E92", "F92"),
        "alarm_5th": tl_entry("D93", "E93", "F93"),
        "tf_alpha": tl_entry("D94", "E94", "F94"),
        "tf_bravo": tl_entry("D95", "E95", "F95"),
        "tf_charlie": tl_entry("D96", "E96", "F96"),
        "tf_delta": tl_entry("D97", "E97", "F97"),
        "general": tl_entry("D98", "E98", "F98"),
        "fuc": tl_entry("D99", "E99", "F99"),
        "fo": tl_entry("D100", "E100", "F100"),
    }

    return {
        "responder_type": responder_type,
        "fire_station_name": get("D20") if responder_type == "First Responder" else get("D21"),
        "notification_date": get("D22"),
        "notification_time": get("D23"),
        "region": get("D24"),
        "province": get("D25"),
        "city": get("D26"),
        "address": get("D27"),
        "landmark": get("D28"),
        "caller_info": get("D29"),
        "receiver": get("D30"),
        "engine": get("D31"),
        "time_dispatched": get("D34"),
        "time_arrived": get("D37"),
        "response_time": get("D40"),
        "distance_km": get("D41"),
        "alarm_level": get("D42"),
        "time_returned": get("D43"),
        "gas_liters": get("D44"),
        "classification": classification,
        "category": cat_val,
        "owner": get("D51"),
        "description": get("D52"),
        "origin": get("D53"),
        "stage": stage,
        "extent": extent,
        "struct_aff": get("D62"),
        "house_aff": get("D63"),
        "fam_aff": get("D64"),
        "indiv_aff": get("D65"),
        "vehic_aff": get("D66"),
        "res_bfp_truck": get("D70"),
        "res_lgu_truck": get("D71"),
        "res_vol_truck": get("D72"),
        "res_bfp_amb": get("D73"),
        "res_non_amb": get("D74"),
        "res_bfp_resc": get("D75"),
        "res_non_resc": get("D76"),
        "icp_present": icp_present,
        "icp_location": get("D102") if icp_present else None,
        "pod_commander": get("D114"),
        "pod_shift": get("D115"),
        "pod_nozzleman": get("D116"),
        "pod_lineman": get("D117"),
        "pod_crew": get("D118"),
        "pod_dpo": get("D119"),
        "pod_safety": get("D120"),
        "pod_fire_arson_inv": get("D121"),
        "others": others,
        "timeline": timeline,
        "narrative": "\n".join(narrative_lines),
        "problems": problems,
        "recommendations": get("B222"),
        "disposition": get("B229"),
        "prepared_by": get("C238"),
        "noted_by": get("F238"),
    }


def render_markdown(data: dict[str, Any], filepath: str) -> str:
    lines: list[str] = []
    lines.append(f"# AFOR Preview: `{os.path.basename(filepath)}`\n")
    lines.append("> Generated by `scripts/afor_preview.py` — verify before uploading.\n")

    def field(label: str, val: object) -> str:
        marker = _fmt(val)
        return f"- **{label}**: {marker}"

    # A. Response Details
    lines.append("\n## A. Response Details\n")
    for key in ["responder_type", "fire_station_name", "notification_date", "notification_time",
                "region", "province", "city", "address", "landmark", "caller_info", "receiver",
                "engine", "time_dispatched", "time_arrived", "response_time", "distance_km",
                "alarm_level", "time_returned", "gas_liters"]:
        lines.append(field(FIELD_LABELS.get(key, key), data.get(key)))

    # B. Classification
    lines.append("\n## B. Classification\n")
    for key in ["classification", "category", "owner", "description", "origin", "stage", "extent"]:
        lines.append(field(FIELD_LABELS.get(key, key), data.get(key)))

    # C. Affected
    lines.append("\n## C. Affected\n")
    for key in ["struct_aff", "house_aff", "fam_aff", "indiv_aff", "vehic_aff"]:
        lines.append(field(FIELD_LABELS.get(key, key), data.get(key)))

    # D. Alarm Timeline
    lines.append("\n## D. Alarm Timeline\n")
    for tl_key, tl_label in TIMELINE_LABELS.items():
        entry = data["timeline"].get(tl_key, {})
        t = entry.get("time") if isinstance(entry, dict) else None
        d = entry.get("date") if isinstance(entry, dict) else None
        cmd = entry.get("commander") if isinstance(entry, dict) else None
        if not _is_empty(t) or not _is_empty(d):
            parts = [f"Time: {_fmt(t, False)}", f"Date: {_fmt(d, False)}"]
            if not _is_empty(cmd):
                parts.append(f"Commander: {_fmt(cmd, False)}")
            lines.append(f"- **{tl_label}**: {' | '.join(parts)}")
        else:
            lines.append(f"- **{tl_label}**: —")

    # E. ICP
    lines.append("\n## E. Incident Command Post\n")
    lines.append(field("ICP Present", "Yes" if data.get("icp_present") else "No"))
    lines.append(field(FIELD_LABELS["icp_location"], data.get("icp_location")))

    # F. Resources
    lines.append("\n## F. Resources Deployed\n")
    for key in ["res_bfp_truck", "res_lgu_truck", "res_vol_truck", "res_bfp_amb",
                "res_non_amb", "res_bfp_resc", "res_non_resc"]:
        lines.append(field(FIELD_LABELS.get(key, key), data.get(key)))

    # G. Personnel on Duty
    lines.append("\n## G. Personnel on Duty\n")
    for key in ["pod_commander", "pod_shift", "pod_nozzleman", "pod_lineman",
                "pod_crew", "pod_dpo", "pod_safety", "pod_fire_arson_inv"]:
        lines.append(field(FIELD_LABELS.get(key, key), data.get(key)))
    if data.get("others"):
        lines.append("\n**Other Personnel at Scene:**\n")
        lines.append("| Name | Designation / Agency |")
        lines.append("|------|----------------------|")
        for p in data["others"]:
            name = _fmt(p.get("name"), False) or "—"
            desig = _fmt(p.get("designation"), False) or "—"
            lines.append(f"| {name} | {desig} |")

    # I. Narrative
    lines.append("\n## I. Narrative Report\n")
    if _is_empty(data.get("narrative")):
        lines.append("⚠️ EMPTY")
    else:
        for i, para in enumerate(str(data["narrative"]).split("\n"), start=1):
            if para.strip():
                lines.append(f"{i}. {para.strip()}")

    # J. Problems
    lines.append("\n## J. Problems Encountered\n")
    problems = data.get("problems") or []
    if not problems:
        lines.append("⚠️ EMPTY (no problems checked)")
    else:
        for p in problems:
            lines.append(f"- ✅ {p}")

    # K–L
    lines.append("\n## K. Recommendations\n")
    lines.append(field(FIELD_LABELS["recommendations"], data.get("recommendations")))
    lines.append("\n## L. Disposition & Signatories\n")
    lines.append(field(FIELD_LABELS["disposition"], data.get("disposition")))
    lines.append(field(FIELD_LABELS["prepared_by"], data.get("prepared_by")))
    lines.append(field(FIELD_LABELS["noted_by"], data.get("noted_by")))

    # Summary: empty field count
    flat_vals = [v for k, v in data.items() if k not in ("timeline", "others", "problems")]
    empty_count = sum(1 for v in flat_vals if _is_empty(v))
    lines.append(f"\n---\n\n> **{empty_count} field(s) flagged as EMPTY** — review before uploading.\n")

    return "\n".join(lines)


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python scripts/afor_preview.py path/to/afor.xlsx", file=sys.stderr)
        sys.exit(1)

    filepath = _validate_input_path(sys.argv[1])

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in (".xlsx", ".xls"):
        print(f"ERROR: Only .xlsx files are supported. Got: {ext}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filepath, data_only=True)

    # Find AFOR sheet
    ws = None
    for name in wb.sheetnames:
        sheet = wb[name]
        if "AFOR" in name.upper() and "WILDLAND" not in name.upper():
            ws = sheet
            break
    if ws is None:
        # Fallback: scan for structural markers
        for name in wb.sheetnames:
            sheet = wb[name]
            t, s = _find_structural_markers(sheet)
            if t and s:
                ws = sheet
                break
    if ws is None:
        ws = wb.active
        print("WARNING: Could not detect AFOR sheet — using active sheet.", file=sys.stderr)

    data = parse_afor_worksheet(ws)
    md = render_markdown(data, filepath)
    print(md)


if __name__ == "__main__":
    main()
