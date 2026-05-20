"""Generate TypeScript city constants and SQL inserts for all PH regions."""

import openpyxl
import re

wb = openpyxl.load_workbook(
    r"c:/proj/WIMS-BFP-PROTOTYPE/AFORs/Proposed-New-AFOR_Nov-2025 (1).xlsx",
    read_only=True,
    data_only=True,
)
ws = wb["List"]

col_c_vals = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[2]:
        col_c_vals.add(str(row[2]).strip())

region_codes = {
    "Region_1",
    "Region_2",
    "Region_3",
    "Region_4A",
    "Region_4B",
    "Region_5",
    "Region_6",
    "Region_7",
    "Region_8",
    "Region_9",
    "Region_10",
    "Region_11",
    "Region_12",
    "BARMM",
    "CAR",
    "CARAGA",
    "NCR",
    "NIR",
}

data = {}
cur_region = None
cur_province = None
for row in ws.iter_rows(min_row=2, values_only=True):
    e_val = row[4]
    if not e_val:
        continue
    estr = str(e_val).strip()
    if estr in region_codes:
        cur_region = estr
        data.setdefault(cur_region, {})
        cur_province = None
    elif estr in col_c_vals:
        cur_province = estr.strip()
        if cur_region:
            data[cur_region].setdefault(cur_province, [])
    elif cur_region and cur_province:
        data[cur_region][cur_province].append(estr)

# Manual overrides for hyphenated names and typos
OVERRIDES = {
    "Can_avid": "Can-avid",
    "Lal_lo": "Lal-lo",
    "Alang_alang": "Alang-alang",
    "Matag_ob": "Matag-ob",
    "Palo_Leyte": "Palo",
    "MacAtrhur": "MacArthur",
    "Merceds": "Mercedes",
    "Conception": "Concepcion",
    "Santa_Barbarra": "Santa Barbara",
    "Santo_Domigo": "Santo Domingo",
    "Trece_Martirez_City": "Trece Martires City",
    "General_Trias": "General Trias City",
    "Libangon": "Libagon",
    "Laua_an": "Laua-an",
    "Anniy": "Anini-y",
    "Sierra_bullones": "Sierra Bullones",
    "Lope_de_Vega": "Lope de Vega",
    "Tagoloan_II": "Tagoloan II",
}

KNOWN_SUFFIXES = {
    "a",
    "b",
    "c",
    "g",
    "i",
    "k",
    "l",
    "m",
    "p",
    "q",
    "r",
    "s",
    "t",
    "z",
    "ak",
    "al",
    "an",
    "ba",
    "bo",
    "ca",
    "cn",
    "cs",
    "di",
    "do",
    "es",
    "ii",
    "in",
    "is",
    "lu",
    "mo",
    "mp",
    "nc",
    "ne",
    "nl",
    "no",
    "ns",
    "nv",
    "om",
    "pa",
    "sc",
    "sk",
    "sl",
    "ws",
    "zs",
    "adn",
    "ads",
    "cvp",
    "ddn",
    "dds",
    "ldn",
    "lds",
    "sdn",
    "sds",
    "zdn",
    "zds",
}


def clean(raw):
    if raw in OVERRIDES:
        return OVERRIDES[raw]
    m = re.search(r"_([a-z]+)$", raw)
    if m and m.group(1) in KNOWN_SUFFIXES:
        raw = raw[: m.start()]
    return raw.replace("_", " ").strip()


def ts_array(cities):
    cleaned = [clean(c) for c in cities]
    # Use double-quotes for entries containing apostrophes
    parts = []
    for c in cleaned:
        if "'" in c:
            parts.append(f'"{c}"')
        else:
            parts.append(f"'{c}'")
    return "[" + ", ".join(parts) + "]"


def print_const(name, province_map):
    """province_map: dict of db_province_name -> list of raw excel city names"""
    lines = [f"const {name}: Record<string, string[]> = {{"]
    for dp, raw_cities in province_map.items():
        arr = ts_array(raw_cities)
        lines.append(f"  '{dp}': {arr},")
    lines.append("};")
    return "\n".join(lines)


# Compute merged/remapped city lists
nir_neg_occ = data["NIR"]["Negros_Occidental"] + data["NIR"]["Isabela"]
nir_neg_ori = data["NIR"]["Negros_Oriental"]
nir_siquijor = data["NIR"]["Siquijor"]
sulu_cities = data["Region_9"]["Sulu"]
barmm_mag = data["BARMM"]["Maguindanao"] + data["BARMM"]["Sultan_Kudarat"]

output_lines = []

# CAR (region_id=2)
car_map = {
    "Abra": data["CAR"]["Abra"],
    "Apayao": data["CAR"]["Apayao"],
    "Benguet": data["CAR"]["Benguet"],
    "Ifugao": data["CAR"]["Ifugao"],
    "Kalinga": data["CAR"]["Kalinga"],
    "Mountain Province": data["CAR"]["Mountain_Province"],
    "Baguio City": ["Baguio_City"],
}
output_lines.append(print_const("REGION_CAR_CITIES", car_map))
output_lines.append("")

# Region V (region_id=8)
v_map = {
    "Albay": data["Region_5"]["Albay"],
    "Camarines Norte": data["Region_5"]["Camarines_Norte"],
    "Camarines Sur": data["Region_5"]["Camarines_Sur"],
    "Catanduanes": data["Region_5"]["Catanduanes"],
    "Masbate": data["Region_5"]["Masbate"],
    "Sorsogon": data["Region_5"]["Sorsogon"],
}
output_lines.append(print_const("REGION_V_CITIES", v_map))
output_lines.append("")

# Region VI (region_id=9) — Negros Occidental from NIR
vi_map = {
    "Aklan": data["Region_6"]["Aklan"],
    "Antique": data["Region_6"]["Antique"],
    "Capiz": data["Region_6"]["Capiz"],
    "Guimaras": data["Region_6"]["Guimaras"],
    "Iloilo": data["Region_6"]["Iloilo"],
    "Negros Occidental": nir_neg_occ,
}
output_lines.append(print_const("REGION_VI_CITIES", vi_map))
output_lines.append("")

# Region VII (region_id=10) — Siquijor + Negros Oriental from NIR
vii_map = {
    "Bohol": data["Region_7"]["Bohol"],
    "Cebu": data["Region_7"]["Cebu"],
    "Negros Oriental": nir_neg_ori,
    "Siquijor": nir_siquijor,
}
output_lines.append(print_const("REGION_VII_CITIES", vii_map))
output_lines.append("")

# Region VIII (region_id=11) — province name remaps
viii_map = {
    "Biliran": data["Region_8"]["Biliran"],
    "Eastern Samar": data["Region_8"]["Eastern_Samar"],
    "Leyte": data["Region_8"]["Northern_Leyte"],
    "Northern Samar": data["Region_8"]["Northern_Samar"],
    "Samar": data["Region_8"]["Western_Samar"],
    "Southern Leyte": data["Region_8"]["Southern_Leyte"],
}
output_lines.append(print_const("REGION_VIII_CITIES", viii_map))
output_lines.append("")

# Region IX (region_id=12) — skip Sulu (goes to BARMM)
ix_map = {
    "Zamboanga del Norte": data["Region_9"]["Zamboanga_del_Norte"],
    "Zamboanga del Sur": data["Region_9"]["Zamboanga_del_Sur"],
    "Zamboanga Sibugay": data["Region_9"]["Zamboanga_Sibugay"],
}
output_lines.append(print_const("REGION_IX_CITIES", ix_map))
output_lines.append("")

# Region X (region_id=13)
x_map = {
    "Bukidnon": data["Region_10"]["Bukidnon"],
    "Camiguin": data["Region_10"]["Camiguin"],
    "Lanao del Norte": data["Region_10"]["Lanao_del_Norte"],
    "Misamis Occidental": data["Region_10"]["Misamis_Occidental"],
    "Misamis Oriental": data["Region_10"]["Misamis_Oriental"],
}
output_lines.append(print_const("REGION_X_CITIES", x_map))
output_lines.append("")

# Region XI (region_id=14)
xi_map = {
    "Davao de Oro": data["Region_11"]["Compostela_Valley_Province"],
    "Davao del Norte": data["Region_11"]["Davao_del_Norte"],
    "Davao del Sur": data["Region_11"]["Davao_del_Sur"],
    "Davao Occidental": data["Region_11"]["Davao_Occidental"],
    "Davao Oriental": data["Region_11"]["Davao_Oriental"],
}
output_lines.append(print_const("REGION_XI_CITIES", xi_map))
output_lines.append("")

# Region XII (region_id=15)
xii_map = {
    "North Cotabato": data["Region_12"]["North_Cotabato"],
    "Sarangani": data["Region_12"]["Sarangani"],
    "South Cotabato": data["Region_12"]["South_Cotabato"],
    "Sultan Kudarat": data["Region_12"]["Sultan_Kudarat"],
}
output_lines.append(print_const("REGION_XII_CITIES", xii_map))
output_lines.append("")

# Region XIII / CARAGA (region_id=16)
xiii_map = {
    "Agusan del Norte": data["CARAGA"]["Agusan_del_Norte"],
    "Agusan del Sur": data["CARAGA"]["Agusan_del_Sur"],
    "Dinagat Islands": data["CARAGA"]["Dinagat_Island"],
    "Surigao del Norte": data["CARAGA"]["Surigao_del_Norte"],
    "Surigao del Sur": data["CARAGA"]["Surigao_del_Sur"],
}
output_lines.append(print_const("REGION_XIII_CITIES", xiii_map))
output_lines.append("")

# BARMM (region_id=17) — Maguindanao covers both del Norte + del Sur; Sulu from Region_9
barmm_map = {
    "Basilan": data["BARMM"]["Basilan"],
    "Lanao del Sur": data["BARMM"]["Lanao_del_Sur"],
    "Maguindanao del Norte": barmm_mag,
    "Maguindanao del Sur": barmm_mag,
    "Sulu": sulu_cities,
    "Tawi-Tawi": data["BARMM"]["Tawi_tawi"],
}
output_lines.append(print_const("BARMM_CITIES", barmm_map))
output_lines.append("")

# NIR (region_id=18) — Isabela section merges into Negros Occidental
nir_map = {
    "Negros Occidental": nir_neg_occ,
    "Negros Oriental": nir_neg_ori,
}
output_lines.append(print_const("NIR_CITIES", nir_map))

with open(
    r"c:/proj/WIMS-BFP-PROTOTYPE/scripts/gen_cities_output.ts", "w", encoding="utf-8"
) as f:
    f.write("\n".join(output_lines) + "\n")
print("Written to gen_cities_output.ts")
