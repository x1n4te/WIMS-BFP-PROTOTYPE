import openpyxl

path = (
    r"e:\WIMS-GIT\WIMS-BFP-PROTOTYPE\src\frontend\public\templates\afor_template.xlsx"
)
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

print("--- DETAILED INSPECTION ---")
# Check rows 20-120
for r in range(20, 121):
    # Only print rows that seem relevant (have some content)
    has_content = False
    for c in ["A", "B", "C", "D", "E"]:
        if ws[f"{c}{r}"].value:
            has_content = True
            break

    if has_content:
        line = f"{r:3}: "
        for c in ["A", "B", "C", "D", "E"]:
            val = ws[f"{c}{r}"].value
            line += f"{c}: {repr(val)[:20]:20} | "
        print(line)

wb.close()
