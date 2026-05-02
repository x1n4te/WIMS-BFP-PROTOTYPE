import csv
import io
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from api.routes.regional import (
    BfpXlsxParser,
    detect_afor_template_kind,
    parse_afor_report_data,
    parse_csv_content,
    parse_wildland_afor_report_data,
    parse_xlsx_content,
)


class _FakeSheet:
    """Minimal worksheet: stable cell values (MagicMock subscripts are not reliable for .value)."""

    def __init__(self, values: dict[str, object]):
        self._values = values

    def __getitem__(self, coord: str):
        class _Cell:
            pass

        c = _Cell()
        c.value = self._values.get(coord)
        return c


def test_bfp_xlsx_parser_mapping():
    """The worksheet parser should read the official AFOR coordinates."""
    mock_ws = _FakeSheet(
        {
            "B20": None,
            "B21": "x",
            "D21": "Quezon City Fire Station",
            "D22": "2025-11-20",
            "D23": "14:30",
            "D26": "Quezon City",
            "D42": "Second Alarm",
            "B50": "/",
            "B60": "1",
            "B102": "x",
            "D102": "Covered court",
            "D62": 2,
            "D70": 3,
            "D89": "14:30",
        }
    )

    data = BfpXlsxParser(mock_ws).parse()

    assert data["responder_type"] == "Augmenting Team"
    assert data["fire_station_name"] == "Quezon City Fire Station"
    assert data["classification"] == "Transportation"
    assert data["extent_of_damage"] == "Total Loss"
    assert data["notification_date"] == "2025-11-20"
    assert data["city"] == "Quezon City"
    assert data["alarm_level"] == "Second Alarm"
    assert data["structures_affected"] == 2
    assert data["res_bfp_trucks"] == 3
    assert data["alarm_1st"] == "14:30"
    assert data["icp_present"] is True
    assert data["icp_location"] == "Covered court"


def test_parse_afor_report_data_maps_canonical_schema():
    """Parsed AFOR fields should land in the schema-aligned payload."""
    raw_data = {
        "notification_date": "2025-11-20",
        "notification_time": "14:30",
        "fire_station_name": "Station A",
        "responder_type": "First Responder",
        "classification": "Structural",
        "category": "Residential",
        "origin": "Kitchen",
        "stage": "Fully Developed",
        "extent": "Total Loss",
        "city": "Manila",
        "address": "123 Sample St",
        "landmark": "Barangay Hall",
        "engine": "Engine 1",
        "time_dispatched": "14:35",
        "time_arrived": "14:42",
        "time_returned": "16:00",
        "distance_km": "3.5",
        "response_time": "12",
        "gas_liters": "20",
        "struct_aff": "5",
        "res_bfp_truck": 2,
        "alarm_level": "1ST ALARM",
        "alarm_1st": "14:45",
        "inj_civ_m": 1,
        "fat_bfp_f": 1,
        "caller_info": "Juan Dela Cruz / 09171234567",
        "receiver": "Duty Watch",
        "owner": "ABC Hardware",
        "pod_commander": "Cmdr",
        "narrative": "Narrative body",
        "recommendations": "Improve hydrant access",
        "disposition": "Forward to investigation",
        "prepared_by": "Shift IC",
        "noted_by": "Engine Commander",
    }

    result = parse_afor_report_data(raw_data, region_id=13)

    assert result.status == "VALID"
    assert result.errors == []

    payload = result.data
    ns = payload["incident_nonsensitive_details"]
    sens = payload["incident_sensitive_details"]
    responding_unit = payload["responding_unit"]

    assert ns["fire_station_name"] == "Station A"
    assert ns["general_category"] == "Structural"
    assert ns["sub_category"] == "Residential"
    assert ns["fire_origin"] == "Kitchen"
    assert ns["structures_affected"] == 5
    assert ns["distance_from_station_km"] == 3.5
    assert ns["resources_deployed"]["trucks"]["bfp"] == 2
    assert "2025-11-20T14:45" in ns["alarm_timeline"]["alarm_1st"]

    assert sens["caller_name"] == "Juan Dela Cruz"
    assert sens["caller_number"] == "09171234567"
    assert sens["receiver_name"] == "Duty Watch"
    assert sens["owner_name"] == "ABC Hardware"
    assert sens["casualty_details"]["injured"]["civilian"]["m"] == 1
    assert sens["casualty_details"]["fatalities"]["firefighter"]["f"] == 1
    assert sens["narrative_report"] == "Narrative body"
    assert sens["disposition_prepared_by"] == "Shift IC"
    assert sens["disposition_noted_by"] == "Engine Commander"

    assert responding_unit["engine_number"] == "Engine 1"
    assert "2025-11-20T14:35" in responding_unit["dispatch_dt"]
    assert "2025-11-20T14:42" in responding_unit["arrival_dt"]
    assert "2025-11-20T16:00" in responding_unit["return_dt"]

    assert payload["_form_kind"] == "STRUCTURAL_AFOR"


def test_parse_afor_report_data_invalid_date():
    """Invalid notification dates should invalidate the import row."""
    result = parse_afor_report_data(
        {"notification_date": "Invalid Date", "city": "Manila"}, region_id=13
    )

    assert result.status == "INVALID"
    assert any("notification_dt" in err for err in result.errors)


def test_parse_afor_report_data_excel_serial_notification_datetime():
    """Excel serial values (e.g., D22/D23) should produce a valid notification_dt."""
    result = parse_afor_report_data(
        {
            "notification_date": 46096.0,
            "notification_time": 0.6041666666666666,
            "city": "Manila",
            "fire_station_name": "Sampaloc Central Fire Station",
        },
        region_id=13,
    )

    assert result.status == "VALID"
    assert result.data["incident_nonsensitive_details"]["notification_dt"] is not None


def test_parse_csv_content_supports_official_form_layout():
    """The official AFOR CSV export should be parsed as a worksheet, not a DictReader table."""
    rows = [["" for _ in range(6)] for _ in range(241)]

    def set_cell(row_number: int, column_index: int, value: str) -> None:
        rows[row_number - 1][column_index] = value

    set_cell(14, 0, "AFTER FIRE OPERATIONS REPORT")
    set_cell(18, 0, "A. RESPONSE DETAILS")
    set_cell(20, 1, "x")
    set_cell(20, 3, "Station A")
    set_cell(22, 3, "2025-11-20")
    set_cell(23, 3, "14:30")
    set_cell(26, 3, "Manila")
    set_cell(42, 3, "1ST ALARM")

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerows(rows)

    rows, form_kind = parse_csv_content(buffer.getvalue(), region_id=1)

    assert form_kind == "STRUCTURAL_AFOR"
    assert len(rows) == 1
    assert rows[0].status == "VALID"
    assert rows[0].data["_city_text"] == "Manila"
    assert (
        rows[0].data["incident_nonsensitive_details"]["fire_station_name"]
        == "Station A"
    )


@patch("openpyxl.load_workbook")
def test_parse_xlsx_content_flow(mock_load):
    """The XLSX flow should produce a schema-aligned preview row."""
    mock_ws = _FakeSheet(
        {
            "A14": "AFTER FIRE OPERATIONS REPORT",
            "A18": "A. RESPONSE DETAILS",
            "B20": "x",
            "D20": "Station",
            "D22": "2025-11-20",
            "D23": "14:30",
            "D26": "Manila",
            "D56": 500,
            "D60": 0.5,
        }
    )
    mock_wb = MagicMock()
    mock_wb.sheetnames = ["AFOR Sheet"]
    mock_wb.__getitem__.side_effect = lambda _name: mock_ws
    mock_wb.close = MagicMock()
    mock_load.return_value = mock_wb

    results, form_kind = parse_xlsx_content(b"fake content", region_id=1)

    assert form_kind == "STRUCTURAL_AFOR"
    assert len(results) == 1
    assert results[0].status == "VALID"
    assert results[0].data["_city_text"] == "Manila"
    assert (
        results[0].data["incident_nonsensitive_details"]["extent_total_floor_area_sqm"]
        == 500
    )


def test_detect_structural_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "AFOR"
    ws["A14"] = "AFTER FIRE OPERATIONS REPORT"
    ws["A18"] = "A. RESPONSE DETAILS"
    assert detect_afor_template_kind(wb) == "STRUCTURAL_AFOR"


def test_detect_structural_workbook_shifted_rows():
    """Filled variants with row-shifted title/section markers should still classify."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Filled AFOR"
    ws["A15"] = "AFTER FIRE OPERATIONS REPORT"
    ws["A19"] = "A. RESPONSE DETAILS"
    assert detect_afor_template_kind(wb) == "STRUCTURAL_AFOR"


def test_bfp_xlsx_parser_shifted_rows_reads_required_fields():
    """Row-shifted structural sheets should still map required cells correctly."""
    mock_ws = _FakeSheet(
        {
            "A15": "AFTER FIRE OPERATIONS REPORT",
            "A19": "A. RESPONSE DETAILS",
            "B21": "x",
            "D21": "Station A",
            "D23": "2025-11-20",
            "D24": "14:30",
            "D27": "Manila",
            "D43": "1ST ALARM",
            "D57": 500,
        }
    )

    data = BfpXlsxParser(mock_ws).parse()
    result = parse_afor_report_data(data, region_id=1)

    assert result.status == "VALID"
    assert result.data["_city_text"] == "Manila"
    assert (
        result.data["incident_nonsensitive_details"]["fire_station_name"] == "Station A"
    )


def test_detect_wildland_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "WILDLAND FIRE AFOR"
    ws["B12"] = "AFTER FIRE OPERATIONS REPORT OF WILDLAND FIRE "
    ws["B13"] = "A. DATES AND TIMES"
    assert detect_afor_template_kind(wb) == "WILDLAND_AFOR"


def test_detect_ambiguous_workbook_returns_none():
    wb = Workbook()
    wb.active["A1"] = "nothing"
    assert detect_afor_template_kind(wb) is None


def test_parse_wildland_afor_report_data_minimal_valid():
    """At least one of primary action, engine, narration, call time, or fire type validates."""
    data = {
        "primary_action_taken": "Direct attack",
        "engine_dispatched": "",
        "narration": "",
        "call_received_at": None,
        "wildland_fire_type": None,
    }
    result = parse_wildland_afor_report_data(data, region_id=13)
    assert result.status == "VALID"
    assert result.data["wildland"]["primary_action_taken"] == "Direct attack"


def test_parse_wildland_afor_report_data_empty_invalid():
    """Empty wildland content should invalidate (matches commit-time rejection)."""
    data = {
        "primary_action_taken": "",
        "engine_dispatched": "",
        "narration": "",
        "call_received_at": None,
        "wildland_fire_type": None,
    }
    result = parse_wildland_afor_report_data(data, region_id=13)
    assert result.status == "INVALID"
    assert any("Missing wildland content" in e for e in result.errors)
