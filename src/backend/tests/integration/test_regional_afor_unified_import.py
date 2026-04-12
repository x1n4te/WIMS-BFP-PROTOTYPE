"""
Unified regional AFOR import — structural vs wildland detection, preview, commit.

Requires DATABASE_URL / running Postgres with WIMS schema (same as test_triage_api).
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

import auth
from main import app


# ---------------------------------------------------------------------------
# Helpers — minimal XLSX bytes (openpyxl)
# ---------------------------------------------------------------------------


def _build_structural_afor_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "AFOR"
    ws["A14"] = "AFTER FIRE OPERATIONS REPORT"
    ws["A18"] = "A. RESPONSE DETAILS"
    ws["B20"] = "x"
    ws["D20"] = "Test FS"
    ws["D22"] = datetime(2025, 1, 15)
    ws["D23"] = "10:00"
    ws["D26"] = "Manila"
    ws["D42"] = "First Alarm"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_wildland_afor_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "WILDLAND FIRE AFOR"
    ws["B12"] = "AFTER FIRE OPERATIONS REPORT OF WILDLAND FIRE "
    ws["B13"] = "A. DATES AND TIMES"
    ws["D15"] = datetime(2025, 3, 10, 8, 30)
    ws["E27"] = "Direct attack on head"
    ws["D23"] = "Engine 99"
    ws["G44"] = "Brush Fire"
    ws["B44"] = "12 ha"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_ambiguous_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws["A1"] = "Not an AFOR workbook"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture(autouse=True)
def _reset_overrides():
    yield
    app.dependency_overrides.clear()


@pytest.fixture
def db_session():
    from database import _SessionLocal  # noqa: SLF001

    db = _SessionLocal()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture
def regional_user_id(db_session: Session):
    """REGIONAL_ENCODER with assigned NCR region (seed)."""
    keycloak_id = uuid.uuid4()
    username = f"regional_test_{keycloak_id.hex[:8]}"
    row = db_session.execute(
        text("""
            INSERT INTO wims.users (keycloak_id, username, role, assigned_region_id)
            SELECT :kid, :username, 'REGIONAL_ENCODER', region_id
            FROM wims.ref_regions WHERE region_code = 'NCR' LIMIT 1
            RETURNING user_id
        """),
        {"kid": keycloak_id, "username": username},
    ).fetchone()
    db_session.commit()
    assert row is not None
    return row[0]


@pytest.fixture
def require_wildland_schema(db_session: Session):
    """Skip DB-dependent tests when Postgres predates wildland AFOR DDL."""
    ok = db_session.execute(
        text(
            """
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_schema = 'wims' AND table_name = 'incident_wildland_afor'
            )
            """
        )
    ).scalar()
    if not ok:
        pytest.skip(
            "wims.incident_wildland_afor missing — apply src/postgres-init/01_wims_initial.sql to the DB"
        )


@pytest.fixture
def client_regional_encoder(client: TestClient, regional_user_id, db_session: Session):
    rid = db_session.execute(
        text(
            "SELECT assigned_region_id FROM wims.users WHERE user_id = CAST(:u AS uuid)"
        ),
        {"u": regional_user_id},
    ).scalar()

    async def mock_regional_encoder_fixed():
        return {
            "user_id": regional_user_id,
            "keycloak_id": str(uuid.uuid4()),
            "role": "REGIONAL_ENCODER",
            "assigned_region_id": rid,
        }

    app.dependency_overrides[auth.get_regional_encoder] = mock_regional_encoder_fixed
    try:
        yield client
    finally:
        app.dependency_overrides.pop(auth.get_regional_encoder, None)


# ---------------------------------------------------------------------------
# Helpers — WGS84 / PostGIS
# ---------------------------------------------------------------------------

# Distinct from legacy placeholder POINT(121.0 14.5) used before real coords.
SAMPLE_LAT = 14.5547
SAMPLE_LON = 121.0244


def _fetch_incident_wgs84(db_session: Session, incident_id: int) -> tuple[float, float]:
    """Return (longitude, latitude) from fire_incidents.location (SRID 4326)."""
    row = db_session.execute(
        text(
            """
            SELECT ST_X(location::geometry), ST_Y(location::geometry)
            FROM wims.fire_incidents WHERE incident_id = :id
            """
        ),
        {"id": incident_id},
    ).fetchone()
    assert row is not None
    return float(row[0]), float(row[1])


def _commit_coords_body() -> dict:
    return {"latitude": SAMPLE_LAT, "longitude": SAMPLE_LON}


def _assert_wgs84_error(res, status: int = 400):
    assert res.status_code == status
    detail = res.json().get("detail")
    if isinstance(detail, dict):
        assert detail.get("code") == "AFOR_WGS84_INVALID"
        assert "message" in detail
    else:
        assert isinstance(detail, str)
        assert "WGS84" in detail or "latitude" in detail.lower()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_regional_import_preview_structural_form_kind(
    client_regional_encoder: TestClient,
):
    response = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "struct.xlsx",
                _build_structural_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 200
    data = response.json()
    assert data.get("form_kind") == "STRUCTURAL_AFOR"
    assert data.get("requires_location") is True


def test_regional_import_preview_wildland_form_kind(
    client_regional_encoder: TestClient,
):
    response = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 200
    data = response.json()
    assert data.get("form_kind") == "WILDLAND_AFOR"
    assert data.get("requires_location") is True


def test_regional_import_ambiguous_returns_400(client_regional_encoder: TestClient):
    response = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "bad.xlsx",
                _build_ambiguous_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 400
    detail = response.json().get("detail", "")
    assert "could not determine AFOR type" in detail
    assert "template" in detail.lower()


def test_commit_wildland_persists_incident_wildland_afor(
    require_wildland_schema,
    client_regional_encoder: TestClient,
    db_session: Session,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    preview = prev.json()
    assert preview["form_kind"] == "WILDLAND_AFOR"
    rows = [r["data"] for r in preview["rows"] if r["status"] == "VALID"]
    assert rows

    commit = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={
            "form_kind": "WILDLAND_AFOR",
            "rows": rows,
            **_commit_coords_body(),
        },
    )
    assert commit.status_code == 200, commit.text
    incident_ids = commit.json()["incident_ids"]
    assert incident_ids

    lon, lat = _fetch_incident_wgs84(db_session, incident_ids[0])
    assert abs(lon - SAMPLE_LON) < 1e-5
    assert abs(lat - SAMPLE_LAT) < 1e-5
    assert not (abs(lon - 121.0) < 1e-9 and abs(lat - 14.5) < 1e-9)

    src = db_session.execute(
        text(
            """
            SELECT iwa.source, iwa.wildland_fire_type
            FROM wims.incident_wildland_afor iwa
            WHERE iwa.incident_id = :iid
            """
        ),
        {"iid": incident_ids[0]},
    ).fetchone()
    assert src is not None
    assert src[0] == "AFOR_IMPORT"
    assert src[1] is not None


def test_commit_wildland_manual_source_sets_manual(
    require_wildland_schema,
    client_regional_encoder: TestClient,
    db_session: Session,
):
    """Manual wildland entry via commit sets incident_wildland_afor.source = MANUAL."""
    row = {
        "_form_kind": "WILDLAND_AFOR",
        "_city_text": "",
        "region_id": 1,
        "wildland": {
            "primary_action_taken": "Direct attack on head",
            "engine_dispatched": "Engine 42",
            "wildland_fire_type": "brush fire",
        },
    }
    commit = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={
            "form_kind": "WILDLAND_AFOR",
            "wildland_row_source": "MANUAL",
            "rows": [row],
            **_commit_coords_body(),
        },
    )
    assert commit.status_code == 200, commit.text
    incident_ids = commit.json()["incident_ids"]
    assert incident_ids

    lon, lat = _fetch_incident_wgs84(db_session, incident_ids[0])
    assert abs(lon - SAMPLE_LON) < 1e-5
    assert abs(lat - SAMPLE_LAT) < 1e-5

    src = db_session.execute(
        text(
            """
            SELECT iwa.source
            FROM wims.incident_wildland_afor iwa
            WHERE iwa.incident_id = :iid
            """
        ),
        {"iid": incident_ids[0]},
    ).fetchone()
    assert src is not None
    assert src[0] == "MANUAL"


def test_commit_wildland_invalid_payload_returns_400(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    """Commit rejects wildland rows that fail parse_wildland_afor_report_data (no minimum content)."""
    row = {
        "_form_kind": "WILDLAND_AFOR",
        "_city_text": "",
        "region_id": 1,
        "wildland": {
            "primary_action_taken": "",
            "engine_dispatched": "",
            "narration": "",
            "call_received_at": None,
            "wildland_fire_type": None,
        },
    }
    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "WILDLAND_AFOR", "rows": [row], **_commit_coords_body()},
    )
    assert res.status_code == 400
    detail = res.json().get("detail", "")
    assert isinstance(detail, str)
    assert "Missing wildland content" in detail or "wildland" in detail.lower()


def test_commit_missing_coordinates_returns_400(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]
    assert rows
    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "WILDLAND_AFOR", "rows": rows},
    )
    _assert_wgs84_error(res)


def test_commit_invalid_latitude_returns_400(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]
    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={
            "form_kind": "WILDLAND_AFOR",
            "rows": rows,
            "latitude": 91.0,
            "longitude": 121.0,
        },
    )
    _assert_wgs84_error(res)


def test_commit_invalid_longitude_returns_400(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]
    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={
            "form_kind": "WILDLAND_AFOR",
            "rows": rows,
            "latitude": 14.5,
            "longitude": 200.0,
        },
    )
    _assert_wgs84_error(res)


def test_commit_structural_persists_wgs84_coordinates(
    client_regional_encoder: TestClient,
    db_session: Session,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "struct.xlsx",
                _build_structural_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]
    assert rows
    commit = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "STRUCTURAL_AFOR", "rows": rows, **_commit_coords_body()},
    )
    assert commit.status_code == 200, commit.text
    iid = commit.json()["incident_ids"][0]
    lon, lat = _fetch_incident_wgs84(db_session, iid)
    assert abs(lon - SAMPLE_LON) < 1e-5
    assert abs(lat - SAMPLE_LAT) < 1e-5


def test_commit_rejects_form_kind_mismatch(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]

    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "STRUCTURAL_AFOR", "rows": rows, **_commit_coords_body()},
    )
    assert res.status_code == 400
