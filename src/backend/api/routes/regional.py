"""Regional Office API — AFOR Import, Regional Incidents, Stats."""

from __future__ import annotations

import csv
import io
import hashlib
import json
import logging
import math
import re
from datetime import datetime, timedelta
from typing import Annotated, Any, Literal, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from auth import get_current_wims_user, get_national_validator, get_regional_encoder
from database import get_db_with_rls
from services.analytics_read_model import (
    sync_incident_to_analytics,
    sync_incidents_batch,
)
from services.duplicate_detection import check_for_duplicate
from utils.crypto import SecurityProvider, SecurityProviderError
from utils.audit import log_system_audit


# ── Lazy SecurityProvider singleton (avoids import-time env check in test mocks) ──
_sp_instance: SecurityProvider | None = None


def _get_security_provider() -> SecurityProvider:
    global _sp_instance  # noqa: PLW0603
    if _sp_instance is None:
        _sp_instance = SecurityProvider()
    return _sp_instance


logger = logging.getLogger("wims.regional")


router = APIRouter(prefix="/api/regional", tags=["regional"])


# ---------------------------------------------------------------------------
# Pydantic Schemas
# ---------------------------------------------------------------------------


class AforParsedRow(BaseModel):
    row_index: int
    status: str  # VALID | INVALID
    errors: list[str]
    data: dict[str, Any]


AforFormKind = Literal["STRUCTURAL_AFOR", "WILDLAND_AFOR"]
WildlandRowSource = Literal["AFOR_IMPORT", "MANUAL"]


class AforParseResponse(BaseModel):
    total_rows: int
    valid_rows: int
    invalid_rows: int
    rows: list[AforParsedRow]
    form_kind: AforFormKind
    # True when the file does not supply reliable WGS84 coordinates; client must collect lat/lon before commit.
    requires_location: bool = True


DuplicateAction = Literal["skip", "merge", "force"]


class RowResolution(BaseModel):
    """Encoder decision for one duplicate row, returned on second commit call."""

    row_index: int
    action: DuplicateAction
    existing_incident_id: int | None = None  # required when action == "merge"


class AforCommitRequest(BaseModel):
    form_kind: AforFormKind
    rows: list[dict[str, Any]]
    # WILDLAND_AFOR: MANUAL for manual entry; omit or AFOR_IMPORT for file import.
    wildland_row_source: WildlandRowSource | None = None
    # WGS84 (SRID 4326). PostGIS stores POINT(longitude latitude) — not GeoJSON [lat, lon].
    latitude: float | None = None
    longitude: float | None = None
    # M4-D: per-row duplicate resolutions on second commit call.
    # When None, the backend runs the duplicate scan and returns DUPLICATE_CHECK_REQUIRED.
    # When provided, the backend applies each row's chosen action.
    resolutions: list[RowResolution] | None = None


class AforCommitResponse(BaseModel):
    status: str
    batch_id: int
    incident_ids: list[int]
    total_committed: int


# M4-D duplicate detection thresholds
DUPLICATE_RADIUS_METERS = 1000  # 1 km
DUPLICATE_MIN_MATCHING_FIELDS = 3


class RegionalStatsResponse(BaseModel):
    total_incidents: int
    by_category: list[dict[str, Any]]
    by_alarm_level: list[dict[str, Any]]
    by_status: list[dict[str, Any]]
    wildland_total: int = 0
    by_wildland_type: list[dict[str, Any]] = []


AFOR_WGS84_INVALID_CODE = "AFOR_WGS84_INVALID"
AFOR_WGS84_INVALID_MESSAGE = (
    "AFOR commit requires valid WGS84 latitude and longitude as JSON numbers "
    "(latitude -90..90, longitude -180..180, both finite). "
    "PostGIS stores POINT(longitude latitude) in SRID 4326; do not confuse with GeoJSON [lat, lon]."
)


def _wgs84_pair_from_raw(latitude: Any, longitude: Any) -> tuple[float, float]:
    """Return (longitude, latitude) for ST_MakePoint. Validates JSON types from the raw request body."""
    if latitude is None or longitude is None:
        raise HTTPException(
            status_code=400,
            detail={
                "code": AFOR_WGS84_INVALID_CODE,
                "message": AFOR_WGS84_INVALID_MESSAGE,
            },
        )
    if type(latitude) is bool or type(longitude) is bool:
        raise HTTPException(
            status_code=400,
            detail={
                "code": AFOR_WGS84_INVALID_CODE,
                "message": AFOR_WGS84_INVALID_MESSAGE,
            },
        )
    if type(latitude) not in (int, float) or type(longitude) not in (int, float):
        raise HTTPException(
            status_code=400,
            detail={
                "code": AFOR_WGS84_INVALID_CODE,
                "message": AFOR_WGS84_INVALID_MESSAGE,
            },
        )
    lat = float(latitude)
    lon = float(longitude)
    if not math.isfinite(lat) or not math.isfinite(lon):
        raise HTTPException(
            status_code=400,
            detail={
                "code": AFOR_WGS84_INVALID_CODE,
                "message": AFOR_WGS84_INVALID_MESSAGE,
            },
        )
    if lat < -90.0 or lat > 90.0 or lon < -180.0 or lon > 180.0:
        raise HTTPException(
            status_code=400,
            detail={
                "code": AFOR_WGS84_INVALID_CODE,
                "message": AFOR_WGS84_INVALID_MESSAGE,
            },
        )
    return lon, lat


def _incident_verification_history_uses_target_columns(db: Session) -> bool:
    """Return True when IVH table already has target_type/target_id columns."""
    return _incident_verification_history_has_column(db, "target_type")


def _incident_verification_history_has_column(db: Session, column_name: str) -> bool:
    """Return True when IVH table has the given column."""
    return bool(
        db.execute(
            text("""
                SELECT EXISTS (
                    SELECT 1
                    FROM information_schema.columns
                    WHERE table_schema = 'wims'
                      AND table_name = 'incident_verification_history'
                      AND column_name = :column_name
                )
            """),
            {"column_name": column_name},
        ).scalar()
    )


def _insert_incident_verification_history(
    db: Session,
    *,
    incident_id: int,
    actor_user_id: str,
    previous_status: str,
    new_status: str,
    notes: str,
    action_label: str | None = None,
) -> None:
    """Insert IVH row with compatibility for both legacy and migrated schemas."""
    has_action_label = _incident_verification_history_has_column(db, "action_label")

    if _incident_verification_history_uses_target_columns(db):
        if has_action_label:
            db.execute(
                text("""
                    INSERT INTO wims.incident_verification_history (
                        target_type, target_id, action_by_user_id,
                        previous_status, new_status, notes, action_label
                    ) VALUES (
                        'OFFICIAL', :iid, CAST(:uid AS uuid),
                        :prev_status, :new_status, :notes, :action_label
                    )
                """),
                {
                    "iid": incident_id,
                    "uid": actor_user_id,
                    "prev_status": previous_status,
                    "new_status": new_status,
                    "notes": notes,
                    "action_label": action_label,
                },
            )
        else:
            db.execute(
                text("""
                    INSERT INTO wims.incident_verification_history (
                        target_type, target_id, action_by_user_id,
                        previous_status, new_status, notes
                    ) VALUES (
                        'OFFICIAL', :iid, CAST(:uid AS uuid),
                        :prev_status, :new_status, :notes
                    )
                """),
                {
                    "iid": incident_id,
                    "uid": actor_user_id,
                    "prev_status": previous_status,
                    "new_status": new_status,
                    "notes": notes,
                },
            )
        return

    db.execute(
        text("""
            INSERT INTO wims.incident_verification_history (
                incident_id, action_by_user_id,
                previous_status, new_status, comments
            ) VALUES (
                :iid, CAST(:uid AS uuid),
                :prev_status, :new_status, :comments
            )
        """),
        {
            "iid": incident_id,
            "uid": actor_user_id,
            "prev_status": previous_status,
            "new_status": new_status,
            "comments": notes,
        },
    )


# ---------------------------------------------------------------------------
# AFOR Parsing Utilities (Official BFP XLSX Refactor)
# ---------------------------------------------------------------------------

ALARM_LEVEL_MAP = {
    "1ST": "First Alarm",
    "1ST ALARM": "First Alarm",
    "FIRST": "First Alarm",
    "FIRST ALARM": "First Alarm",
    "2ND": "Second Alarm",
    "2ND ALARM": "Second Alarm",
    "SECOND": "Second Alarm",
    "SECOND ALARM": "Second Alarm",
    "3RD": "Third Alarm",
    "3RD ALARM": "Third Alarm",
    "THIRD": "Third Alarm",
    "THIRD ALARM": "Third Alarm",
    "4TH": "Fourth Alarm",
    "4TH ALARM": "Fourth Alarm",
    "FOURTH": "Fourth Alarm",
    "FOURTH ALARM": "Fourth Alarm",
    "5TH": "Fifth Alarm",
    "5TH ALARM": "Fifth Alarm",
    "FIFTH": "Fifth Alarm",
    "FIFTH ALARM": "Fifth Alarm",
    "TF ALPHA": "Task Force Alpha",
    "TASK FORCE ALPHA": "Task Force Alpha",
    "TF BRAVO": "Task Force Bravo",
    "TASK FORCE BRAVO": "Task Force Bravo",
    "TF CHARLIE": "Task Force Charlie",
    "TASK FORCE CHARLIE": "Task Force Charlie",
    "TF DELTA": "Task Force Delta",
    "TASK FORCE DELTA": "Task Force Delta",
    "GENERAL": "General Alarm",
    "GENERAL ALARM": "General Alarm",
}

_CATEGORY_CANONICAL: dict[str, str] = {
    "STRUCTURAL": "STRUCTURAL",
    "NON_STRUCTURAL": "NON_STRUCTURAL",
    "NON-STRUCTURAL": "NON_STRUCTURAL",
    "VEHICULAR": "VEHICULAR",
    "TRANSPORTATION": "VEHICULAR",
}

# All known DB values for each canonical category (covers legacy form submissions)
_CATEGORY_DB_VARIANTS: dict[str, list[str]] = {
    "STRUCTURAL": ["STRUCTURAL", "Structural"],
    "NON_STRUCTURAL": ["NON_STRUCTURAL", "Non-Structural", "NON-STRUCTURAL"],
    "VEHICULAR": ["VEHICULAR", "Transportation", "TRANSPORTATION", "Vehicular"],
}


def _normalize_general_category(val: str) -> str:
    key = val.strip().upper().replace("-", "_").replace(" ", "_")
    return _CATEGORY_CANONICAL.get(key, val)


def _safe_int(val: Any, default: int = 0) -> int:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        if isinstance(val, (int, float)):
            return int(val)
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default


def _safe_dt(val: Any) -> str | None:
    """Safe datetime string conversion."""
    if isinstance(val, datetime):
        return val.isoformat()
    if not val:
        return None

    # Excel stores dates/times as serial floats in many filled templates.
    # Serial date epoch (Windows): 1899-12-30.
    if isinstance(val, (int, float)):
        try:
            serial = float(val)
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=serial)
            if serial < 1:
                return dt.strftime("%H:%M:%S")
            return dt.isoformat()
        except Exception:
            return None

    raw_numeric = str(val).strip()
    try:
        serial = float(raw_numeric)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        if serial < 1:
            return dt.strftime("%H:%M:%S")
        return dt.isoformat()
    except (ValueError, TypeError):
        pass

    dt_str = str(val).strip()
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m-%d-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M",
        "%H:%M",
        "%H:%M:%S",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
    ]:
        try:
            return datetime.strptime(dt_str, fmt).isoformat()
        except ValueError:
            continue
    return None


_COORD_RE = re.compile(r"^([A-Z]+)(\d+)$")


class _SheetCell:
    def __init__(self, value: Any):
        self.value = value


def _column_letters_to_index(letters: str) -> int:
    index = 0
    for char in letters:
        index = (index * 26) + (ord(char) - ord("A") + 1)
    return index - 1


class CsvWorksheetAdapter:
    """Expose CSV cells through worksheet-like `A1` coordinates."""

    def __init__(self, rows: list[list[str]]):
        self.rows = rows

    def __getitem__(self, coord: str) -> _SheetCell:
        match = _COORD_RE.match(coord.upper())
        if not match:
            raise KeyError(f"Invalid coordinate: {coord}")

        column_letters, row_number = match.groups()
        row_idx = int(row_number) - 1
        col_idx = _column_letters_to_index(column_letters)

        value = None
        if 0 <= row_idx < len(self.rows) and 0 <= col_idx < len(self.rows[row_idx]):
            raw_value = self.rows[row_idx][col_idx]
            if isinstance(raw_value, str):
                raw_value = raw_value.strip()
            value = raw_value or None

        return _SheetCell(value)


def _looks_like_official_afor_csv(rows: list[list[str]]) -> bool:
    if not rows:
        return False

    first_column_values = [
        (row[0].strip().upper() if row and isinstance(row[0], str) else "") for row in rows
    ]
    return (
        "AFTER FIRE OPERATIONS REPORT" in first_column_values
        and "A. RESPONSE DETAILS" in first_column_values
    )


def _cell_str(ws: Any, coord: str) -> str:
    try:
        v = ws[coord].value
    except Exception:
        return ""
    if v is None:
        return ""
    return str(v).strip()


def _sheet_has_structural_markers(ws: Any) -> bool:
    """Structural AFOR marker detection, tolerant to row shifts in filled templates."""
    title_row, section_row = _find_structural_marker_rows(ws)
    if title_row is None or section_row is None:
        return False
    # In official templates, section header appears a few rows after title.
    return 2 <= (section_row - title_row) <= 8


def _find_structural_marker_rows(ws: Any) -> tuple[int | None, int | None]:
    """Find title/section marker rows by scanning the top-left block of the sheet."""
    title_row: int | None = None
    section_row: int | None = None

    for row in range(1, 161):
        row_values = [
            _cell_str(ws, f"{col}{row}").upper() for col in ("A", "B", "C", "D", "E", "F")
        ]
        combined = " ".join(v for v in row_values if v).strip()

        if title_row is None and "AFTER FIRE OPERATIONS REPORT" in combined:
            title_row = row
        if section_row is None and "A. RESPONSE DETAILS" in combined:
            section_row = row

        if title_row is not None and section_row is not None:
            break

    return title_row, section_row


def _sheet_has_wildland_markers(ws: Any) -> bool:
    """
    Wildland workbook: main sheet title in B12 and section A header in B13.
    Tie-break: sheet name containing 'WILDLAND FIRE AFOR' wins over structural when both match.
    """
    b12 = _cell_str(ws, "B12").upper()
    b13 = _cell_str(ws, "B13").upper()
    if "WILDLAND" in b12 and "A. DATES" in b13:
        return True
    if "WILDLAND FIRE" in b12:
        return True
    return False


def detect_afor_template_kind(wb: Any) -> AforFormKind | None:
    """
    Classify uploaded workbook as structural vs wildland AFOR.

    Detection:
        Wildland markers also match when B12 contains WILDLAND FIRE even if B13 is not the usual
        "A. DATES…" line (see `_sheet_has_wildland_markers`).

    Rules (order):
    1. If any sheet name contains 'WILDLAND FIRE AFOR' (case-insensitive) and that sheet
       has wildland markers (B12/B13 or title containing WILDLAND) → WILDLAND_AFOR.
    2. Else if any sheet has structural markers (A14 + A18) → STRUCTURAL_AFOR.
    3. Else if any sheet has wildland markers without relying on sheet name → WILDLAND_AFOR.
    4. Else None (ambiguous).
    """
    sheets: list[tuple[str, Any]] = [(n, wb[n]) for n in wb.sheetnames]

    for name, ws in sheets:
        if "WILDLAND FIRE AFOR" in name.upper() and _sheet_has_wildland_markers(ws):
            return "WILDLAND_AFOR"

    for name, ws in sheets:
        if _sheet_has_structural_markers(ws):
            return "STRUCTURAL_AFOR"

    for _name, ws in sheets:
        if _sheet_has_wildland_markers(ws):
            return "WILDLAND_AFOR"

    return None


def _pick_structural_worksheet(wb: Any) -> Any:
    for name in wb.sheetnames:
        ws = wb[name]
        if _sheet_has_structural_markers(ws):
            return ws
    for name in wb.sheetnames:
        if "AFOR" in name.upper():
            return wb[name]
    return wb.active


def _pick_wildland_worksheet(wb: Any) -> Any:
    for name in wb.sheetnames:
        if "WILDLAND" in name.upper() and "AFOR" in name.upper():
            return wb[name]
    for name in wb.sheetnames:
        if _sheet_has_wildland_markers(wb[name]):
            return wb[name]
    return wb.active


_WILDLAND_FIRE_TYPES_LOWER = {
    "fire",
    "agricultural land fire",
    "brush fire",
    "forest fire",
    "grassland fire",
    "grazing land fire",
    "mineral land fire",
    "peatland fire",
}


def _normalize_wildland_fire_type(raw: Any) -> str | None:
    if raw is None:
        return None
    t = str(raw).strip().lower()
    if t in _WILDLAND_FIRE_TYPES_LOWER:
        return t
    return None


def _parse_ha_from_area_text(raw: Any) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip()
    m = re.search(r"([\d.]+)\s*ha", s, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


class WildlandXlsxParser:
    """Parser for BFP wildland AFOR workbook (sheet 'WILDLAND FIRE AFOR')."""

    def __init__(self, ws: Any):
        self.ws = ws

    def get(self, coord: str) -> Any:
        val = self.ws[coord].value
        if val is None:
            return None
        if isinstance(val, str):
            return val.strip()
        return val

    def parse(self) -> dict[str, Any]:
        def _dt_cell(coord: str) -> datetime | None:
            v = self.get(coord)
            if isinstance(v, datetime):
                return v
            return None

        call_received = _dt_cell("D15")
        fire_started = _dt_cell("D17")
        fire_arrival = _dt_cell("D19")
        fire_controlled = _dt_cell("D21")

        extras: list[str] = []
        for coord in ("E28", "E29"):
            v = self.get(coord)
            if v:
                extras.append(str(v))

        fire_behavior = {
            "elevation_ft": _safe_float(self.get("D51"), 0.0) or None,
            "relative_position_slope": self.get("D52"),
            "aspect": self.get("D53"),
            "flame_length_ft": _safe_float(self.get("D54"), 0.0) or None,
            "rate_of_spread_chains_per_hour": _safe_float(self.get("D55"), 0.0) or None,
        }

        problems: list[str] = []
        for r in range(76, 80):
            line = self.get(f"B{r}")
            if line and str(line).strip():
                problems.append(str(line).strip())

        recommendations: list[str] = []
        for r in range(83, 87):
            line = self.get(f"B{r}")
            if line and str(line).strip():
                recommendations.append(str(line).strip())

        alarm_rows: list[dict[str, Any]] = []
        for r in range(50, 65):
            status = self.get(f"J{r}")
            if not status or not str(status).strip():
                continue
            time_declared = self.get(f"K{r}")
            commander = self.get(f"L{r}")
            alarm_rows.append(
                {
                    "alarm_status": str(status).strip(),
                    "time_declared": str(time_declared).strip() if time_declared else "",
                    "ground_commander": str(commander).strip() if commander else "",
                }
            )

        raw_type = self.get("G44")
        wft = _normalize_wildland_fire_type(raw_type)

        return {
            "call_received_at": call_received,
            "fire_started_at": fire_started,
            "fire_arrival_at": fire_arrival,
            "fire_controlled_at": fire_controlled,
            "caller_transmitted_by": self.get("B33") or self.get("B32"),
            "caller_office_address": self.get("D33") or self.get("D32"),
            "call_received_by_personnel": self.get("F33") or self.get("F32"),
            "engine_dispatched": self.get("D23"),
            "incident_location_description": self.get("D31") or self.get("B31"),
            "distance_to_fire_station_km": _safe_float(self.get("D32"), 0.0)
            if self.get("D32") not in (None, "")
            else None,
            "primary_action_taken": self.get("E27"),
            "assistance_combined_summary": " | ".join(extras) if extras else None,
            "buildings_involved": _safe_int(self.get("B40")),
            "buildings_threatened": _safe_int(self.get("G40")),
            "ownership_and_property_notes": self.get("B41") or self.get("B39"),
            "total_area_burned_display": self.get("B44"),
            "total_area_burned_hectares": _parse_ha_from_area_text(self.get("B44")),
            "wildland_fire_type": wft,
            "raw_wildland_fire_type": raw_type,
            "area_type_summary": {},
            "causes_and_ignition_factors": {},
            "suppression_factors": {},
            "weather": {},
            "fire_behavior": {k: v for k, v in fire_behavior.items() if v not in (None, "", 0.0)},
            "peso_losses": {},
            "casualties": {},
            "narration": self.get("B68"),
            "problems_encountered": problems,
            "recommendations_list": recommendations,
            "prepared_by": self.get("B91"),
            "prepared_by_title": self.get("B92"),
            "noted_by": self.get("E88") or self.get("F88"),
            "noted_by_title": self.get("E91"),
            "wildland_alarm_statuses": alarm_rows,
            "wildland_assistance_rows": [],
        }


def parse_wildland_afor_report_data(data: dict[str, Any], region_id: int) -> AforParsedRow:
    """Map wildland workbook dict into commit payload + validation."""
    errors: list[str] = []

    primary = (data.get("primary_action_taken") or "").strip()
    engine = (data.get("engine_dispatched") or "").strip()
    narration = (data.get("narration") or "").strip()
    call_at = data.get("call_received_at")
    wft = data.get("wildland_fire_type")

    if not primary and not engine and not narration and not call_at and not wft:
        errors.append(
            "Missing wildland content: need at least one of call time (D15), primary action (E27), "
            "engine (D23), narration (B68), or wildland fire type (G44)."
        )

    if data.get("raw_wildland_fire_type") and not wft:
        errors.append(
            f"Wildland fire type value is not allowed: {data.get('raw_wildland_fire_type')!r}. "
            "Use the Sheet1 list (e.g. Brush Fire, Forest Fire)."
        )

    wl_payload = {
        k: v for k, v in data.items() if k not in ("raw_wildland_fire_type", "recommendations_list")
    }
    wl_payload["recommendations"] = data.get("recommendations_list") or []

    mapped: dict[str, Any] = {
        "_form_kind": "WILDLAND_AFOR",
        "_city_text": "",
        "region_id": region_id,
        "wildland": wl_payload,
    }

    status = "VALID" if not errors else "INVALID"
    return AforParsedRow(row_index=0, status=status, errors=errors, data=mapped)


def _combine_date_and_time(notification_dt: str | None, time_value: Any) -> str | None:
    if not notification_dt or not time_value:
        return None

    date_part = str(notification_dt).split("T", 1)[0]
    return _safe_dt(f"{date_part} {str(time_value).strip()}")


class BfpXlsxParser:
    """Parser for the official BFP manual entry form (AFOR)."""

    def __init__(self, ws):
        self.ws = ws
        self._row_offset = self._infer_row_offset()

    def _infer_row_offset(self) -> int:
        """Infer row offset when users fill a structurally identical AFOR with shifted rows."""
        title_row, section_row = _find_structural_marker_rows(self.ws)
        if title_row is None:
            return 0

        offset = title_row - 14
        # Validate offset with section marker when available.
        if section_row is not None and (section_row - 18) != offset:
            return 0
        return offset

    def _coord_with_offset(self, coord: str) -> str:
        match = _COORD_RE.match(coord.upper())
        if not match or self._row_offset == 0:
            return coord

        col, row_str = match.groups()
        shifted_row = max(1, int(row_str) + self._row_offset)
        return f"{col}{shifted_row}"

    def get(self, coord: str) -> Any:
        shifted_coord = self._coord_with_offset(coord)
        val = self.ws[shifted_coord].value
        if val is None and shifted_coord != coord:
            # Fallback to canonical location to support mixed/custom sheets.
            val = self.ws[coord].value
        if val is None:
            return None
        if isinstance(val, str):
            return val.strip()
        return val

    def _is_marked(self, coord: str) -> bool:
        raw = self.get(coord)
        if raw is None:
            return False

        if isinstance(raw, bool):
            return raw

        if isinstance(raw, (int, float)):
            return raw != 0

        val = str(raw).strip().lower()
        if not val:
            return False

        if val.startswith("="):
            expr = val.lstrip("=").strip().lower()
            if expr in {"true", "1"}:
                return True

        return val in {
            "x",
            "1",
            "true",
            "v",
            "/",
            "yes",
            "checked",
            "☑",
            "☒",
            "✓",
            "✔",
            "✅",
        }

    def _first_nonempty(self, *coords: str) -> Any:
        for coord in coords:
            val = self.get(coord)
            if val is None:
                continue
            if isinstance(val, str) and not val.strip():
                continue
            return val
        return None

    def _male_female_pair(self, row: int) -> tuple[Any, Any]:
        # Some AFOR variants shift M/F columns by one; try common adjacent pairs.
        candidate_pairs = [("D", "E"), ("C", "D"), ("E", "F"), ("F", "G")]
        fallback_pair = (None, None)
        for male_col, female_col in candidate_pairs:
            male_val = self.get(f"{male_col}{row}")
            female_val = self.get(f"{female_col}{row}")
            if fallback_pair == (None, None):
                fallback_pair = (male_val, female_val)

            has_male = male_val not in (None, "")
            has_female = female_val not in (None, "")
            if has_male or has_female:
                return male_val, female_val

        return fallback_pair

    def _is_marked_on_row(self, row: int, cols: tuple[str, ...] = ("B", "C", "D")) -> bool:
        return any(self._is_marked(f"{col}{row}") for col in cols)

    def parse(self) -> dict[str, Any]:
        """Extract sections A through L into a comprehensive data dictionary."""

        # Section A: Response Details
        responder_type = (
            "First Responder"
            if self._is_marked("B20")
            else ("Augmenting Team" if self._is_marked("B21") else "First Responder")
        )

        # Section B: Classification
        classification = "Structural"
        cat_val = self.get("D48")
        if self._is_marked_on_row(49):
            classification = "Non-Structural"
            cat_val = self.get("D49")
        elif self._is_marked_on_row(50):
            classification = "Transportation"
            cat_val = self.get("D50")
        elif self.get("D49") not in (None, ""):
            classification = "Non-Structural"
            cat_val = self.get("D49")
        elif self.get("D50") not in (None, ""):
            classification = "Transportation"
            cat_val = self.get("D50")

        stage = self.get("D54") or self.get("B54")
        if stage and "pick from dropdown" in str(stage).lower():
            stage = None

        # Extent of Damage
        extent = "None / Minor"
        if self._is_marked_on_row(57):
            extent = "Confined to Object"
        elif self._is_marked_on_row(58):
            extent = "Confined to Room"
        elif self._is_marked_on_row(59):
            extent = "Confined to Structure"
        elif self._is_marked_on_row(60):
            extent = "Total Loss"
        elif self._is_marked_on_row(61):
            extent = "Extended Beyond Structure"
        else:
            extent_text = str(self._first_nonempty("D57", "D58", "D59", "D60", "D61") or "").strip()
            if extent_text:
                extent = extent_text

        # Section J: Problems
        problems = []
        prob_map = {
            "B195": "Inaccurate address",
            "B196": "Geographically challenged",
            "B197": "Road conditions",
            "B198": "Road under construction",
            "B199": "Traffic congestion",
            "B200": "Road accidents",
            "B201": "Vehicles failure to yield",
            "B202": "Natural Disasters",
            "B203": "Civil Disturbance",
            "B204": "Uncooperative or panicked residents",
            "B205": "Safety and security threats",
            "B206": "Property security or owner delays",
            "B207": "Engine failure",
            "B208": "Uncooperative fire auxiliary",
            "B209": "Poor water supply access",
            "B210": "Intense heat and smoke",
            "B211": "Structural hazards",
            "B212": "Equipment malfunction",
            "B213": "Poor inter-agency coordination",
            "B214": "Radio communication breakdown",
            "B215": "HazMat risks",
            "B216": "Physical exhaustion and injuries",
            "B217": "Emotional and psychological effects",
            "B218": "Community complaints",
            "B219": "Others",
        }
        for c, flavor in prob_map.items():
            row_num = int(c[1:])
            if self._is_marked_on_row(row_num):
                problems.append(flavor)

        icp_present = self._is_marked_on_row(102)
        icp_location = self.get("D102") if icp_present else None

        # Section I: Narrative joining (Rows 160 to 190)
        narrative_lines = []
        for r in range(160, 191):
            line = self.get(f"B{r}")
            if line:
                narrative_lines.append(str(line))

        # Section G: Other Personnel (Rows 124 to 132)
        others = []
        for r in range(124, 133):
            name = self.get(f"B{r}")
            rem = self.get(f"E{r}")
            if name and "N/A" not in str(name).upper():
                others.append({"name": name, "designation": rem or ""})

        inj_civ_m, inj_civ_f = self._male_female_pair(106)
        inj_bfp_m, inj_bfp_f = self._male_female_pair(107)
        inj_aux_m, inj_aux_f = self._male_female_pair(108)
        fat_civ_m, fat_civ_f = self._male_female_pair(109)
        fat_bfp_m, fat_bfp_f = self._male_female_pair(110)
        fat_aux_m, fat_aux_f = self._male_female_pair(111)

        return {
            "responder_type": responder_type,
            "fire_station_name": self.get("D20")
            if responder_type == "First Responder"
            else self.get("D21"),
            "notification_date": self.get("D22"),
            "notification_time": self.get("D23"),
            "region": self.get("D24"),
            "province": self.get("D25"),
            "city": self.get("D26"),
            "address": self.get("D27"),
            "landmark": self.get("D28"),
            "caller_info": self.get("D29"),
            "receiver": self.get("D30"),
            "engine": self.get("D31"),
            "time_dispatched": self.get("D34"),
            "time_arrived": self.get("D37"),
            "response_time": self.get("D40"),
            "distance_km": self.get("D41"),
            "alarm_level": self.get("D42"),
            "time_returned": self.get("D43"),
            "gas_liters": self.get("D44"),
            "classification": classification,
            "category": cat_val,
            "owner": self.get("D51"),
            "description": self.get("D52"),
            "origin": self.get("D53"),
            "stage": stage,
            "extent": extent,
            "extent_total_floor_area_sqm": self.get("D56")
            or self.get("D57")
            or self.get("D58")
            or self.get("D59")
            or self.get("D60"),
            "extent_total_land_area_hectares": self.get("D59") or self.get("D60"),
            "struct_aff": self.get("D62"),
            "house_aff": self.get("D63"),
            "fam_aff": self.get("D64"),
            "indiv_aff": self.get("D65"),
            "vehic_aff": self.get("D66"),
            "res_bfp_truck": self.get("D70"),
            "res_lgu_truck": self.get("D71"),
            "res_vol_truck": self.get("D72"),
            "res_bfp_amb": self.get("D73"),
            "res_non_amb": self.get("D74"),
            "res_bfp_resc": self.get("D75"),
            "res_non_resc": self.get("D76"),
            "res_others": self.get("D77"),
            "tool_scba": self.get("D79"),
            "tool_rope": self.get("D80"),
            "tool_ladder": self.get("D81"),
            "tool_hose": self.get("D82"),
            "tool_hydra": self.get("D83"),
            "tool_others": self.get("D84"),
            "hydrant_dist": self.get("D85"),
            "timeline": {
                "alarm_1st": {"time": self.get("D89"), "date": self.get("E89")},
                "alarm_2nd": {"time": self.get("D90"), "date": self.get("E90")},
                "alarm_3rd": {"time": self.get("D91"), "date": self.get("E91")},
                "alarm_4th": {"time": self.get("D92"), "date": self.get("E92")},
                "alarm_5th": {"time": self.get("D93"), "date": self.get("E93")},
                "tf_alpha": {"time": self.get("D94"), "date": self.get("E94")},
                "tf_bravo": {"time": self.get("D95"), "date": self.get("E95")},
                "tf_charlie": {"time": self.get("D96"), "date": self.get("E96")},
                "tf_delta": {"time": self.get("D97"), "date": self.get("E97")},
                "general": {"time": self.get("D98"), "date": self.get("E98")},
                "fuc": {"time": self.get("D99"), "date": self.get("E99")},
                "fo": {"time": self.get("D100"), "date": self.get("E100")},
            },
            "icp_present": icp_present,
            "icp_location": icp_location,
            "inj_civ_m": inj_civ_m,
            "inj_civ_f": inj_civ_f,
            "inj_bfp_m": inj_bfp_m,
            "inj_bfp_f": inj_bfp_f,
            "inj_aux_m": inj_aux_m,
            "inj_aux_f": inj_aux_f,
            "fat_civ_m": fat_civ_m,
            "fat_civ_f": fat_civ_f,
            "fat_bfp_m": fat_bfp_m,
            "fat_bfp_f": fat_bfp_f,
            "fat_aux_m": fat_aux_m,
            "fat_aux_f": fat_aux_f,
            "pod_commander": self.get("D114"),
            "pod_shift": self.get("D115"),
            "pod_nozzleman": self.get("D116"),
            "pod_lineman": self.get("D117"),
            "pod_crew": self.get("D118"),
            "pod_dpo": self.get("D119"),
            "pod_safety": self.get("D120"),
            "others_list": others,
            "narrative": "\n".join(narrative_lines),
            "problems": problems,
            "recommendations": self.get("B222"),
            "disposition": self.get("B229"),
            "prepared_by": self.get("C238"),
            "noted_by": self.get("F238"),
            # Backward-compatible aliases used by older tests/scripts.
            "extent_of_damage": extent,
            "structures_affected": self.get("D62"),
            "res_bfp_trucks": self.get("D70"),
            "alarm_1st": self.get("D89"),
        }


def parse_afor_report_data(data: dict, region_id: int) -> AforParsedRow:
    """Map the extracted AFOR dictionary into the strict database schema."""
    errors: list[str] = []

    def _dt(d: Any, t: Any = None) -> str | None:
        if not d:
            return None

        if t:
            # Native Excel conversions often give datetime/date + datetime.time objects.
            if isinstance(d, datetime) and hasattr(t, "hour") and hasattr(t, "minute"):
                try:
                    return datetime.combine(d.date(), t).isoformat()
                except Exception:
                    pass

            # Excel serial date/time support for real filled XLSX exports.
            d_serial: float | None = None
            t_serial: float | None = None
            try:
                d_serial = float(d)
                t_serial = float(t)
            except (TypeError, ValueError):
                d_serial = None
                t_serial = None

            if d_serial is not None and t_serial is not None:
                try:
                    base = datetime(1899, 12, 30)
                    date_dt = base + timedelta(days=d_serial)
                    time_dt = base + timedelta(days=t_serial)
                    merged = datetime.combine(date_dt.date(), time_dt.time())
                    return merged.isoformat()
                except Exception:
                    pass

            date_part = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d).split(" ")[0]
            return _safe_dt(f"{date_part} {str(t).strip()}")

        return _safe_dt(d)

    notif_dt = _dt(data.get("notification_date"), data.get("notification_time"))

    # Split caller_info: "Name / 0917-..."
    ci = str(data.get("caller_info") or "")
    c_name = ci.split("/")[0].strip() if "/" in ci else ci
    c_num = ci.split("/")[1].strip() if "/" in ci else ""

    casualty_details = {
        "injured": {
            "civilian": {
                "m": _safe_int(data.get("inj_civ_m")),
                "f": _safe_int(data.get("inj_civ_f")),
            },
            "firefighter": {
                "m": _safe_int(data.get("inj_bfp_m")),
                "f": _safe_int(data.get("inj_bfp_f")),
            },
            "auxiliary": {
                "m": _safe_int(data.get("inj_aux_m")),
                "f": _safe_int(data.get("inj_aux_f")),
            },
        },
        "fatalities": {
            "civilian": {
                "m": _safe_int(data.get("fat_civ_m")),
                "f": _safe_int(data.get("fat_civ_f")),
            },
            "firefighter": {
                "m": _safe_int(data.get("fat_bfp_m")),
                "f": _safe_int(data.get("fat_bfp_f")),
            },
            "auxiliary": {
                "m": _safe_int(data.get("fat_aux_m")),
                "f": _safe_int(data.get("fat_aux_f")),
            },
        },
    }

    timeline = data.get("timeline") or {
        "alarm_1st": {
            "time": data.get("alarm_1st"),
            "date": data.get("notification_date"),
        },
        "alarm_2nd": {"time": None, "date": data.get("notification_date")},
        "alarm_3rd": {"time": None, "date": data.get("notification_date")},
        "alarm_4th": {"time": None, "date": data.get("notification_date")},
        "alarm_5th": {"time": None, "date": data.get("notification_date")},
        "tf_alpha": {"time": None, "date": data.get("notification_date")},
        "tf_bravo": {"time": None, "date": data.get("notification_date")},
        "tf_charlie": {"time": None, "date": data.get("notification_date")},
        "tf_delta": {"time": None, "date": data.get("notification_date")},
        "general": {"time": None, "date": data.get("notification_date")},
        "fuc": {"time": None, "date": data.get("notification_date")},
        "fo": {"time": None, "date": data.get("notification_date")},
    }

    incident_nonsensitive_details = {
        "notification_dt": notif_dt,
        "responder_type": data.get("responder_type"),
        "fire_station_name": data.get("fire_station_name") or "",
        "alarm_level": ALARM_LEVEL_MAP.get(
            str(data.get("alarm_level") or "").strip().upper(), data.get("alarm_level")
        ),
        "general_category": data.get("classification") or data.get("classification_of_involved"),
        "sub_category": data.get("category") or data.get("type_of_involved_general_category"),
        "fire_origin": data.get("origin") or data.get("area_of_origin"),
        "extent_of_damage": data.get("extent") or data.get("extent_of_damage"),
        "stage_of_fire": data.get("stage") or data.get("stage_of_fire_upon_arrival"),
        "structures_affected": _safe_int(
            data.get("struct_aff")
            if data.get("struct_aff") is not None
            else data.get("structures_affected")
        ),
        "households_affected": _safe_int(data.get("house_aff")),
        "families_affected": _safe_int(data.get("fam_aff")),
        "individuals_affected": _safe_int(data.get("indiv_aff")),
        "vehicles_affected": _safe_int(data.get("vehic_aff")),
        "distance_from_station_km": _safe_float(
            data.get("distance_km")
            if data.get("distance_km") is not None
            else data.get("distance_from_station_km")
        ),
        "total_response_time_minutes": _safe_int(data.get("response_time")),
        "total_gas_consumed_liters": _safe_float(data.get("gas_liters")),
        "extent_total_floor_area_sqm": _safe_float(data.get("extent_total_floor_area_sqm")),
        "extent_total_land_area_hectares": _safe_float(data.get("extent_total_land_area_hectares")),
        "resources_deployed": {
            "trucks": {
                "bfp": _safe_int(
                    data.get("res_bfp_truck")
                    if data.get("res_bfp_truck") is not None
                    else data.get("res_bfp_trucks")
                ),
                "lgu": _safe_int(data.get("res_lgu_truck")),
                "volunteer": _safe_int(data.get("res_vol_truck")),
            },
            "medical": {
                "bfp": _safe_int(data.get("res_bfp_amb")),
                "non_bfp": _safe_int(data.get("res_non_amb")),
            },
            "special_assets": {
                "rescue_bfp": _safe_int(data.get("res_bfp_resc")),
                "rescue_non_bfp": _safe_int(data.get("res_non_resc")),
                "others": str(data.get("res_others") or ""),
            },
            "tools": {
                "scba": _safe_int(data.get("tool_scba")),
                "rope": str(data.get("tool_rope") or ""),
                "ladder": _safe_int(data.get("tool_ladder")),
                "hoseline": str(data.get("tool_hose") or ""),
                "hydraulic": _safe_int(data.get("tool_hydra")),
                "others": str(data.get("tool_others") or ""),
            },
            "hydrant_distance": str(data.get("hydrant_dist") or ""),
        },
        "alarm_timeline": {
            "alarm_1st": _dt(timeline["alarm_1st"]["date"], timeline["alarm_1st"]["time"]),
            "alarm_2nd": _dt(timeline["alarm_2nd"]["date"], timeline["alarm_2nd"]["time"]),
            "alarm_3rd": _dt(timeline["alarm_3rd"]["date"], timeline["alarm_3rd"]["time"]),
            "alarm_4th": _dt(timeline["alarm_4th"]["date"], timeline["alarm_4th"]["time"]),
            "alarm_5th": _dt(timeline["alarm_5th"]["date"], timeline["alarm_5th"]["time"]),
            "alarm_tf_alpha": _dt(timeline["tf_alpha"]["date"], timeline["tf_alpha"]["time"]),
            "alarm_tf_bravo": _dt(timeline["tf_bravo"]["date"], timeline["tf_bravo"]["time"]),
            "alarm_tf_charlie": _dt(timeline["tf_charlie"]["date"], timeline["tf_charlie"]["time"]),
            "alarm_tf_delta": _dt(timeline["tf_delta"]["date"], timeline["tf_delta"]["time"]),
            "alarm_general": _dt(timeline["general"]["date"], timeline["general"]["time"]),
            "alarm_fuc": _dt(timeline["fuc"]["date"], timeline["fuc"]["time"]),
            "alarm_fo": _dt(timeline["fo"]["date"], timeline["fo"]["time"]),
        },
        "problems_encountered": data.get("problems", []),
        "recommendations": data.get("recommendations") or "",
    }

    mapped = {
        "region_id": region_id,
        "incident_nonsensitive_details": incident_nonsensitive_details,
        "incident_sensitive_details": {
            "caller_name": c_name,
            "caller_number": c_num,
            "receiver_name": data.get("receiver") or "",
            "owner_name": data.get("owner") or "",
            "establishment_name": data.get("owner") or "",
            "street_address": data.get("address") or "",
            "landmark": data.get("landmark") or "",
            "personnel_on_duty": {
                "engine_commander": data.get("pod_commander") or "",
                "shift_in_charge": data.get("pod_shift") or "",
                "nozzleman": data.get("pod_nozzleman") or "",
                "lineman": data.get("pod_lineman") or "",
                "engine_crew": data.get("pod_crew") or "",
                "driver": data.get("pod_dpo") or "",
                "pump_operator": data.get("pod_dpo") or "",
                "safety_officer": {"name": data.get("pod_safety") or "", "contact": ""},
            },
            "other_personnel": data.get("others_list", []),
            "casualty_details": casualty_details,
            "narrative_report": data.get("narrative") or "",
            "disposition": data.get("disposition") or "",
            "disposition_prepared_by": data.get("prepared_by") or "",
            "disposition_noted_by": data.get("noted_by") or "",
            "prepared_by_officer": data.get("prepared_by") or "",
            "noted_by_officer": data.get("noted_by") or "",
            "is_icp_present": bool(data.get("icp_present")),
            "icp_location": data.get("icp_location") or "",
        },
        "responding_unit": {
            "station_name": data.get("fire_station_name") or "",
            "engine_number": data.get("engine") or "",
            "responder_type": data.get("responder_type") or "",
            "dispatch_dt": _combine_date_and_time(notif_dt, data.get("time_dispatched")),
            "arrival_dt": _combine_date_and_time(notif_dt, data.get("time_arrived")),
            "return_dt": _combine_date_and_time(notif_dt, data.get("time_returned")),
        },
        "_city_text": data.get("city") or "",
        "_province_text": data.get("province") or "",
    }

    if not notif_dt:
        errors.append("Missing required fields: notification_dt (Check D22/D23 in XLSX)")
    if not mapped["_city_text"]:
        errors.append("Missing required fields: _city_text (City/Municipality)")

    mapped["_form_kind"] = "STRUCTURAL_AFOR"

    status = "VALID" if not errors else "INVALID"
    return AforParsedRow(row_index=0, status=status, errors=errors, data=mapped)


def parse_csv_content(content: str, region_id: int) -> tuple[list[AforParsedRow], AforFormKind]:
    """Parse either the official AFOR form-style CSV or a flat tabular CSV (structural only)."""
    rows = list(csv.reader(io.StringIO(content)))
    if _looks_like_official_afor_csv(rows):
        parser = BfpXlsxParser(CsvWorksheetAdapter(rows))
        return [parse_afor_report_data(parser.parse(), region_id)], "STRUCTURAL_AFOR"

    reader = csv.DictReader(io.StringIO(content))
    results = []
    for row in reader:
        if not any(row.values()):
            continue
        results.append(parse_afor_report_data(row, region_id))
    return results, "STRUCTURAL_AFOR"


def parse_xlsx_content(content: bytes, region_id: int) -> tuple[list[AforParsedRow], AforFormKind]:
    """Parse XLSX: detect structural vs wildland, then dispatch."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True)
    try:
        kind = detect_afor_template_kind(wb)
        if kind is None:
            raise ValueError(
                "could not determine AFOR type. Expected either the official structural AFOR "
                "(column A: 'AFTER FIRE OPERATIONS REPORT' and 'A. RESPONSE DETAILS') or the wildland "
                "template (sheet 'WILDLAND FIRE AFOR' with section A dates in column B). "
                "See public/templates/ for sample workbooks."
            )

        if kind == "STRUCTURAL_AFOR":
            ws = _pick_structural_worksheet(wb)
            parser = BfpXlsxParser(ws)
            report_data = parser.parse()
            parsed_row = parse_afor_report_data(report_data, region_id)
            return [parsed_row], kind

        ws = _pick_wildland_worksheet(wb)
        parser = WildlandXlsxParser(ws)
        report_data = parser.parse()
        parsed_row = parse_wildland_afor_report_data(report_data, region_id)
        return [parsed_row], kind
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/afor/import", response_model=AforParseResponse)
async def import_afor_file(
    file: UploadFile = File(...),
    user: dict = Depends(get_regional_encoder),
    db: Session = Depends(get_db_with_rls),
):
    """
    Upload and parse an AFOR file (.xlsx or .csv).
    Returns parsed rows with validation status for preview before commit.
    """
    region_id = user["assigned_region_id"]

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(
            status_code=400, detail="Only .xlsx, .xls, and .csv files are supported"
        )

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File is empty")

    try:
        if ext == "csv":
            decoded = content.decode("utf-8-sig")  # Handle BOM
            rows, form_kind = parse_csv_content(decoded, region_id)
        else:
            rows, form_kind = parse_xlsx_content(content, region_id)
    except ValueError as e:
        logger.warning("AFOR type detection failed: %s", e)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception:
        logger.exception("Failed to parse AFOR file")
        raise HTTPException(status_code=400, detail="Failed to parse file")

    if len(rows) == 0:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    valid_count = sum(1 for r in rows if r.status == "VALID")

    return AforParseResponse(
        total_rows=len(rows),
        valid_rows=valid_count,
        invalid_rows=len(rows) - valid_count,
        rows=rows,
        form_kind=form_kind,
    )


_WILDLAND_ALARM_STATUS_ALLOWED = {
    "1st Alarm",
    "2nd Alarm",
    "3rd Alarm",
    "4th Alarm",
    "Task Force Alpha",
    "Task Force Bravo",
    "General Alarm",
    "Ongoing",
    "Fire Out",
    "Fire Under Control",
    "Fire Out Upon Arrival",
    "Fire Under Investigation",
    "Late Reported",
    "Unresponded",
    "No Firefighting Conducted",
}


def _dt_for_sql(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return val


def _commit_wildland_afor_row(
    db: Session,
    row_data: dict[str, Any],
    batch_id: int,
    user_id: Any,
    region_id: int,
    incident_ids: list[int],
    lon: float,
    lat: float,
    *,
    source: WildlandRowSource = "AFOR_IMPORT",
) -> None:
    """Insert fire_incident + incident_wildland_afor + optional alarm/assistance children."""
    wl = dict(row_data.get("wildland") or {})
    alarm_statuses: list[dict[str, Any]] = list(wl.pop("wildland_alarm_statuses", []) or [])
    assistance_rows: list[dict[str, Any]] = list(wl.pop("wildland_assistance_rows", []) or [])

    inc_row = db.execute(
        text("""
            INSERT INTO wims.fire_incidents
                (import_batch_id, encoder_id, region_id, location, verification_status)
            VALUES
                (:batch_id, CAST(:uid AS uuid), :region_id,
                 ST_SetSRID(ST_MakePoint(:lon, :lat), 4326)::geography,
                 'DRAFT')
            RETURNING incident_id
        """),
        {
            "batch_id": batch_id,
            "uid": user_id,
            "region_id": region_id,
            "lon": lon,
            "lat": lat,
        },
    ).fetchone()

    if not inc_row:
        return

    incident_id = inc_row[0]
    incident_ids.append(incident_id)

    params = {
        "incident_id": incident_id,
        "batch_id": batch_id,
        "source": source,
        "call_received_at": _dt_for_sql(wl.get("call_received_at")),
        "fire_started_at": _dt_for_sql(wl.get("fire_started_at")),
        "fire_arrival_at": _dt_for_sql(wl.get("fire_arrival_at")),
        "fire_controlled_at": _dt_for_sql(wl.get("fire_controlled_at")),
        "caller_transmitted_by": wl.get("caller_transmitted_by") or "",
        "caller_office_address": wl.get("caller_office_address") or "",
        "call_received_by_personnel": wl.get("call_received_by_personnel") or "",
        "engine_dispatched": wl.get("engine_dispatched") or "",
        "incident_location_description": wl.get("incident_location_description") or "",
        "distance_to_fire_station_km": wl.get("distance_to_fire_station_km"),
        "primary_action_taken": wl.get("primary_action_taken") or "",
        "assistance_combined_summary": wl.get("assistance_combined_summary") or "",
        "buildings_involved": wl.get("buildings_involved") or 0,
        "buildings_threatened": wl.get("buildings_threatened") or 0,
        "ownership_and_property_notes": wl.get("ownership_and_property_notes") or "",
        "total_area_burned_display": wl.get("total_area_burned_display") or "",
        "total_area_burned_hectares": wl.get("total_area_burned_hectares"),
        "wildland_fire_type": wl.get("wildland_fire_type") or None,
        "area_type_summary": json.dumps(wl.get("area_type_summary") or {}),
        "causes_and_ignition_factors": json.dumps(wl.get("causes_and_ignition_factors") or {}),
        "suppression_factors": json.dumps(wl.get("suppression_factors") or {}),
        "weather": json.dumps(wl.get("weather") or {}),
        "fire_behavior": json.dumps(wl.get("fire_behavior") or {}),
        "peso_losses": json.dumps(wl.get("peso_losses") or {}),
        "casualties": json.dumps(wl.get("casualties") or {}),
        "narration": wl.get("narration") or "",
        "problems_encountered": json.dumps(wl.get("problems_encountered") or []),
        "recommendations": json.dumps(wl.get("recommendations") or []),
        "prepared_by": wl.get("prepared_by") or "",
        "prepared_by_title": wl.get("prepared_by_title") or "",
        "noted_by": wl.get("noted_by") or "",
        "noted_by_title": wl.get("noted_by_title") or "",
    }

    iwa_row = db.execute(
        text("""
            INSERT INTO wims.incident_wildland_afor (
                incident_id, import_batch_id, source,
                call_received_at, fire_started_at, fire_arrival_at, fire_controlled_at,
                caller_transmitted_by, caller_office_address, call_received_by_personnel,
                engine_dispatched, incident_location_description, distance_to_fire_station_km,
                primary_action_taken, assistance_combined_summary,
                buildings_involved, buildings_threatened, ownership_and_property_notes,
                total_area_burned_display, total_area_burned_hectares, wildland_fire_type,
                area_type_summary, causes_and_ignition_factors, suppression_factors,
                weather, fire_behavior, peso_losses, casualties,
                narration, problems_encountered, recommendations,
                prepared_by, prepared_by_title, noted_by, noted_by_title
            ) VALUES (
                :incident_id, :batch_id, :source,
                CAST(:call_received_at AS timestamptz),
                CAST(:fire_started_at AS timestamptz),
                CAST(:fire_arrival_at AS timestamptz),
                CAST(:fire_controlled_at AS timestamptz),
                :caller_transmitted_by, :caller_office_address, :call_received_by_personnel,
                :engine_dispatched, :incident_location_description, :distance_to_fire_station_km,
                :primary_action_taken, :assistance_combined_summary,
                :buildings_involved, :buildings_threatened, :ownership_and_property_notes,
                :total_area_burned_display, :total_area_burned_hectares, :wildland_fire_type,
                CAST(:area_type_summary AS jsonb), CAST(:causes_and_ignition_factors AS jsonb),
                CAST(:suppression_factors AS jsonb),
                CAST(:weather AS jsonb), CAST(:fire_behavior AS jsonb),
                CAST(:peso_losses AS jsonb), CAST(:casualties AS jsonb),
                :narration, CAST(:problems_encountered AS jsonb), CAST(:recommendations AS jsonb),
                :prepared_by, :prepared_by_title, :noted_by, :noted_by_title
            )
            RETURNING incident_wildland_afor_id
        """),
        params,
    ).fetchone()

    if not iwa_row:
        return

    iwa_id = iwa_row[0]

    for order, a in enumerate(alarm_statuses):
        status = (a.get("alarm_status") or "").strip()
        if status not in _WILDLAND_ALARM_STATUS_ALLOWED:
            continue
        db.execute(
            text("""
                INSERT INTO wims.wildland_afor_alarm_statuses (
                    incident_wildland_afor_id, sort_order, alarm_status, time_declared, ground_commander
                ) VALUES (
                    :iwa_id, :sort_order, :alarm_status, :time_declared, :ground_commander
                )
            """),
            {
                "iwa_id": iwa_id,
                "sort_order": order,
                "alarm_status": status,
                "time_declared": a.get("time_declared") or "",
                "ground_commander": a.get("ground_commander") or "",
            },
        )

    for order, row in enumerate(assistance_rows):
        org = (row.get("organization_or_unit") or row.get("organization") or "").strip()
        if not org:
            continue
        db.execute(
            text("""
                INSERT INTO wims.wildland_afor_assistance_rows (
                    incident_wildland_afor_id, sort_order, organization_or_unit, detail
                ) VALUES (
                    :iwa_id, :sort_order, :organization_or_unit, :detail
                )
            """),
            {
                "iwa_id": iwa_id,
                "sort_order": order,
                "organization_or_unit": org,
                "detail": row.get("detail") or "",
            },
        )


def _extract_row_match_fields(row_data: dict[str, Any], form_kind: AforFormKind) -> dict[str, Any]:
    """Extract the fields used for duplicate matching from one parsed row.

    Returns a dict with: alarm_level, general_category, notification_dt (date), fire_station_name.
    Missing values are returned as None — only present fields participate in match counting.
    """
    if form_kind == "WILDLAND_AFOR":
        wl = row_data.get("wildland") or {}
        notification_dt = wl.get("call_received_at") or wl.get("incident_date")
        return {
            "alarm_level": wl.get("alarm_level"),
            "general_category": "WILDLAND",
            "notification_dt": str(notification_dt)[:10] if notification_dt else None,
            "fire_station_name": wl.get("fire_station_name") or wl.get("station_name"),
        }
    ns = row_data.get("incident_nonsensitive_details") or {}
    notification_dt = ns.get("notification_dt")
    return {
        "alarm_level": (ns.get("alarm_level") or "").strip() or None,
        "general_category": _normalize_general_category(ns.get("general_category", "") or "")
        or None,
        "notification_dt": str(notification_dt)[:10] if notification_dt else None,
        "fire_station_name": (ns.get("fire_station_name") or "").strip() or None,
    }


def _find_duplicates(
    db: Session,
    rows: list[dict[str, Any]],
    region_id: int,
    lon: float,
    lat: float,
    form_kind: AforFormKind,
) -> list[dict[str, Any]]:
    """M4-D: For each incoming row, find existing fire_incidents within 1km that
    match on at least DUPLICATE_MIN_MATCHING_FIELDS fields. Returns one entry per
    duplicate row with the matched incident_id, distance, and matched fields.
    """
    duplicates: list[dict[str, Any]] = []

    candidates = db.execute(
        text("""
            SELECT
                fi.incident_id,
                ST_Distance(
                    fi.location::geography,
                    ST_SetSRID(ST_MakePoint(:lon, :lat), 4326)::geography
                ) AS distance_m,
                nd.alarm_level,
                nd.general_category,
                nd.notification_dt,
                nd.fire_station_name
            FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE fi.region_id = :region_id
              AND fi.is_archived = FALSE
              AND fi.verification_status != 'REJECTED'
              AND ST_DWithin(
                  fi.location::geography,
                  ST_SetSRID(ST_MakePoint(:lon, :lat), 4326)::geography,
                  :radius
              )
        """),
        {"lon": lon, "lat": lat, "region_id": region_id, "radius": DUPLICATE_RADIUS_METERS},
    ).fetchall()

    if not candidates:
        return duplicates

    for row_index, row_data in enumerate(rows):
        incoming = _extract_row_match_fields(row_data, form_kind)
        best_match: dict[str, Any] | None = None
        for cand in candidates:
            cand_existing = {
                "alarm_level": cand[2],
                "general_category": cand[3],
                "notification_dt": str(cand[4])[:10] if cand[4] else None,
                "fire_station_name": cand[5],
            }
            matched_fields: list[str] = []
            for key, incoming_val in incoming.items():
                cand_val = cand_existing.get(key)
                if (
                    incoming_val is not None
                    and cand_val is not None
                    and str(incoming_val).strip().lower() == str(cand_val).strip().lower()
                ):
                    matched_fields.append(key)
            if len(matched_fields) >= DUPLICATE_MIN_MATCHING_FIELDS:
                if best_match is None or len(matched_fields) > len(best_match["matched_fields"]):
                    best_match = {
                        "row_index": row_index,
                        "existing_incident_id": cand[0],
                        "distance_m": float(cand[1]) if cand[1] is not None else 0.0,
                        "matched_fields": matched_fields,
                        "incoming_values": incoming,
                        "existing_values": cand_existing,
                    }
        if best_match is not None:
            duplicates.append(best_match)

    return duplicates


@router.post("/afor/commit")
async def commit_afor_import(
    request: Request,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """
    Commit validated AFOR rows to the database.
    Creates a data_import_batch and inserts fire_incidents with details.
    """
    try:
        raw_body: dict[str, Any] = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body") from None

    body = AforCommitRequest.model_validate(raw_body)
    lon, lat = _wgs84_pair_from_raw(raw_body.get("latitude"), raw_body.get("longitude"))

    region_id = user["assigned_region_id"]
    user_id = user["user_id"]

    if not body.rows:
        raise HTTPException(status_code=400, detail="No rows to commit")

    for row_data in body.rows:
        rk = row_data.get("_form_kind")
        if rk != body.form_kind:
            raise HTTPException(
                status_code=400,
                detail="form_kind mismatch: preview rows do not match commit form_kind",
            )

    validated_wildland_rows: list[dict[str, Any]] | None = None
    if body.form_kind == "WILDLAND_AFOR":
        wildland_errors: list[str] = []
        validated_wildland_rows = []
        for idx, row_data in enumerate(body.rows):
            wl_dict = row_data.get("wildland") or {}
            parsed = parse_wildland_afor_report_data(wl_dict, region_id)
            if parsed.status != "VALID":
                for err in parsed.errors:
                    wildland_errors.append(f"Row {idx + 1}: {err}")
            else:
                validated_wildland_rows.append(parsed.data)
        if wildland_errors:
            raise HTTPException(status_code=400, detail=" ".join(wildland_errors))

    # ── M4-D: Multi-factor duplicate pre-check ───────────────────────────────
    # First call (resolutions=None): scan; if duplicates found, return without inserting.
    # Second call (resolutions=[...]): apply per-row decisions (skip/merge/force).
    if body.resolutions is None:
        duplicates = _find_duplicates(db, body.rows, region_id, lon, lat, body.form_kind)
        if duplicates:
            return {
                "status": "DUPLICATE_CHECK_REQUIRED",
                "duplicates": duplicates,
                "radius_meters": DUPLICATE_RADIUS_METERS,
                "min_matching_fields": DUPLICATE_MIN_MATCHING_FIELDS,
            }

    # Build the resolution map keyed by row_index for fast lookup.
    resolution_map: dict[int, RowResolution] = {}
    if body.resolutions:
        for r in body.resolutions:
            resolution_map[r.row_index] = r

    # Create import batch
    batch_row = db.execute(
        text("""
            INSERT INTO wims.data_import_batches (region_id, uploaded_by, record_count)
            VALUES (:region_id, CAST(:uid AS uuid), :count)
            RETURNING batch_id
        """),
        {"region_id": region_id, "uid": user_id, "count": len(body.rows)},
    ).fetchone()

    if not batch_row:
        raise HTTPException(status_code=500, detail="Failed to create import batch")

    batch_id = batch_row[0]
    incident_ids: list[int] = []

    wildland_source: WildlandRowSource = (
        "MANUAL" if body.wildland_row_source == "MANUAL" else "AFOR_IMPORT"
    )

    def _group_total(groups: dict[str, Any], key: str) -> int:
        bucket = groups.get(key, {}) if isinstance(groups, dict) else {}
        return _safe_int(bucket.get("m")) + _safe_int(bucket.get("f"))

    for idx, row_data in enumerate(body.rows):
        # M4-D: skip rows the encoder explicitly chose to skip
        resolution = resolution_map.get(idx)
        if resolution is not None and resolution.action == "skip":
            continue

        if body.form_kind == "WILDLAND_AFOR":
            assert validated_wildland_rows is not None
            _commit_wildland_afor_row(
                db,
                validated_wildland_rows[idx],
                batch_id,
                user_id,
                region_id,
                incident_ids,
                lon,
                lat,
                source=wildland_source,
            )
            continue

        # M4-D: merge into existing incident — UPDATE rather than INSERT
        if resolution is not None and resolution.action == "merge":
            if resolution.existing_incident_id is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Row {idx}: merge action requires existing_incident_id",
                )
            existing_id = resolution.existing_incident_id
            ns_merge = row_data.get("incident_nonsensitive_details", {}) or {}
            db.execute(
                text("""
                    UPDATE wims.incident_nonsensitive_details SET
                        notification_dt = COALESCE(CAST(:notification_dt AS timestamptz), notification_dt),
                        alarm_level = COALESCE(NULLIF(:alarm_level, ''), alarm_level),
                        general_category = COALESCE(NULLIF(:general_category, ''), general_category),
                        sub_category = COALESCE(NULLIF(:sub_category, ''), sub_category),
                        fire_station_name = COALESCE(NULLIF(:fire_station_name, ''), fire_station_name),
                        structures_affected = COALESCE(:structures_affected, structures_affected),
                        households_affected = COALESCE(:households_affected, households_affected),
                        individuals_affected = COALESCE(:individuals_affected, individuals_affected),
                        families_affected = COALESCE(:families_affected, families_affected)
                    WHERE incident_id = :iid
                """),
                {
                    "iid": existing_id,
                    "notification_dt": ns_merge.get("notification_dt"),
                    "alarm_level": ns_merge.get("alarm_level", "") or "",
                    "general_category": _normalize_general_category(
                        ns_merge.get("general_category", "") or ""
                    )
                    or "",
                    "sub_category": ns_merge.get("sub_category", "") or "",
                    "fire_station_name": ns_merge.get("fire_station_name", "") or "",
                    "structures_affected": ns_merge.get("structures_affected"),
                    "households_affected": ns_merge.get("households_affected"),
                    "individuals_affected": ns_merge.get("individuals_affected"),
                    "families_affected": ns_merge.get("families_affected"),
                },
            )
            db.execute(
                text("UPDATE wims.fire_incidents SET updated_at = now() WHERE incident_id = :iid"),
                {"iid": existing_id},
            )
            incident_ids.append(existing_id)
            continue

        ns = row_data.get("incident_nonsensitive_details", {})
        sens = row_data.get("incident_sensitive_details", {})
        casualty_details = (
            sens.get("casualty_details", {})
            if isinstance(sens.get("casualty_details", {}), dict)
            else {}
        )
        injured_groups = (
            casualty_details.get("injured", {})
            if isinstance(casualty_details.get("injured", {}), dict)
            else {}
        )
        fatal_groups = (
            casualty_details.get("fatalities", {}) or casualty_details.get("fatal", {}) or {}
        )

        inc_row = db.execute(
            text("""
                INSERT INTO wims.fire_incidents
                    (import_batch_id, encoder_id, region_id, location, verification_status)
                VALUES
                    (:batch_id, CAST(:uid AS uuid), :region_id,
                     ST_SetSRID(ST_MakePoint(:lon, :lat), 4326)::geography,
                     'DRAFT')
                RETURNING incident_id
            """),
            {
                "batch_id": batch_id,
                "uid": user_id,
                "region_id": region_id,
                "lon": lon,
                "lat": lat,
            },
        ).fetchone()

        if not inc_row:
            continue

        incident_id = inc_row[0]
        incident_ids.append(incident_id)

        city_text = row_data.get("_city_text", "")
        geo_ids = db.execute(
            text("""
                SELECT c.city_id
                FROM wims.ref_cities c
                WHERE LOWER(c.city_name) = LOWER(:city)
                LIMIT 1
            """),
            {"city": city_text},
        ).fetchone()
        city_id = geo_ids[0] if geo_ids else None

        db.execute(
            text("""
                INSERT INTO wims.incident_nonsensitive_details (
                    incident_id, city_id, distance_from_station_km, notification_dt,
                    alarm_level, general_category, sub_category,
                    civilian_injured, civilian_deaths, firefighter_injured, firefighter_deaths,
                    families_affected, responder_type, fire_origin, extent_of_damage,
                    structures_affected, households_affected, individuals_affected,
                    resources_deployed, alarm_timeline, problems_encountered, recommendations,
                    fire_station_name, total_response_time_minutes, total_gas_consumed_liters,
                    stage_of_fire, extent_total_floor_area_sqm, extent_total_land_area_hectares,
                    vehicles_affected
                ) VALUES (
                    :incident_id, :city_id, :distance_from_station_km, CAST(:notification_dt AS timestamptz),
                    :alarm_level, :general_category, :sub_category,
                    :civ_inj, :civ_fat, :ff_inj, :ff_fat,
                    :families_affected, :responder_type, :fire_origin, :extent_of_damage,
                    :structures_affected, :households_affected, :individuals_affected,
                    CAST(:resources_deployed AS jsonb), CAST(:alarm_timeline AS jsonb),
                    CAST(:problems_encountered AS jsonb), :recommendations,
                    :fire_station_name, :total_response_time_minutes, :total_gas_consumed_liters,
                    :stage_of_fire, :floor_area, :land_area, :vehicles_affected
                )
            """),
            {
                "incident_id": incident_id,
                "city_id": city_id,
                "distance_from_station_km": ns.get("distance_from_station_km"),
                "notification_dt": ns.get("notification_dt"),
                "alarm_level": ns.get("alarm_level", ""),
                "general_category": _normalize_general_category(
                    ns.get("general_category", "") or ""
                ),
                "sub_category": ns.get("sub_category", ""),
                "civ_inj": _group_total(injured_groups, "civilian"),
                "civ_fat": _group_total(fatal_groups, "civilian"),
                "ff_inj": _group_total(injured_groups, "firefighter"),
                "ff_fat": _group_total(fatal_groups, "firefighter"),
                "families_affected": ns.get("families_affected", 0),
                "responder_type": ns.get("responder_type", ""),
                "fire_origin": ns.get("fire_origin", ""),
                "extent_of_damage": ns.get("extent_of_damage", ""),
                "structures_affected": ns.get("structures_affected", 0),
                "households_affected": ns.get("households_affected", 0),
                "individuals_affected": ns.get("individuals_affected", 0),
                "resources_deployed": json.dumps(ns.get("resources_deployed", {})),
                "alarm_timeline": json.dumps(ns.get("alarm_timeline", {})),
                "problems_encountered": json.dumps(ns.get("problems_encountered", [])),
                "recommendations": ns.get("recommendations", ""),
                "fire_station_name": ns.get("fire_station_name", ""),
                "total_response_time_minutes": ns.get("total_response_time_minutes", 0),
                "total_gas_consumed_liters": ns.get("total_gas_consumed_liters", 0),
                "stage_of_fire": ns.get("stage_of_fire", ""),
                "floor_area": ns.get("extent_total_floor_area_sqm", 0),
                "land_area": ns.get("extent_total_land_area_hectares", 0),
                "vehicles_affected": ns.get("vehicles_affected", 0),
            },
        )

        # ── Encrypt PII fields before INSERT ─────────────────────────────────────
        # PII fields (caller_name, caller_number, owner_name, occupant_name) are
        # stored ONLY in the encrypted blob. Plaintext columns are set to NULL.
        # receiver_name is NOT encrypted (public / internal use only).
        #
        # caller_info arrives as "Name / Number" at the top-level row_data field;
        # owner_name / occupant_name arrive in sens (incident_sensitive_details body).
        ci = str(row_data.get("caller_info") or "").strip()
        caller_name_row, caller_number_row = "", ""

        if ci:
            if "/" in ci:
                left, right = ci.split("/", 1)
                caller_name_row = left.strip()
                caller_number_row = right.strip()
            else:
                caller_name_row = ci

        pii_for_blob = {
            k: v
            for k, v in (
                ("caller_name", caller_name_row),
                ("caller_number", caller_number_row),
                ("owner_name", sens.get("owner_name")),
                ("occupant_name", sens.get("occupant_name")),
            )
            if v  # omit None and empty strings
        }
        # Always produce a dict (empty or populated) so decrypt never raises on None
        if not pii_for_blob:
            pii_for_blob = {}

        aad = f"incident_id:{incident_id}".encode("utf-8")
        nonce_b64: str | None = None
        ct_b64: str | None = None
        try:
            sp = _get_security_provider()
            nonce_b64, ct_b64 = sp.encrypt_json(pii_for_blob, aad)
        except SecurityProviderError as exc:
            logger.warning(
                "PII encryption unavailable for incident_id=%s during AFOR commit; proceeding without encrypted blob (%s)",
                incident_id,
                exc,
            )

        db.execute(
            text("""
                INSERT INTO wims.incident_sensitive_details (
                    incident_id, street_address, landmark,
                    caller_name, caller_number, receiver_name,
                    owner_name, establishment_name,
                    narrative_report, disposition,
                    disposition_prepared_by, disposition_noted_by,
                    prepared_by_officer, noted_by_officer,
                    personnel_on_duty, other_personnel, casualty_details,
                    is_icp_present, icp_location,
                    pii_blob_enc, encryption_iv
                ) VALUES (
                    :incident_id, :street_address, :landmark,
                    NULL, NULL, :receiver_name,
                    NULL, :establishment_name,
                    :narrative_report, :disposition,
                    :disposition_prepared_by, :disposition_noted_by,
                    :prepared_by_officer, :noted_by_officer,
                    CAST(:personnel_on_duty AS jsonb),
                    CAST(:other_personnel AS jsonb),
                    CAST(:casualty_details AS jsonb),
                    :is_icp_present, :icp_location,
                    :pii_blob_enc, :pii_nonce
                )
            """),
            {
                "incident_id": incident_id,
                "street_address": sens.get("street_address", ""),
                "landmark": sens.get("landmark", ""),
                # Plaintext PII columns → NULL; only pii_blob_enc is authoritative
                "receiver_name": sens.get("receiver_name", ""),
                "establishment_name": sens.get("establishment_name", ""),
                "narrative_report": sens.get("narrative_report", ""),
                "disposition": sens.get("disposition", ""),
                "disposition_prepared_by": sens.get("disposition_prepared_by", ""),
                "disposition_noted_by": sens.get("disposition_noted_by", ""),
                "prepared_by_officer": sens.get("prepared_by_officer", ""),
                "noted_by_officer": sens.get("noted_by_officer", ""),
                "personnel_on_duty": json.dumps(sens.get("personnel_on_duty", {})),
                "other_personnel": json.dumps(sens.get("other_personnel", [])),
                "casualty_details": json.dumps(casualty_details),
                "is_icp_present": sens.get("is_icp_present", False),
                "icp_location": sens.get("icp_location", ""),
                # Encrypted PII blob
                "pii_blob_enc": ct_b64,
                "pii_nonce": nonce_b64,
            },
        )

        responding_unit = row_data.get("responding_unit", {})
        if any(
            responding_unit.get(key)
            for key in (
                "station_name",
                "engine_number",
                "dispatch_dt",
                "arrival_dt",
                "return_dt",
            )
        ):
            db.execute(
                text("""
                    INSERT INTO wims.responding_units (
                        incident_id, station_name, engine_number, responder_type,
                        dispatch_dt, arrival_dt, return_dt
                    ) VALUES (
                        :incident_id, :station_name, :engine_number, :responder_type,
                        CAST(:dispatch_dt AS timestamptz),
                        CAST(:arrival_dt AS timestamptz),
                        CAST(:return_dt AS timestamptz)
                    )
                """),
                {
                    "incident_id": incident_id,
                    "station_name": responding_unit.get("station_name", ""),
                    "engine_number": responding_unit.get("engine_number", ""),
                    "responder_type": responding_unit.get("responder_type", ""),
                    "dispatch_dt": responding_unit.get("dispatch_dt"),
                    "arrival_dt": responding_unit.get("arrival_dt"),
                    "return_dt": responding_unit.get("return_dt"),
                },
            )

    db.commit()

    # Sync analytics read model (only VERIFIED non-archived will appear in facts)
    sync_incidents_batch(db, incident_ids)
    db.commit()

    return AforCommitResponse(
        status="ok",
        batch_id=batch_id,
        incident_ids=incident_ids,
        total_committed=len(incident_ids),
    )


@router.get("/incidents")
def get_regional_incidents(
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    category: Optional[str] = None,
    status: Optional[str] = None,
):
    """
    Fetch fire incidents scoped to the current encoder.
    Joins nonsensitive details for summary view.
    """
    encoder_id = user["user_id"]

    where_clauses = [
        "fi.encoder_id = CAST(:encoder_id AS uuid)",
        "fi.is_archived = FALSE",
    ]
    params: dict[str, Any] = {
        "encoder_id": str(encoder_id),
        "limit": limit,
        "offset": offset,
    }

    if category:
        cat_key = category.strip().upper().replace("-", "_").replace(" ", "_")
        if cat_key == "TRANSPORTATION":
            cat_key = "VEHICULAR"
        variants = _CATEGORY_DB_VARIANTS.get(cat_key, [category])
        where_clauses.append("nd.general_category = ANY(:categories)")
        params["categories"] = variants
    if status:
        where_clauses.append("fi.verification_status = :status")
        params["status"] = status

    where_sql = " AND ".join(where_clauses)

    rows = db.execute(
        text(f"""
            SELECT fi.incident_id, fi.verification_status, fi.created_at,
                   nd.notification_dt, nd.general_category, nd.alarm_level,
                   nd.fire_station_name, nd.structures_affected,
                   nd.households_affected, nd.individuals_affected,
                   nd.responder_type, nd.fire_origin, nd.extent_of_damage,
                   sd.owner_name, sd.establishment_name, sd.caller_name,
                   CASE WHEN iwa.incident_id IS NOT NULL THEN TRUE ELSE FALSE END AS is_wildland,
                   fi.updated_at,
                   c.city_name, p.province_name, rr.region_name
            FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            LEFT JOIN wims.incident_sensitive_details sd ON sd.incident_id = fi.incident_id
            LEFT JOIN wims.incident_wildland_afor iwa ON iwa.incident_id = fi.incident_id
            LEFT JOIN wims.ref_cities c ON c.city_id = nd.city_id
            LEFT JOIN wims.ref_provinces p ON p.province_id = c.province_id
            LEFT JOIN wims.ref_regions rr ON rr.region_id = fi.region_id
            WHERE {where_sql}
            ORDER BY fi.updated_at DESC NULLS LAST, fi.created_at DESC
            LIMIT :limit OFFSET :offset
        """),
        params,
    ).fetchall()

    total = (
        db.execute(
            text(f"""
            SELECT COUNT(*) FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE {where_sql}
        """),
            {k: v for k, v in params.items() if k not in ("limit", "offset")},
        ).scalar()
        or 0
    )

    def _location_display(city: str | None, province: str | None, region: str | None) -> str | None:
        parts = [p for p in (region, province, city) if p]
        return ", ".join(parts) if parts else None

    return {
        "items": [
            {
                "incident_id": r[0],
                "verification_status": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "notification_dt": r[3].isoformat() if r[3] else None,
                "general_category": r[4],
                "alarm_level": r[5],
                "fire_station_name": r[6],
                "structures_affected": r[7],
                "households_affected": r[8],
                "individuals_affected": r[9],
                "responder_type": r[10],
                "fire_origin": r[11],
                "extent_of_damage": r[12],
                "owner_name": r[13],
                "establishment_name": r[14],
                "caller_name": r[15],
                "is_wildland": bool(r[16]),
                "updated_at": r[17].isoformat() if r[17] else None,
                "location_display": _location_display(r[18], r[19], r[20]),
            }
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


# ---------------------------------------------------------------------------
# M4-E: Dedicated Draft Management Endpoints
#
# IMPORTANT: These routes must be registered BEFORE /incidents/{incident_id}
# so that "drafts" / "draft/{id}" are not matched as the {incident_id} param.
# ---------------------------------------------------------------------------


@router.get("/incidents/drafts")
def list_encoder_drafts(
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
):
    """List the current encoder's DRAFT incidents (most-recently-updated first)."""
    encoder_id = user["user_id"]
    rows = db.execute(
        text(
            """
            SELECT
                fi.incident_id, fi.region_id, fi.created_at, fi.updated_at,
                nd.notification_dt, nd.general_category, nd.alarm_level,
                nd.fire_station_name
            FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd
                   ON nd.incident_id = fi.incident_id
            WHERE fi.encoder_id = CAST(:eid AS uuid)
              AND fi.verification_status = 'DRAFT'
              AND fi.is_archived = FALSE
            ORDER BY fi.updated_at DESC NULLS LAST, fi.created_at DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        {"eid": str(encoder_id), "limit": limit, "offset": offset},
    ).fetchall()
    total = (
        db.execute(
            text(
                """
            SELECT COUNT(*) FROM wims.fire_incidents
            WHERE encoder_id = CAST(:eid AS uuid)
              AND verification_status = 'DRAFT'
              AND is_archived = FALSE
            """
            ),
            {"eid": str(encoder_id)},
        ).scalar()
        or 0
    )
    return {
        "items": [
            {
                "incident_id": r[0],
                "region_id": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "updated_at": r[3].isoformat() if r[3] else None,
                "notification_dt": r[4].isoformat() if r[4] else None,
                "general_category": r[5],
                "alarm_level": r[6],
                "fire_station_name": r[7],
            }
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.get("/incidents/check-duplicate")
def check_incident_duplicate(
    region_id: int,
    fire_date: str,
    incident_type_code: Optional[str] = None,
    general_category: Optional[str] = None,
    user: Annotated[dict, Depends(get_regional_encoder)] = None,
    db: Annotated[Session, Depends(get_db_with_rls)] = None,
):
    """Return existing non-archived incidents that could be duplicates.

    Detection criteria (OR logic — any match triggers a warning):
      1. Same region + type_code + same calendar month + year (reference number space collision)
      2. Same region + type_code + exact fire date
      3. Same region + general_category + exact fire date (when no type_code available)
    """
    try:
        fire_dt = datetime.fromisoformat(str(fire_date))
    except (ValueError, TypeError):
        raise HTTPException(status_code=422, detail="fire_date must be a valid YYYY-MM-DD date")

    fire_month = fire_dt.month
    fire_year = fire_dt.year

    # Build WHERE conditions with explicit Python-side checks to avoid NULL pitfalls
    where_conditions = [
        "fi.region_id = :rid",
        "fi.is_archived = FALSE",
        "fi.verification_status = 'VERIFIED'",
    ]

    # Build OR sub-conditions
    or_parts: list[str] = []
    params: dict[str, Any] = {
        "rid": region_id,
        "fire_date": fire_date,
        "fire_month": fire_month,
        "fire_year": fire_year,
    }

    if incident_type_code:
        params["type_code"] = incident_type_code
        # Same reference number space (same type + same month + year)
        or_parts.append(
            "(fi.incident_type_code = :type_code"
            " AND EXTRACT(MONTH FROM nd.notification_dt AT TIME ZONE 'Asia/Manila') = :fire_month"
            " AND EXTRACT(YEAR FROM nd.notification_dt AT TIME ZONE 'Asia/Manila') = :fire_year)"
        )
        # Exact date + type (catches same day, different month edge case from above)
        or_parts.append(
            "(fi.incident_type_code = :type_code"
            " AND DATE(nd.notification_dt AT TIME ZONE 'Asia/Manila') = CAST(:fire_date AS DATE))"
        )

    if general_category:
        params["general_category"] = general_category
        # Same category + exact date (fallback when no type code)
        or_parts.append(
            "(nd.general_category = :general_category"
            " AND DATE(nd.notification_dt AT TIME ZONE 'Asia/Manila') = CAST(:fire_date AS DATE))"
        )

    if not or_parts:
        # Nothing to match on — can't run a useful check
        return {"duplicates": []}

    where_conditions.append(f"({' OR '.join(or_parts)})")
    where_sql = " AND ".join(where_conditions)

    rows = db.execute(
        text(f"""
            SELECT
                fi.incident_id,
                fi.reference_number,
                fi.verification_status,
                fi.incident_type_code,
                nd.notification_dt,
                nd.alarm_level,
                nd.general_category,
                nd.sub_category,
                nd.fire_station_name,
                nd.station_code,
                c.city_name,
                p.province_name,
                rr.region_name,
                sd.street_address
            FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd
                ON nd.incident_id = fi.incident_id
            LEFT JOIN wims.ref_cities c ON c.city_id = nd.city_id
            LEFT JOIN wims.ref_provinces p ON p.province_id = c.province_id
            LEFT JOIN wims.ref_regions rr ON rr.region_id = fi.region_id
            LEFT JOIN wims.incident_sensitive_details sd ON sd.incident_id = fi.incident_id
            WHERE {where_sql}
            ORDER BY
                fi.verification_status DESC,  -- VERIFIED first, then PENDING
                fi.created_at DESC
            LIMIT 10
        """),
        params,
    ).fetchall()

    return {
        "duplicates": [
            {
                "incident_id": r[0],
                "reference_number": r[1],
                "verification_status": r[2],
                "incident_type_code": r[3],
                "notification_dt": str(r[4]) if r[4] else None,
                "alarm_level": r[5],
                "general_category": r[6],
                "type_of_involved": r[7],
                "fire_station_name": r[8],
                "station_code": r[9],
                "city_municipality": r[10],
                "province_district": r[11],
                "region_name": r[12],
                "street_address": r[13],
            }
            for r in rows
        ]
    }


@router.get("/incidents/{incident_id}")
def get_regional_incident_detail(
    incident_id: int,
    user: Annotated[dict, Depends(get_current_wims_user)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Fetch a single incident detail. Encoders see only their own; validators see any."""
    role = user.get("role", "")
    is_validator = role in ("NATIONAL_VALIDATOR", "SYSTEM_ADMIN", "NATIONAL_ANALYST")

    if is_validator:
        row = db.execute(
            text("""
                SELECT fi.incident_id, fi.verification_status, fi.created_at,
                       fi.region_id, fi.encoder_id,
                       ST_Y(fi.location::geometry) AS latitude,
                       ST_X(fi.location::geometry) AS longitude,
                       fi.reference_number, fi.incident_type_code,
                       fi.parent_incident_id,
                       fi.is_duplicate, fi.duplicate_of, fi.updated_at
                FROM wims.fire_incidents fi
                WHERE fi.incident_id = :iid
                  AND fi.is_archived = FALSE
            """),
            {"iid": incident_id},
        ).fetchone()
    else:
        encoder_id = user["user_id"]
        row = db.execute(
            text("""
                SELECT fi.incident_id, fi.verification_status, fi.created_at,
                       fi.region_id, fi.encoder_id,
                       ST_Y(fi.location::geometry) AS latitude,
                       ST_X(fi.location::geometry) AS longitude,
                       fi.reference_number, fi.incident_type_code,
                       fi.parent_incident_id,
                       fi.is_duplicate, fi.duplicate_of, fi.updated_at
                FROM wims.fire_incidents fi
                WHERE fi.incident_id = :iid
                  AND fi.encoder_id = CAST(:encoder_id AS uuid)
                  AND fi.is_archived = FALSE
            """),
            {"iid": incident_id, "encoder_id": str(encoder_id)},
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Incident not found or access denied")

    # Fetch nonsensitive
    ns = db.execute(
        text("SELECT * FROM wims.incident_nonsensitive_details WHERE incident_id = :iid"),
        {"iid": incident_id},
    ).fetchone()

    loc_row = db.execute(
        text("""
            SELECT c.city_name, p.province_name
            FROM wims.incident_nonsensitive_details nd
            LEFT JOIN wims.ref_cities c ON c.city_id = nd.city_id
            LEFT JOIN wims.ref_provinces p ON p.province_id = c.province_id
            WHERE nd.incident_id = :iid
        """),
        {"iid": incident_id},
    ).fetchone()

    # Fetch sensitive
    sd_row = db.execute(
        text("SELECT * FROM wims.incident_sensitive_details WHERE incident_id = :iid"),
        {"iid": incident_id},
    ).fetchone()

    def row_to_dict(r, keys=None):
        if r is None:
            return {}
        if keys:
            return {k: r[i] for i, k in enumerate(keys)}
        return dict(r._mapping) if hasattr(r, "_mapping") else {}

    sd_dict = row_to_dict(sd_row)

    # ── Decrypt PII blob if present (new writes use encrypted blob; old rows fall back) ──
    if sd_dict.get("pii_blob_enc") and sd_dict.get("encryption_iv"):
        try:
            aad = f"incident_id:{incident_id}".encode("utf-8")
            pii_plaintext = _get_security_provider().decrypt_json(
                sd_dict["encryption_iv"],
                sd_dict["pii_blob_enc"],
                aad,
            )
            # Inject decrypted PII fields so frontend contract is unchanged
            sd_dict["caller_name"] = pii_plaintext.get("caller_name")
            sd_dict["caller_number"] = pii_plaintext.get("caller_number")
            sd_dict["owner_name"] = pii_plaintext.get("owner_name")
            sd_dict["occupant_name"] = pii_plaintext.get("occupant_name")
        except SecurityProviderError:
            # Auth/key failure on a blob that claims to be valid — possible tampering
            # or key rotation without re-encrypt. Log with incident_id; never log
            # nonce, ciphertext, or plaintext. Return legacy plaintext as fallback.
            logger.error(
                "CRITICAL: PII blob decryption failed (possible tamper or key mismatch). "
                "incident_id=%s",
                incident_id,
            )
            pass

    # Do not expose internal blob columns in API response
    sd_dict.pop("pii_blob_enc", None)
    sd_dict.pop("encryption_iv", None)

    nonsensitive = row_to_dict(ns)
    # Prefer the stored text columns; fall back to the ref-table JOIN for old rows
    if loc_row:
        if not nonsensitive.get("city_municipality") and loc_row[0]:
            nonsensitive["city_municipality"] = loc_row[0]
        if not nonsensitive.get("province_district") and loc_row[1]:
            nonsensitive["province_district"] = loc_row[1]
    nonsensitive["_city_text"] = nonsensitive.get("city_municipality") or ""
    nonsensitive["_province_text"] = nonsensitive.get("province_district") or ""

    # Fetch the most recent rejection reason with compatibility across IVH schemas.
    ivh_has_notes = _incident_verification_history_has_column(db, "notes")
    ivh_has_comments = _incident_verification_history_has_column(db, "comments")
    ivh_has_action_timestamp = _incident_verification_history_has_column(db, "action_timestamp")
    ivh_has_created_at = _incident_verification_history_has_column(db, "created_at")
    ivh_uses_target_columns = _incident_verification_history_uses_target_columns(db)

    rejection_reason = None
    rejection_at = None

    if (ivh_has_notes or ivh_has_comments) and (ivh_has_action_timestamp or ivh_has_created_at):
        notes_column = "notes" if ivh_has_notes else "comments"
        timestamp_column = "action_timestamp" if ivh_has_action_timestamp else "created_at"
        incident_filter = (
            "target_type = 'OFFICIAL' AND target_id = :iid"
            if ivh_uses_target_columns
            else "incident_id = :iid"
        )
        rejection_row = db.execute(
            text(f"""
                SELECT {notes_column}, {timestamp_column}
                FROM wims.incident_verification_history
                WHERE {incident_filter}
                  AND new_status = 'REJECTED'
                ORDER BY {timestamp_column} DESC
                LIMIT 1
            """),
            {"iid": incident_id},
        ).fetchone()
        rejection_reason = rejection_row[0] if rejection_row else None
        rejection_at = rejection_row[1].isoformat() if rejection_row and rejection_row[1] else None
    else:
        logger.warning(
            "IVH schema missing notes/comments or timestamp columns; skipping rejection history lookup."
        )

    # Check if incident has a wildland AFOR record (separate form from structural AFOR)
    wildland_row = db.execute(
        text("""
            SELECT wildland_fire_type, total_area_burned_hectares, total_area_burned_display
            FROM wims.incident_wildland_afor
            WHERE incident_id = :iid
        """),
        {"iid": incident_id},
    ).fetchone()
    is_wildland = wildland_row is not None
    wildland_fire_type = wildland_row[0] if wildland_row else None
    wildland_area_hectares = (
        float(wildland_row[1]) if wildland_row and wildland_row[1] is not None else None
    )
    wildland_area_display = wildland_row[2] if wildland_row else None

    return {
        "incident_id": row[0],
        "verification_status": row[1],
        "created_at": row[2].isoformat() if row[2] else None,
        "region_id": row[3],
        "latitude": float(row[5]) if row[5] is not None else None,
        "longitude": float(row[6]) if row[6] is not None else None,
        "reference_number": row[7],
        "incident_type_code": row[8],
        "parent_incident_id": row[9],
        "is_duplicate": bool(row[10]) if row[10] is not None else False,
        "duplicate_of": row[11],
        "updated_at": row[12].isoformat() if row[12] else None,
        "is_wildland": is_wildland,
        "wildland_fire_type": wildland_fire_type,
        "wildland_area_hectares": wildland_area_hectares,
        "wildland_area_display": wildland_area_display,
        "nonsensitive": nonsensitive,
        "sensitive": sd_dict,
        "rejection_reason": rejection_reason,
        "rejection_at": rejection_at,
    }


@router.get("/validator/stats")
def get_validator_stats(
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Counts of VERIFIED incidents by category visible to this validator."""
    by_cat_rows = db.execute(
        text("""
            SELECT nd.general_category, COUNT(*) as cnt
            FROM wims.fire_incidents fi
            JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE fi.verification_status = 'VERIFIED' AND fi.is_archived = FALSE
            GROUP BY nd.general_category
            ORDER BY cnt DESC
        """),
    ).fetchall()

    pending_count = (
        db.execute(
            text("""
            SELECT COUNT(*) FROM wims.fire_incidents
            WHERE verification_status = 'PENDING_VALIDATION' AND is_archived = FALSE
        """),
        ).scalar()
        or 0
    )

    total_verified = sum(r[1] for r in by_cat_rows)
    return {
        "total_verified": total_verified,
        "pending_validation": pending_count,
        "by_category": [{"category": r[0], "count": r[1]} for r in by_cat_rows],
    }


@router.get("/stats", response_model=RegionalStatsResponse)
def get_regional_stats(
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Quick summary stats scoped to the current encoder."""
    encoder_id = user["user_id"]

    total = (
        db.execute(
            text(
                "SELECT COUNT(*) FROM wims.fire_incidents WHERE encoder_id = CAST(:eid AS uuid) AND is_archived = FALSE"
            ),
            {"eid": str(encoder_id)},
        ).scalar()
        or 0
    )

    by_cat_rows = db.execute(
        text("""
            SELECT nd.general_category, COUNT(*) as cnt
            FROM wims.fire_incidents fi
            JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE fi.encoder_id = CAST(:eid AS uuid) AND fi.is_archived = FALSE
            GROUP BY nd.general_category
            ORDER BY cnt DESC
        """),
        {"eid": str(encoder_id)},
    ).fetchall()

    by_alarm_rows = db.execute(
        text("""
            SELECT nd.alarm_level, COUNT(*) as cnt
            FROM wims.fire_incidents fi
            JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE fi.encoder_id = CAST(:eid AS uuid) AND fi.is_archived = FALSE
            GROUP BY nd.alarm_level
            ORDER BY cnt DESC
        """),
        {"eid": str(encoder_id)},
    ).fetchall()

    by_status_rows = db.execute(
        text("""
            SELECT verification_status, COUNT(*) as cnt
            FROM wims.fire_incidents
            WHERE encoder_id = CAST(:eid AS uuid) AND is_archived = FALSE
            GROUP BY verification_status
            ORDER BY cnt DESC
        """),
        {"eid": str(encoder_id)},
    ).fetchall()

    # Wildland fire stats (separate AFOR form)
    wildland_total = (
        db.execute(
            text("""
                SELECT COUNT(*)
                FROM wims.incident_wildland_afor iwa
                JOIN wims.fire_incidents fi ON fi.incident_id = iwa.incident_id
                WHERE fi.encoder_id = CAST(:eid AS uuid) AND fi.is_archived = FALSE
            """),
            {"eid": str(encoder_id)},
        ).scalar()
        or 0
    )

    wildland_type_rows = db.execute(
        text("""
            SELECT iwa.wildland_fire_type, COUNT(*) as cnt
            FROM wims.incident_wildland_afor iwa
            JOIN wims.fire_incidents fi ON fi.incident_id = iwa.incident_id
            WHERE fi.encoder_id = CAST(:eid AS uuid) AND fi.is_archived = FALSE
            GROUP BY iwa.wildland_fire_type
            ORDER BY cnt DESC
        """),
        {"eid": str(encoder_id)},
    ).fetchall()

    return RegionalStatsResponse(
        total_incidents=total,
        by_category=[{"category": r[0], "count": r[1]} for r in by_cat_rows],
        by_alarm_level=[{"alarm_level": r[0], "count": r[1]} for r in by_alarm_rows],
        by_status=[{"status": r[0], "count": r[1]} for r in by_status_rows],
        wildland_total=wildland_total,
        by_wildland_type=[{"fire_type": r[0], "count": r[1]} for r in wildland_type_rows],
    )


# ---------------------------------------------------------------------------
# CRUD — Direct Incident Create / Update / Delete
# ---------------------------------------------------------------------------

_AFOR_MONTH_CODES = [
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
]


_REGION_CODE_TO_AFOR: dict[str, str] = {
    "NCR": "RGN-NCR",
    "CAR": "RGN-CAR",
    "NIR": "RGN-NIR",
    "BARMM": "RGN-BARMM",
    "I": "RGN-1",
    "II": "RGN-2",
    "III": "RGN-3",
    "IV-A": "RGN-4A",
    "IV-B": "RGN-4B",
    "V": "RGN-5",
    "VI": "RGN-6",
    "VII": "RGN-7",
    "VIII": "RGN-8",
    "IX": "RGN-9",
    "X": "RGN-10",
    "XI": "RGN-11",
    "XII": "RGN-12",
    "XIII": "RGN-13",
}


def _generate_reference_number(
    db: Session,
    region_id: int,
    incident_type_code: str,
    station_code: str,
    notification_dt: str | None,
) -> str:
    """Generate AFOR-{RGN-CODE}-{station}-{type}-{MMM}-{YYYY}-{NNNN}.

    The sequence number is globally unique across all incidents — not per-region
    or per-type — so no two incidents share the same trailing number.
    """
    region_row = db.execute(
        text("SELECT region_code FROM wims.ref_regions WHERE region_id = :rid"),
        {"rid": region_id},
    ).fetchone()
    raw_code = region_row[0] if region_row else "UNK"
    rgn_code = _REGION_CODE_TO_AFOR.get(raw_code, f"RGN-{raw_code}")

    try:
        dt = (
            datetime.fromisoformat(str(notification_dt).replace("Z", "+00:00"))
            if notification_dt
            else datetime.now()
        )
    except (ValueError, TypeError):
        dt = datetime.now()

    month = _AFOR_MONTH_CODES[dt.month - 1]
    year = dt.year
    station = (station_code or "TBA").strip() or "TBA"

    # Atomic monotonic counter — persists across archive/replace flows.
    # Fails loudly if the sequence row is missing, preventing duplicate risk.
    seq_row = db.execute(
        text("""
            UPDATE wims.reference_sequence
            SET current_value = current_value + 1
            WHERE id = 0
            RETURNING current_value
        """),
    ).fetchone()
    if not seq_row:
        raise RuntimeError(
            "reference_sequence row id=0 is missing — cannot generate reference number"
        )
    seq = int(seq_row[0])
    return f"AFOR-{rgn_code}-{station}-{incident_type_code}-{month}-{year}-{seq:04d}"


class IncidentCreateRequest(BaseModel):
    """Create a new fire incident with nonsensitive + optional sensitive details."""

    latitude: float
    longitude: float
    region_id: int | None = None
    # Nonsensitive details
    notification_dt: str | None = None
    alarm_level: str | None = None
    general_category: str | None = None
    sub_category: str | None = None
    specific_type: str | None = None
    occupancy_type: str | None = None
    city_id: int | None = None
    distance_from_station_km: float | None = None
    estimated_damage_php: float | None = None
    civilian_injured: int = 0
    civilian_deaths: int = 0
    firefighter_injured: int = 0
    firefighter_deaths: int = 0
    families_affected: int = 0
    structures_affected: int = 0
    households_affected: int = 0
    individuals_affected: int = 0
    responder_type: str | None = None
    fire_origin: str | None = None
    extent_of_damage: str | None = None
    stage_of_fire: str | None = None
    fire_station_name: str | None = None
    total_response_time_minutes: int | None = None
    recommendations: str | None = None
    # Location text (free-text, replaces city_id/province join for display)
    province_district: str | None = None
    city_municipality: str | None = None
    # Reference number fields
    station_code: str | None = "TBA"
    incident_type_code: str | None = None
    # Update-request tracking
    parent_incident_id: int | None = None
    # Sensitive details (optional — PII fields)
    street_address: str | None = None
    landmark: str | None = None
    caller_name: str | None = None
    caller_number: str | None = None
    narrative_report: str | None = None
    owner_name: str | None = None
    occupant_name: str | None = None
    establishment_name: str | None = None
    receiver_name: str | None = None
    prepared_by_officer: str | None = None
    noted_by_officer: str | None = None
    remarks: str | None = None


class IncidentUpdateRequest(BaseModel):
    """Update an existing DRAFT/PENDING incident."""

    # Nonsensitive fields
    notification_dt: str | None = None
    alarm_level: str | None = None
    general_category: str | None = None
    sub_category: str | None = None
    specific_type: str | None = None
    occupancy_type: str | None = None
    city_id: int | None = None
    distance_from_station_km: float | None = None
    estimated_damage_php: float | None = None
    civilian_injured: int | None = None
    civilian_deaths: int | None = None
    firefighter_injured: int | None = None
    firefighter_deaths: int | None = None
    families_affected: int | None = None
    structures_affected: int | None = None
    households_affected: int | None = None
    individuals_affected: int | None = None
    responder_type: str | None = None
    fire_origin: str | None = None
    extent_of_damage: str | None = None
    stage_of_fire: str | None = None
    fire_station_name: str | None = None
    total_response_time_minutes: int | None = None
    recommendations: str | None = None
    # Location text (free-text, replaces city_id/province join for display)
    province_district: str | None = None
    city_municipality: str | None = None
    # Reference number fields
    station_code: str | None = None
    incident_type_code: str | None = None
    # Sensitive fields
    street_address: str | None = None
    landmark: str | None = None
    caller_name: str | None = None
    caller_number: str | None = None
    narrative_report: str | None = None
    owner_name: str | None = None
    occupant_name: str | None = None
    establishment_name: str | None = None
    receiver_name: str | None = None
    prepared_by_officer: str | None = None
    noted_by_officer: str | None = None
    remarks: str | None = None
    # JSONB fields for full-form edit
    alarm_timeline: dict | None = None
    resources_deployed: dict | None = None
    problems_encountered: list | None = None
    other_personnel: list | None = None
    personnel_on_duty: dict | None = None
    casualty_details: dict | None = None
    disposition: str | None = None
    latitude: float | None = None
    longitude: float | None = None


@router.post("/incidents", status_code=201)
def create_incident(
    body: IncidentCreateRequest,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Create a new fire incident (DRAFT) with nonsensitive + optional sensitive details."""
    region_id = body.region_id or user.get("assigned_region_id")
    if region_id is None:
        raise HTTPException(
            status_code=400,
            detail="region_id is required when no assigned region is set",
        )
    assigned_region_id = user.get("assigned_region_id")
    if assigned_region_id and region_id != assigned_region_id:
        raise HTTPException(
            status_code=403,
            detail={
                "code": "REGION_MISMATCH",
                "message": "You can only create incidents in your assigned region.",
            },
        )
    encoder_id = user["user_id"]

    # Reference number is assigned only at validator approval — not at create time
    type_code = (body.incident_type_code or "").strip().upper() or None

    # Insert fire_incidents core row
    incident_row = db.execute(
        text("""
            INSERT INTO wims.fire_incidents
                (encoder_id, region_id, location, verification_status, incident_type_code, parent_incident_id)
            VALUES (:eid, :rid, ST_SetSRID(ST_MakePoint(:lon, :lat), 4326), 'DRAFT', :type_code, :parent_id)
            RETURNING incident_id
        """),
        {
            "eid": encoder_id,
            "rid": region_id,
            "lon": body.longitude,
            "lat": body.latitude,
            "type_code": type_code,
            "parent_id": body.parent_incident_id,
        },
    ).fetchone()
    incident_id = incident_row[0]

    # Insert nonsensitive details
    ns_fields = {
        "notification_dt",
        "alarm_level",
        "general_category",
        "sub_category",
        "specific_type",
        "occupancy_type",
        "city_id",
        "barangay_id",
        "province_district",
        "city_municipality",
        "distance_from_station_km",
        "estimated_damage_php",
        "civilian_injured",
        "civilian_deaths",
        "firefighter_injured",
        "firefighter_deaths",
        "families_affected",
        "structures_affected",
        "households_affected",
        "individuals_affected",
        "responder_type",
        "fire_origin",
        "extent_of_damage",
        "stage_of_fire",
        "fire_station_name",
        "total_response_time_minutes",
        "recommendations",
        "station_code",
    }
    ns_params = {"iid": incident_id}
    ns_cols = ["incident_id"]
    ns_vals = [":iid"]
    for field in ns_fields:
        val = getattr(body, field, None)
        if val is not None:
            if field == "alarm_level" and isinstance(val, str):
                val = ALARM_LEVEL_MAP.get(val.upper().strip(), val)
            elif field == "general_category" and isinstance(val, str):
                val = _normalize_general_category(val)
            ns_cols.append(field)
            ns_vals.append(f":{field}")
            ns_params[field] = val

    if len(ns_cols) > 1:
        db.execute(
            text(
                f"INSERT INTO wims.incident_nonsensitive_details ({', '.join(ns_cols)}) VALUES ({', '.join(ns_vals)})"
            ),
            ns_params,
        )

    # Insert sensitive details (with PII encryption if caller_name/caller_number provided)
    pii_fields = ["caller_name", "caller_number", "owner_name", "occupant_name"]
    has_pii = any(getattr(body, f, None) for f in pii_fields)

    sd_fields = {
        "street_address",
        "landmark",
        "narrative_report",
        "establishment_name",
        "receiver_name",
        "prepared_by_officer",
        "noted_by_officer",
        "remarks",
    }
    sd_params = {"iid": incident_id}
    sd_cols = ["incident_id"]
    sd_vals = [":iid"]

    if has_pii:
        pii_dict = {f: getattr(body, f) or "" for f in pii_fields}
        try:
            sp = _get_security_provider()
            nonce_b64, ct_b64 = sp.encrypt_json(pii_dict, f"incident_id:{incident_id}".encode())
            sd_cols.extend(["pii_blob_enc", "encryption_iv"])
            sd_vals.extend([":pii_blob", ":enc_iv"])
            sd_params["pii_blob"] = ct_b64
            sd_params["enc_iv"] = nonce_b64
        except SecurityProviderError:
            logger.warning(
                "PII encryption failed — storing without blob (incident_id=%s)",
                incident_id,
            )

    for field in sd_fields:
        val = getattr(body, field, None)
        if val is not None:
            sd_cols.append(field)
            sd_vals.append(f":{field}")
            sd_params[field] = val

    if len(sd_cols) > 1:
        db.execute(
            text(
                f"INSERT INTO wims.incident_sensitive_details ({', '.join(sd_cols)}) VALUES ({', '.join(sd_vals)})"
            ),
            sd_params,
        )

    _insert_incident_verification_history(
        db,
        incident_id=incident_id,
        actor_user_id=str(encoder_id),
        previous_status="DRAFT",
        new_status="DRAFT",
        notes="Encoder created new draft",
        action_label="CREATED_DRAFT",
    )

    db.commit()
    logger.info(
        "Created incident %s in region %s by encoder %s",
        incident_id,
        region_id,
        encoder_id,
    )
    return {
        "status": "created",
        "incident_id": incident_id,
        "verification_status": "DRAFT",
        "incident_type_code": type_code,
        "parent_incident_id": body.parent_incident_id,
    }


def _apply_incident_field_updates(
    db: Session, incident_id: int, body: "IncidentUpdateRequest"
) -> None:
    """Apply nonsensitive/sensitive/JSONB/coords field updates from an
    IncidentUpdateRequest to the given incident_id. Caller is responsible
    for status checks, audit-trail writes, and committing the transaction.
    """
    # Ensure child rows exist so UPDATE statements never silently affect 0 rows.
    db.execute(
        text(
            """
            INSERT INTO wims.incident_nonsensitive_details (incident_id)
            SELECT :iid
            WHERE NOT EXISTS (
                SELECT 1 FROM wims.incident_nonsensitive_details WHERE incident_id = :iid
            )
            """
        ),
        {"iid": incident_id},
    )
    db.execute(
        text(
            """
            INSERT INTO wims.incident_sensitive_details (incident_id)
            SELECT :iid
            WHERE NOT EXISTS (
                SELECT 1 FROM wims.incident_sensitive_details WHERE incident_id = :iid
            )
            """
        ),
        {"iid": incident_id},
    )

    ns_fields = {
        "notification_dt",
        "alarm_level",
        "general_category",
        "sub_category",
        "specific_type",
        "occupancy_type",
        "city_id",
        "barangay_id",
        "province_district",
        "city_municipality",
        "distance_from_station_km",
        "estimated_damage_php",
        "civilian_injured",
        "civilian_deaths",
        "firefighter_injured",
        "firefighter_deaths",
        "families_affected",
        "structures_affected",
        "households_affected",
        "individuals_affected",
        "responder_type",
        "fire_origin",
        "extent_of_damage",
        "stage_of_fire",
        "fire_station_name",
        "total_response_time_minutes",
        "recommendations",
        "station_code",
    }
    ns_updates: list[str] = []
    ns_params: dict[str, Any] = {"iid": incident_id}
    for field in ns_fields:
        val = getattr(body, field, None)
        if val is not None:
            if field == "alarm_level" and isinstance(val, str):
                val = ALARM_LEVEL_MAP.get(val.upper().strip(), val)
            elif field == "general_category" and isinstance(val, str):
                val = _normalize_general_category(val)
            ns_updates.append(f"{field} = :{field}")
            ns_params[field] = val
    if ns_updates:
        db.execute(
            text(
                f"UPDATE wims.incident_nonsensitive_details SET {', '.join(ns_updates)} WHERE incident_id = :iid"
            ),
            ns_params,
        )

    # Update incident_type_code on the fire_incidents core row if provided
    new_type_code = (getattr(body, "incident_type_code", None) or "").strip().upper() or None
    if new_type_code:
        db.execute(
            text(
                "UPDATE wims.fire_incidents SET incident_type_code = :tc WHERE incident_id = :iid"
            ),
            {"tc": new_type_code, "iid": incident_id},
        )

    sd_fields = {
        "street_address",
        "landmark",
        "narrative_report",
        "establishment_name",
        "receiver_name",
        "prepared_by_officer",
        "noted_by_officer",
        "remarks",
    }
    pii_fields = ["caller_name", "caller_number", "owner_name", "occupant_name"]
    sd_updates: list[str] = []
    sd_params: dict[str, Any] = {"iid": incident_id}
    has_pii_update = False
    for field in sd_fields | set(pii_fields):
        val = getattr(body, field, None)
        if val is not None:
            if field in pii_fields:
                has_pii_update = True
            else:
                sd_updates.append(f"{field} = :{field}")
                sd_params[field] = val
    if has_pii_update:
        existing = db.execute(
            text(
                "SELECT pii_blob_enc, encryption_iv FROM wims.incident_sensitive_details WHERE incident_id = :iid"
            ),
            {"iid": incident_id},
        ).fetchone()
        existing_pii: dict[str, Any] = {}
        if existing and existing[0] and existing[1]:
            try:
                sp = _get_security_provider()
                existing_pii = sp.decrypt_json(
                    existing[1], existing[0], f"incident_id:{incident_id}".encode()
                )
            except SecurityProviderError:
                logger.warning(
                    "Failed to decrypt existing PII for incident %s — overwriting",
                    incident_id,
                )
        for field in pii_fields:
            val = getattr(body, field, None)
            if val is not None:
                existing_pii[field] = val
        try:
            sp = _get_security_provider()
            nonce_b64, ct_b64 = sp.encrypt_json(existing_pii, f"incident_id:{incident_id}".encode())
            sd_updates.extend(["pii_blob_enc = :pii_blob", "encryption_iv = :enc_iv"])
            sd_params["pii_blob"] = ct_b64
            sd_params["enc_iv"] = nonce_b64
        except SecurityProviderError:
            logger.warning("PII re-encryption failed for incident %s", incident_id)
    if sd_updates:
        db.execute(
            text(
                f"UPDATE wims.incident_sensitive_details SET {', '.join(sd_updates)} WHERE incident_id = :iid"
            ),
            sd_params,
        )

    jsonb_ns = {
        "alarm_timeline": body.alarm_timeline,
        "resources_deployed": body.resources_deployed,
        "problems_encountered": body.problems_encountered,
    }
    jsonb_ns_updates: list[str] = []
    jsonb_ns_params: dict[str, Any] = {"iid": incident_id}
    for field, val in jsonb_ns.items():
        if val is not None:
            jsonb_ns_updates.append(f"{field} = CAST(:{field} AS jsonb)")
            jsonb_ns_params[field] = json.dumps(val)
    if jsonb_ns_updates:
        db.execute(
            text(
                f"UPDATE wims.incident_nonsensitive_details SET {', '.join(jsonb_ns_updates)} WHERE incident_id = :iid"
            ),
            jsonb_ns_params,
        )

    jsonb_sd = {
        "personnel_on_duty": body.personnel_on_duty,
        "other_personnel": body.other_personnel,
        "casualty_details": body.casualty_details,
        "disposition": body.disposition,
    }
    jsonb_sd_updates: list[str] = []
    jsonb_sd_params: dict[str, Any] = {"iid": incident_id}
    for field, val in jsonb_sd.items():
        if val is not None:
            if field == "disposition":
                jsonb_sd_updates.append(f"{field} = :{field}")
            else:
                jsonb_sd_updates.append(f"{field} = CAST(:{field} AS jsonb)")
            jsonb_sd_params[field] = json.dumps(val) if field != "disposition" else val
    if jsonb_sd_updates:
        db.execute(
            text(
                f"UPDATE wims.incident_sensitive_details SET {', '.join(jsonb_sd_updates)} WHERE incident_id = :iid"
            ),
            jsonb_sd_params,
        )

    if body.latitude is not None and body.longitude is not None:
        db.execute(
            text(
                """
                UPDATE wims.fire_incidents
                SET updated_at = now(),
                    location = ST_SetSRID(ST_MakePoint(:lon, :lat), 4326)
                WHERE incident_id = :iid
                """
            ),
            {"lon": body.longitude, "lat": body.latitude, "iid": incident_id},
        )
    else:
        db.execute(
            text("UPDATE wims.fire_incidents SET updated_at = now() WHERE incident_id = :iid"),
            {"iid": incident_id},
        )


@router.put("/incidents/{incident_id}")
def update_incident(
    incident_id: int,
    body: IncidentUpdateRequest,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Update a DRAFT or REJECTED incident owned by the current encoder.

    PENDING incidents cannot be edited directly — the encoder must withdraw
    them first (PATCH /incidents/{id}/unpend) which transitions PENDING → DRAFT.
    """
    encoder_id = user["user_id"]

    # Verify ownership + editable status
    incident = db.execute(
        text("""
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
        """),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()

    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found or not owned by you")

    if incident[1] == "PENDING":
        raise HTTPException(
            status_code=403,
            detail="This incident is PENDING review. Withdraw it first to edit.",
        )
    if incident[1] not in ("DRAFT", "REJECTED"):
        raise HTTPException(
            status_code=403,
            detail=f"Cannot edit incident with status '{incident[1]}'. Only DRAFT or REJECTED incidents can be edited.",
        )

    # Apply field updates (extracted helper — shared with /incidents/draft/{id})
    _apply_incident_field_updates(db, incident_id, body)

    # M4-B Issue #4: log every encoder edit to the audit trail
    try:
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(encoder_id),
            previous_status=incident[1],
            new_status=incident[1],
            notes="Encoder edit — fields updated",
            action_label="EDITED",
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to update incident_id=%s", incident_id)
        raise HTTPException(status_code=500, detail="Failed to save incident draft update")
    logger.info("Updated incident %s by encoder %s", incident_id, encoder_id)
    return {"status": "updated", "incident_id": incident_id}


@router.post("/incidents/{incident_id}/force-replace")
def force_replace_incident(
    incident_id: int,
    body: IncidentUpdateRequest,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Replace a PENDING incident's data without requiring withdraw first.

    Used when duplicate detection identifies a PENDING incident that the encoder
    wants to overwrite with the current form data.  The PENDING incident remains
    in PENDING status so the validator sees the updated data.  Every call is
    audited in incident_verification_history.
    """
    encoder_id = user["user_id"]

    incident = db.execute(
        text("""
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
        """),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()

    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found or not owned by you")
    if incident[1] != "PENDING":
        raise HTTPException(
            status_code=403,
            detail=f"Force-replace only applies to PENDING incidents. Current status: {incident[1]}",
        )

    _apply_incident_field_updates(db, incident_id, body)

    try:
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(encoder_id),
            previous_status="PENDING",
            new_status="PENDING",
            notes="Encoder force-replaced PENDING incident data (duplicate resolution)",
            action_label="EDITED",
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to force-replace incident_id=%s", incident_id)
        raise HTTPException(status_code=500, detail="Failed to replace incident data")

    logger.info("Force-replaced PENDING incident %s by encoder %s", incident_id, encoder_id)
    return {"status": "replaced", "incident_id": incident_id}


# ---------------------------------------------------------------------------
# M4-E: Dedicated draft management endpoints (PATCH/DELETE).
# These have 3-segment paths so they do not conflict with /incidents/{id}.
# The list endpoint (GET /incidents/drafts) is registered separately above.
# ---------------------------------------------------------------------------


@router.patch("/incidents/draft/{incident_id}")
def update_draft(
    incident_id: int,
    body: IncidentUpdateRequest,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Update a DRAFT incident owned by the current encoder.

    Mirrors update_incident() but enforces verification_status = 'DRAFT'.
    Drafts do NOT get an audit trail entry — they are not under review.
    """
    encoder_id = user["user_id"]
    incident = db.execute(
        text(
            """
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
            """
        ),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()
    if not incident:
        raise HTTPException(status_code=404, detail="Draft not found or not owned by you")
    if incident[1] != "DRAFT":
        raise HTTPException(
            status_code=403,
            detail=f"Endpoint accepts DRAFT only. Current status: {incident[1]}",
        )
    _apply_incident_field_updates(db, incident_id, body)
    try:
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(encoder_id),
            previous_status="DRAFT",
            new_status="DRAFT",
            notes="Encoder updated draft fields",
            action_label="EDITED",
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to update draft incident_id=%s", incident_id)
        raise HTTPException(status_code=500, detail="Failed to save draft")
    logger.info("Draft updated for incident %s by encoder %s", incident_id, encoder_id)
    return {"status": "draft_updated", "incident_id": incident_id}


@router.delete("/incidents/draft/{incident_id}", status_code=200)
def delete_draft(
    incident_id: int,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Soft-archive a DRAFT incident (sets is_archived = TRUE)."""
    encoder_id = user["user_id"]
    incident = db.execute(
        text(
            """
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
            """
        ),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()
    if not incident:
        raise HTTPException(status_code=404, detail="Draft not found or not owned by you")
    if incident[1] != "DRAFT":
        raise HTTPException(
            status_code=403,
            detail=f"Endpoint accepts DRAFT only. Current status: {incident[1]}",
        )
    db.execute(
        text(
            "UPDATE wims.fire_incidents SET is_archived = TRUE, updated_at = now() WHERE incident_id = :iid"
        ),
        {"iid": incident_id},
    )
    _insert_incident_verification_history(
        db,
        incident_id=incident_id,
        actor_user_id=str(encoder_id),
        previous_status="DRAFT",
        new_status="DRAFT",
        notes="Encoder deleted draft",
        action_label="DELETED_DRAFT",
    )
    db.commit()
    logger.info("Draft deleted (archived) incident %s by encoder %s", incident_id, encoder_id)
    return {"status": "deleted", "incident_id": incident_id}


@router.patch("/incidents/{incident_id}/unpend")
def unpend_incident(
    incident_id: int,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Allow encoder to withdraw a PENDING submission back to DRAFT."""
    encoder_id = user["user_id"]

    row = db.execute(
        text("""
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
        """),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Incident not found or not owned by you")

    if row[1] != "PENDING":
        raise HTTPException(status_code=400, detail=f"Incident is {row[1]}, not PENDING")

    try:
        db.execute(
            text(
                "UPDATE wims.fire_incidents SET verification_status = 'DRAFT', updated_at = now() WHERE incident_id = :iid"
            ),
            {"iid": incident_id},
        )
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(encoder_id),
            previous_status="PENDING",
            new_status="DRAFT",
            notes="Encoder withdrew incident for editing",
            action_label="WITHDRAWN",
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to unpend incident_id=%s", incident_id)
        raise HTTPException(status_code=500, detail="Failed to withdraw incident")
    logger.info("Unpended incident %s by encoder %s", incident_id, encoder_id)
    return {"status": "unpended", "incident_id": incident_id, "new_status": "DRAFT"}


@router.delete("/incidents/{incident_id}")
def delete_incident(
    incident_id: int,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Soft-delete a DRAFT incident. Sets is_archived = TRUE."""
    encoder_id = user["user_id"]

    incident = db.execute(
        text("""
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
        """),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()

    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found or not owned by you")

    if incident[1] not in ("DRAFT", "REJECTED"):
        raise HTTPException(
            status_code=403,
            detail=f"Cannot delete incident with status '{incident[1]}'. Only DRAFT or REJECTED incidents can be deleted.",
        )

    db.execute(
        text(
            "UPDATE wims.fire_incidents SET is_archived = TRUE, updated_at = now() WHERE incident_id = :iid"
        ),
        {"iid": incident_id},
    )
    _insert_incident_verification_history(
        db,
        incident_id=incident_id,
        actor_user_id=str(encoder_id),
        previous_status=incident[1],
        new_status=incident[1],
        notes="Encoder deleted incident",
        action_label="DELETED_DRAFT",
    )
    db.commit()
    logger.info("Soft-deleted incident %s by encoder %s", incident_id, encoder_id)
    return {"status": "deleted", "incident_id": incident_id}


@router.patch("/incidents/{incident_id}/submit", status_code=200)
def submit_incident_for_review(
    incident_id: int,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    ack_duplicate: bool = False,
    force: bool = False,
):
    """Submit a DRAFT or REJECTED incident for validator review (DRAFT/REJECTED → PENDING).

    Duplicate check
    ---------------
    On first call (ack_duplicate=False, force=False), if a PENDING or VERIFIED incident
    with the same region + location + fire date exists, returns HTTP 409 with
    {code: "DUPLICATE_DETECTED", matched_incident_id, matched_status} without submitting.

    The caller may:
    - Re-call with ack_duplicate=True: sets is_duplicate=TRUE + duplicate_of before PENDING.
    - Re-call with force=True: bypasses detection entirely, submits without flagging.
    """
    encoder_id = user["user_id"]

    incident = db.execute(
        text("""
            SELECT incident_id, verification_status, encoder_id
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND encoder_id = CAST(:eid AS uuid)
              AND is_archived = FALSE
        """),
        {"iid": incident_id, "eid": str(encoder_id)},
    ).fetchone()

    if not incident:
        raise HTTPException(status_code=404, detail="Incident not found or not owned by you")

    current_status = incident[1]
    inc_encoder_id = str(incident[2]) if incident[2] else None

    if inc_encoder_id != str(encoder_id):
        raise HTTPException(status_code=403, detail="You can only submit your own incidents")

    if current_status not in ("DRAFT", "REJECTED"):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot submit incident with status '{current_status}'. Only DRAFT or REJECTED incidents can be submitted.",
        )

    # Required-field gate — province/district and city/municipality must be set before PENDING
    required_check = db.execute(
        text("""
            SELECT notification_dt, general_category, province_district, city_municipality
            FROM wims.incident_nonsensitive_details
            WHERE incident_id = :iid
        """),
        {"iid": incident_id},
    ).fetchone()

    missing_fields = []
    if required_check:
        if not required_check[0]:
            missing_fields.append("notification_dt (Date of Notification)")
        if not required_check[1]:
            missing_fields.append("general_category (Classification)")
        if not required_check[2]:
            missing_fields.append("province_district (Province / District)")
        if not required_check[3]:
            missing_fields.append("city_municipality (City / Municipality)")

    if missing_fields:
        raise HTTPException(
            status_code=422,
            detail=f"Cannot submit: required fields are missing — {', '.join(missing_fields)}",
        )

    # Duplicate detection — skip when force=True or already acknowledged.
    matched_duplicate_id: int | None = None
    already_flagged = db.execute(
        text("SELECT is_duplicate FROM wims.fire_incidents WHERE incident_id = :iid"),
        {"iid": incident_id},
    ).scalar()

    if not force and not already_flagged and not ack_duplicate:
        geo_meta = db.execute(
            text("""
                SELECT nd.notification_dt, nd.general_category, fi.incident_type_code,
                       fi.region_id, nd.alarm_level,
                       ST_Y(fi.location::geometry) AS lat,
                       ST_X(fi.location::geometry) AS lon
                FROM wims.fire_incidents fi
                LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
                WHERE fi.incident_id = :iid
            """),
            {"iid": incident_id},
        ).fetchone()

        if geo_meta:
            # Use the incident's fire date when available; None skips date filter (spatial-only).
            fire_date_str: str | None = None
            if geo_meta[0]:
                notif_dt = geo_meta[0]
                fire_date_str = (
                    str(notif_dt.date()) if hasattr(notif_dt, "date") else str(notif_dt)[:10]
                )

            # Check against VERIFIED incidents
            verified_dup = check_for_duplicate(
                db,
                incident_id=incident_id,
                region_id=geo_meta[3],
                alarm_level=geo_meta[4],
                incident_date=fire_date_str,
                lat=geo_meta[5],
                lon=geo_meta[6],
                general_category=geo_meta[1],
                incident_type_code=geo_meta[2],
                exclude_statuses=("DRAFT", "REJECTED", "REPLACED"),
            )
            if verified_dup:
                matched_status = (
                    db.execute(
                        text(
                            "SELECT verification_status FROM wims.fire_incidents WHERE incident_id = :iid"
                        ),
                        {"iid": verified_dup},
                    ).scalar()
                    or "UNKNOWN"
                )
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "DUPLICATE_DETECTED",
                        "incident_id": incident_id,
                        "matched_incident_id": verified_dup,
                        "matched_status": matched_status,
                    },
                )

    try:
        # If acknowledged duplicate: flag the incident and resolve duplicate_of before submitting
        if ack_duplicate and not already_flagged:
            geo_meta = db.execute(
                text("""
                    SELECT nd.notification_dt, nd.general_category, fi.incident_type_code,
                           fi.region_id, nd.alarm_level,
                           ST_Y(fi.location::geometry) AS lat,
                           ST_X(fi.location::geometry) AS lon
                    FROM wims.fire_incidents fi
                    LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
                    WHERE fi.incident_id = :iid
                """),
                {"iid": incident_id},
            ).fetchone()
            if geo_meta:
                ack_date_str: str | None = None
                if geo_meta[0]:
                    ack_date_str = (
                        str(geo_meta[0].date())
                        if hasattr(geo_meta[0], "date")
                        else str(geo_meta[0])[:10]
                    )
                matched_duplicate_id = check_for_duplicate(
                    db,
                    incident_id=incident_id,
                    region_id=geo_meta[3],
                    alarm_level=geo_meta[4],
                    incident_date=ack_date_str,
                    lat=geo_meta[5],
                    lon=geo_meta[6],
                    general_category=geo_meta[1],
                    incident_type_code=geo_meta[2],
                    exclude_statuses=("DRAFT", "REJECTED", "REPLACED"),
                )
                if matched_duplicate_id:
                    db.execute(
                        text("""
                            UPDATE wims.fire_incidents
                            SET is_duplicate = TRUE, duplicate_of = :did
                            WHERE incident_id = :iid
                        """),
                        {"did": matched_duplicate_id, "iid": incident_id},
                    )

        update_result = db.execute(
            text(
                "UPDATE wims.fire_incidents SET verification_status = 'PENDING', updated_at = now() WHERE incident_id = :iid"
            ),
            {"iid": incident_id},
        )
        if update_result.rowcount != 1:
            raise HTTPException(status_code=409, detail="Incident status update failed")
        # M4-G: Snapshot the nonsensitive details on first PENDING transition only.
        # WHERE submitted_snapshot IS NULL ensures we never overwrite the original
        # submission state — re-submissions after rejection keep the first snapshot.
        db.execute(
            text(
                """
                UPDATE wims.fire_incidents fi
                SET submitted_snapshot = (
                    SELECT to_jsonb(nd) - 'detail_id'
                    FROM wims.incident_nonsensitive_details nd
                    WHERE nd.incident_id = fi.incident_id
                )
                WHERE fi.incident_id = :iid
                  AND fi.submitted_snapshot IS NULL
                """
            ),
            {"iid": incident_id},
        )
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(encoder_id),
            previous_status=current_status,
            new_status="PENDING",
            notes="Submitted for review",
            action_label="SUBMITTED",
        )
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        logger.exception("Failed to submit incident_id=%s for review", incident_id)
        raise HTTPException(
            status_code=500,
            detail="Failed to submit incident — transaction rolled back",
        )

    logger.info(
        "Encoder user_id=%s submitted incident_id=%s for review (%s → PENDING)",
        encoder_id,
        incident_id,
        current_status,
    )
    return {
        "status": "submitted",
        "incident_id": incident_id,
        "verification_status": "PENDING",
        "is_duplicate": ack_duplicate and matched_duplicate_id is not None,
        "duplicate_of": matched_duplicate_id,
    }


# ---------------------------------------------------------------------------
# Validator Workflow
# ---------------------------------------------------------------------------

# Allowed actions a NATIONAL_VALIDATOR can submit and their target DB status.
# accept_replace: approve a duplicate by inheriting the matched incident's ref_num and archiving it.
_VALIDATOR_ACTION_MAP: dict[str, str] = {
    "accept": "VERIFIED",
    "accept_replace": "VERIFIED",
    "pending": "PENDING",
    "reject": "REJECTED",
}

# Statuses a validator is allowed to transition an incident INTO.
_VALIDATOR_TARGET_STATUSES = frozenset(_VALIDATOR_ACTION_MAP.values())

# Statuses shown in the validator queue by default (encoder-submitted, awaiting review).
_VALIDATOR_DEFAULT_QUEUE_STATUSES = ("PENDING", "PENDING_VALIDATION")


class VerificationActionRequest(BaseModel):
    """Body for PATCH /api/regional/incidents/{incident_id}/verification."""

    action: str  # "accept" | "accept_replace" | "pending" | "reject"
    notes: str | None = None  # Optional reason / validator notes
    # When the validator chooses "Replace Existing" from the duplicate modal, pass the ID
    # of the incident to supersede. Takes priority over the stored duplicate_of value.
    original_incident_id: int | None = None


@router.get("/validator/incidents")
def get_validator_incident_queue(
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    status: Optional[str] = None,
    show_all: bool = Query(default=False),
    encoder_id: Optional[str] = None,
    archived: bool = Query(default=False),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
):
    """Validator incident queue — NATIONAL_VALIDATOR only.

    Returns encoder-submitted fire incidents across all regions.
    encoder_id IS NOT NULL is always enforced so public/DMZ submissions
    (encoder_id = NULL) are never surfaced here.

    Query params
    ------------
    status      — filter to a single verification_status value.
                  Defaults to PENDING and PENDING_VALIDATION when omitted.
    show_all    — when true and status is omitted, include all statuses
                  (DRAFT/PENDING/PENDING_VALIDATION/VERIFIED/REJECTED/REPLACED)
                  for encoder-submitted incidents across all regions.
    encoder_id  — filter to incidents submitted by one specific encoder UUID.
    archived    — when true, return only archived incidents. Default: active only.
    limit/offset — pagination.

    """
    # M4-F: NATIONAL_VALIDATOR has cross-region authority; no region gate here.
    # The role check is enforced by get_national_validator dependency.

    archive_clause = "fi.is_archived = TRUE" if archived else "fi.is_archived = FALSE"
    where_clauses = [
        archive_clause,
        "fi.encoder_id IS NOT NULL",  # encoder-submitted only — never public DMZ rows
        "fi.verification_status != 'DRAFT'",  # validators never see drafts
    ]
    params: dict[str, Any] = {
        "limit": limit,
        "offset": offset,
    }

    if status:
        where_clauses.append("fi.verification_status = :status")
        params["status"] = status
    elif not show_all and not archived:
        # Default: show the two awaiting-review statuses
        where_clauses.append("fi.verification_status = ANY(:default_statuses)")
        params["default_statuses"] = list(_VALIDATOR_DEFAULT_QUEUE_STATUSES)

    if encoder_id:
        where_clauses.append("fi.encoder_id = CAST(:encoder_id AS uuid)")
        params["encoder_id"] = encoder_id

    where_sql = " AND ".join(where_clauses)

    rows = db.execute(
        text(f"""
            SELECT
                fi.incident_id,
                fi.verification_status,
                fi.encoder_id,
                fi.region_id,
                fi.created_at,
                nd.notification_dt,
                nd.general_category,
                nd.alarm_level,
                nd.fire_station_name,
                nd.structures_affected,
                nd.households_affected,
                nd.responder_type,
                nd.fire_origin,
                nd.extent_of_damage,
                fi.parent_incident_id,
                fi.is_duplicate,
                fi.duplicate_of,
                fi.updated_at,
                fi.reference_number
            FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd
                   ON nd.incident_id = fi.incident_id
            WHERE {where_sql}
            ORDER BY fi.created_at DESC
            LIMIT :limit OFFSET :offset
        """),
        params,
    ).fetchall()

    total = (
        db.execute(
            text(f"""
                SELECT COUNT(*)
                FROM wims.fire_incidents fi
                WHERE {where_sql}
            """),
            {k: v for k, v in params.items() if k not in ("limit", "offset")},
        ).scalar()
        or 0
    )

    return {
        "items": [
            {
                "incident_id": r[0],
                "verification_status": r[1],
                "encoder_id": str(r[2]) if r[2] else None,
                "region_id": r[3],
                "created_at": r[4].isoformat() if r[4] else None,
                "submitted_at": r[4].isoformat() if r[4] else None,
                "notification_dt": r[5].isoformat() if r[5] else None,
                "general_category": r[6],
                "alarm_level": r[7],
                "fire_station_name": r[8],
                "structures_affected": r[9],
                "households_affected": r[10],
                "responder_type": r[11],
                "fire_origin": r[12],
                "extent_of_damage": r[13],
                "parent_incident_id": r[14],
                "is_duplicate": bool(r[15]) if r[15] is not None else False,
                "duplicate_of": r[16],
                "updated_at": r[17].isoformat() if r[17] else None,
                "reference_number": r[18],
            }
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.patch("/incidents/{incident_id}/verification")
def verify_incident(
    incident_id: int,
    body: VerificationActionRequest,
    request: Request,
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    force: bool = Query(default=False),
):
    """Apply a validator decision to one encoder-submitted incident.

    NATIONAL_VALIDATOR only. Enforces encoder linkage before writing.

    Allowed actions
    ---------------
    accept  → VERIFIED
    pending → PENDING
    reject  → REJECTED

    Audit trail
    -----------
    Every call inserts one row into wims.incident_verification_history in the
    same transaction as the status update — if either write fails, both roll back.

    Error responses
    ---------------
    400 — unknown action value
    403 — incident has no encoder_id (public DMZ row)
    404 — incident not found or is archived
    409 — incident already has the requested target status (idempotency guard)
    """
    validator_user_id = user["user_id"]
    # M4-F: NATIONAL_VALIDATOR has cross-region authority; no region gate.

    # --- 1. Validate action value before touching the DB ---
    action = (body.action or "").strip().lower()
    if action not in _VALIDATOR_ACTION_MAP:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Unknown action '{body.action}'. "
                f"Allowed values: {sorted(_VALIDATOR_ACTION_MAP.keys())}"
            ),
        )
    target_status = _VALIDATOR_ACTION_MAP[action]

    # --- 2. Fetch the incident (existence + archive check) ---
    incident_row = db.execute(
        text("""
            SELECT
                fi.incident_id,
                fi.verification_status,
                fi.region_id,
                fi.encoder_id,
                u.keycloak_id,
                fi.created_at
            FROM wims.fire_incidents fi
            JOIN wims.users u ON u.user_id = fi.encoder_id
            WHERE fi.incident_id = :iid AND fi.is_archived = FALSE
        """),
        {"iid": incident_id},
    ).fetchone()

    if incident_row is None:
        raise HTTPException(status_code=404, detail="Incident not found")

    inc_region_id = incident_row[2]
    inc_encoder_id = incident_row[3]
    inc_keycloak_id = incident_row[4]
    inc_created_at = incident_row[5]
    current_status = incident_row[1]

    # --- 3. Encoder linkage — reject public/DMZ rows ---
    if inc_encoder_id is None:
        raise HTTPException(
            status_code=403,
            detail="This incident was submitted via public DMZ (no encoder) and cannot be processed through the validator workflow",
        )

    # --- 4. Idempotency guard ---
    if current_status == target_status:
        raise HTTPException(
            status_code=409,
            detail=f"Incident is already in status '{current_status}'",
        )

    # --- 4a. Prevent invalid state transitions ---
    if current_status == "VERIFIED" and action == "reject":
        raise HTTPException(
            status_code=403,
            detail="Cannot reject an incident that is already verified",
        )
    if current_status == "REJECTED" and action == "accept":
        raise HTTPException(
            status_code=403,
            detail="Cannot accept an incident that has been rejected. It must be resubmitted by the encoder.",
        )
    if current_status in ("VERIFIED", "REJECTED") and action in ("accept", "reject"):
        if current_status == "VERIFIED" and action == "accept":
            pass
        elif current_status == "REJECTED" and action == "reject":
            pass
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot transition incident from '{current_status}' via action '{action}'. Only Archive is available for finalized incidents.",
            )

    # --- 4b. Duplicate check on plain accept (skipped when force=True) ---
    if action == "accept" and not force:
        geo_row = db.execute(
            text("""
                SELECT ST_Y(fi.location::geometry), ST_X(fi.location::geometry),
                       nd.notification_dt, nd.general_category, fi.incident_type_code,
                       fi.region_id, nd.alarm_level, fi.parent_incident_id
                FROM wims.fire_incidents fi
                LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
                WHERE fi.incident_id = :iid
            """),
            {"iid": incident_id},
        ).fetchone()
        if geo_row and geo_row[7] is None:
            verify_date_str: str | None = None
            if geo_row[2]:
                verify_date_str = (
                    str(geo_row[2].date()) if hasattr(geo_row[2], "date") else str(geo_row[2])[:10]
                )
            dup_id = check_for_duplicate(
                db,
                incident_id=incident_id,
                region_id=geo_row[5],
                alarm_level=geo_row[6],
                incident_date=verify_date_str,
                lat=geo_row[0],
                lon=geo_row[1],
                general_category=geo_row[3],
                incident_type_code=geo_row[4],
                exclude_statuses=("DRAFT", "REJECTED", "REPLACED"),
            )
            if dup_id:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "DUPLICATE_DETECTED",
                        "matched_incident_id": dup_id,
                    },
                )

    # --- 5. Compute verification hash + prepare reference number ---
    data_hash = None
    ref_num: str | None = None
    parent_to_archive: int | None = None
    if target_status == "VERIFIED":
        # M6-D: compute immutable SHA-256 verification hash
        canonical = {
            "encoder_id": str(inc_encoder_id),
            "keycloak_id": str(inc_keycloak_id),
            "incident_id": str(incident_id),
            "region_id": str(inc_region_id),
            "verification_status": "VERIFIED",
            "created_at": inc_created_at.isoformat(),
        }
        data_hash = hashlib.sha256(json.dumps(canonical, sort_keys=True).encode()).hexdigest()

        meta_row = db.execute(
            text("""
                SELECT fi.incident_type_code, fi.parent_incident_id, fi.duplicate_of
                FROM wims.fire_incidents fi
                WHERE fi.incident_id = :iid
            """),
            {"iid": incident_id},
        ).fetchone()
        type_code = meta_row[0] if meta_row else None
        parent_incident_id_val = meta_row[1] if meta_row else None
        duplicate_of_val = meta_row[2] if meta_row else None

        effective_original_id = body.original_incident_id or duplicate_of_val
        if action == "accept_replace" and effective_original_id:
            orig_ref_row = db.execute(
                text("SELECT reference_number FROM wims.fire_incidents WHERE incident_id = :pid"),
                {"pid": effective_original_id},
            ).fetchone()
            ref_num = orig_ref_row[0] if orig_ref_row else None
            parent_to_archive = effective_original_id
        elif parent_incident_id_val:
            orig_ref_row = db.execute(
                text("SELECT reference_number FROM wims.fire_incidents WHERE incident_id = :pid"),
                {"pid": parent_incident_id_val},
            ).fetchone()
            ref_num = orig_ref_row[0] if orig_ref_row else None
            parent_to_archive = parent_incident_id_val
        elif type_code:
            ns_meta = db.execute(
                text("""
                    SELECT notification_dt, station_code
                    FROM wims.incident_nonsensitive_details
                    WHERE incident_id = :iid
                """),
                {"iid": incident_id},
            ).fetchone()
            notification_dt = str(ns_meta[0]) if ns_meta and ns_meta[0] else None
            station_code = (ns_meta[1] if ns_meta else None) or "TBA"
            ref_num = _generate_reference_number(
                db, inc_region_id, type_code, station_code, notification_dt
            )

    # --- 6. Apply update + audit in one transaction ---
    try:
        db.execute(
            text("""
                UPDATE wims.fire_incidents
                SET verification_status = :new_status,
                    data_hash = COALESCE(:data_hash, data_hash),
                    updated_at = now()
                WHERE incident_id = :iid
            """),
            {"new_status": target_status, "iid": incident_id, "data_hash": data_hash},
        )

        if parent_to_archive:
            # Archive original first AND clear its reference_number so the unique
            # constraint is released before we assign that ref_num to the update incident.
            # Also set status to REPLACED so it appears correctly in the archive view.
            db.execute(
                text("""
                    UPDATE wims.fire_incidents
                    SET is_archived = TRUE,
                        verification_status = 'REPLACED',
                        reference_number = NULL,
                        archived_at = now(),
                        updated_at = now()
                    WHERE incident_id = :pid
                """),
                {"pid": parent_to_archive},
            )
            _insert_incident_verification_history(
                db,
                incident_id=parent_to_archive,
                actor_user_id=str(validator_user_id),
                previous_status="VERIFIED",
                new_status="REPLACED",
                notes=f"Archived — superseded by replacement incident #{incident_id}",
                action_label="REPLACED_EXISTING",
            )

        if action == "accept_replace" and effective_original_id:
            db.execute(
                text("""
                    UPDATE wims.fire_incidents
                    SET is_duplicate = FALSE,
                        duplicate_of = NULL,
                        updated_at = now()
                    WHERE incident_id = :iid
                """),
                {"iid": incident_id},
            )

        if ref_num:
            db.execute(
                text("""
                    UPDATE wims.fire_incidents
                    SET reference_number = :ref
                    WHERE incident_id = :iid
                """),
                {"ref": ref_num, "iid": incident_id},
            )

        _action_label_map = {
            "accept": "APPROVED",
            "accept_replace": "ACCEPTED_AS_NEW",
            "reject": "REJECTED",
            "pending": "RETURNED_TO_PENDING",
        }
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(validator_user_id),
            previous_status=current_status,
            new_status=target_status,
            notes=body.notes or "Validator action",
            action_label=_action_label_map.get(body.action, body.action.upper()),
        )

        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to apply verification action for incident_id=%s", incident_id)
        raise HTTPException(
            status_code=500,
            detail="Failed to apply verification action — transaction rolled back",
        )

    # analytics sync after primary commit
    sync_incident_to_analytics(db, incident_id)
    if parent_to_archive:
        sync_incident_to_analytics(db, parent_to_archive)
    db.commit()

    # non-critical audit after primary state is durable
    log_system_audit(
        db=db,
        user_id=validator_user_id,
        action_type=f"VERIFY_{action.upper()}",
        table_affected="fire_incidents",
        record_id=incident_id,
        request=request,
    )
    try:
        db.commit()
    except Exception:
        db.rollback()

    logger.info(
        "Validator user_id=%s applied action='%s' to incident_id=%s (region_id=%s, %s → %s)",
        validator_user_id,
        action,
        incident_id,
        inc_region_id,
        current_status,
        target_status,
    )

    return {
        "incident_id": incident_id,
        "previous_status": current_status,
        "new_status": target_status,
        "action": action,
        "encoder_id": str(inc_encoder_id),
        "region_id": inc_region_id,
        "reference_number": ref_num,
        "parent_archived": parent_to_archive,
    }


# ---------------------------------------------------------------------------
# M4-H: Bulk approve
# ---------------------------------------------------------------------------


class BulkApproveRequest(BaseModel):
    incident_ids: list[int]
    notes: str | None = None


@router.post("/validator/incidents/bulk-approve")
def bulk_approve_incidents(
    body: BulkApproveRequest,
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Atomically approve multiple PENDING incidents.

    All-or-nothing: if any incident is missing, archived, or not in PENDING status,
    the entire batch is rejected (422) and no incidents are modified.
    """
    validator_user_id = user["user_id"]

    if not body.incident_ids:
        raise HTTPException(status_code=400, detail="incident_ids must not be empty")

    rows = db.execute(
        text(
            """
            SELECT incident_id, verification_status, encoder_id, created_at,
                   nd.notification_dt, nd.general_category, fi2.incident_type_code,
                   fi2.region_id, nd.alarm_level,
                   ST_Y(fi2.location::geometry), ST_X(fi2.location::geometry)
            FROM wims.fire_incidents fi2
            LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi2.incident_id
            WHERE fi2.incident_id = ANY(:ids) AND fi2.is_archived = FALSE
            """
        ),
        {"ids": body.incident_ids},
    ).fetchall()

    found_ids = {r[0] for r in rows}
    missing = sorted(set(body.incident_ids) - found_ids)
    if missing:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Some incidents were not found or are archived. Transaction aborted.",
                "missing_ids": missing,
            },
        )

    not_pending = [r[0] for r in rows if r[1] != "PENDING"]
    no_encoder = [r[0] for r in rows if r[2] is None]
    if not_pending or no_encoder:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "All incidents must be PENDING and encoder-submitted. Transaction aborted.",
                "failed_ids": sorted(set(not_pending) | set(no_encoder)),
            },
        )

    # Sort by created_at ASC so oldest incidents are approved first (FIFO).
    rows = sorted(rows, key=lambda r: r[3] or datetime.min.replace(tzinfo=None))

    approved: list[int] = []
    held_for_review: list[dict] = []

    try:
        for row in rows:
            (
                iid,
                prev_status,
                _,
                created_at,
                notif_dt,
                gen_cat,
                type_code,
                region_id,
                alarm,
                lat,
                lon,
            ) = row

            # Check for duplicates including recently-VERIFIED in the last 60 seconds.
            # fire_date_str may be None — check_for_duplicate handles that with spatial-only match.
            fire_date_str = str(notif_dt.date()) if notif_dt and hasattr(notif_dt, "date") else None
            dup_id = check_for_duplicate(
                db,
                incident_id=iid,
                region_id=region_id,
                alarm_level=alarm,
                incident_date=fire_date_str,
                lat=lat,
                lon=lon,
                general_category=gen_cat,
                incident_type_code=type_code,
                exclude_statuses=("DRAFT", "REJECTED", "REPLACED"),
                verified_window_seconds=60,
            )
            if dup_id:
                held_for_review.append({"id": iid, "matching_incident_id": dup_id})
                continue

            db.execute(
                text(
                    """
                    UPDATE wims.fire_incidents
                    SET verification_status = 'VERIFIED', updated_at = now()
                    WHERE incident_id = :iid
                    """
                ),
                {"iid": iid},
            )
            _insert_incident_verification_history(
                db,
                incident_id=iid,
                actor_user_id=str(validator_user_id),
                previous_status=prev_status,
                new_status="VERIFIED",
                notes=body.notes or "Bulk approve",
                action_label="BULK_APPROVED",
            )
            approved.append(iid)

        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Bulk approve failed")
        raise HTTPException(status_code=500, detail="Bulk approve failed — transaction rolled back")

    logger.info(
        "Validator user_id=%s bulk-approved %d incidents: %s; held: %d",
        validator_user_id,
        len(approved),
        sorted(approved),
        len(held_for_review),
    )
    return {
        "approved": len(approved),
        "incident_ids": sorted(approved),
        "held_for_review": held_for_review,
    }


# ---------------------------------------------------------------------------
# B4: Archive endpoint for validators
# ---------------------------------------------------------------------------


@router.patch("/validator/incidents/{incident_id}/archive")
def archive_incident(
    incident_id: int,
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Archive a finalized (VERIFIED, REJECTED, or REPLACED) incident.

    Sets is_archived=TRUE, archived_at=NOW(), verification_status unchanged.
    Returns 400 if the incident is in DRAFT or PENDING status.
    """
    validator_user_id = user["user_id"]

    incident = db.execute(
        text("""
            SELECT incident_id, verification_status
            FROM wims.fire_incidents
            WHERE incident_id = :iid
              AND is_archived = FALSE
        """),
        {"iid": incident_id},
    ).fetchone()

    if incident is None:
        raise HTTPException(status_code=404, detail="Incident not found or already archived")

    current_status = incident[1]
    archivable_statuses = ("VERIFIED", "REJECTED", "REPLACED")
    if current_status not in archivable_statuses:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Only {', '.join(archivable_statuses)} incidents can be archived. "
                f"Current status: '{current_status}'."
            ),
        )

    try:
        db.execute(
            text("""
                UPDATE wims.fire_incidents
                SET is_archived = TRUE,
                    archived_at  = now(),
                    updated_at   = now()
                WHERE incident_id = :iid
            """),
            {"iid": incident_id},
        )
        _insert_incident_verification_history(
            db,
            incident_id=incident_id,
            actor_user_id=str(validator_user_id),
            previous_status=current_status,
            new_status="ARCHIVED",
            notes="Archived by validator",
            action_label="ARCHIVED",
        )
        db.commit()
    except Exception:
        db.rollback()
        logger.exception("Failed to archive incident_id=%s", incident_id)
        raise HTTPException(status_code=500, detail="Archive failed — transaction rolled back")

    return {"status": "archived", "incident_id": incident_id}


# ---------------------------------------------------------------------------
# M4-G: Side-by-side diff for validators
# ---------------------------------------------------------------------------


# Field keys included in the diff. PII fields from incident_sensitive_details
# are intentionally excluded — only nonsensitive operational details are diffed.
_DIFF_FIELDS = (
    "notification_dt",
    "alarm_level",
    "general_category",
    "sub_category",
    "specific_type",
    "occupancy_type",
    "city_id",
    "barangay_id",
    "distance_from_station_km",
    "estimated_damage_php",
    "civilian_injured",
    "civilian_deaths",
    "firefighter_injured",
    "firefighter_deaths",
    "families_affected",
    "structures_affected",
    "households_affected",
    "individuals_affected",
    "responder_type",
    "fire_origin",
    "extent_of_damage",
    "stage_of_fire",
    "fire_station_name",
    "total_response_time_minutes",
    "recommendations",
    "vehicles_affected",
    "extent_total_floor_area_sqm",
    "extent_total_land_area_hectares",
    "alarm_timeline",
    "resources_deployed",
    "problems_encountered",
)


@router.get("/validator/incidents/{incident_id}/diff")
def get_incident_diff(
    incident_id: int,
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
):
    """Return the original-vs-current diff for an incident's nonsensitive fields.

    Original = wims.fire_incidents.submitted_snapshot (JSONB written on first
    DRAFT/REJECTED → PENDING transition).
    Current  = wims.incident_nonsensitive_details (live row).
    """
    incident_row = db.execute(
        text(
            """
            SELECT incident_id, submitted_snapshot
            FROM wims.fire_incidents
            WHERE incident_id = :iid AND is_archived = FALSE
            """
        ),
        {"iid": incident_id},
    ).fetchone()
    if incident_row is None:
        raise HTTPException(status_code=404, detail="Incident not found")

    snapshot: dict[str, Any] | None = incident_row[1]

    current_row = db.execute(
        text(
            """
            SELECT to_jsonb(nd) - 'detail_id' AS doc
            FROM wims.incident_nonsensitive_details nd
            WHERE nd.incident_id = :iid
            """
        ),
        {"iid": incident_id},
    ).fetchone()
    current: dict[str, Any] = current_row[0] if current_row and current_row[0] else {}

    if snapshot is None:
        return {
            "original": None,
            "current": {k: current.get(k) for k in _DIFF_FIELDS if k in current},
            "changed_fields": [],
            "note": "No snapshot available — incident submitted before diff tracking was enabled.",
        }

    original_subset: dict[str, Any] = {k: snapshot.get(k) for k in _DIFF_FIELDS if k in snapshot}
    current_subset: dict[str, Any] = {k: current.get(k) for k in _DIFF_FIELDS if k in current}
    all_keys = set(original_subset.keys()) | set(current_subset.keys())
    changed_fields = sorted(k for k in all_keys if original_subset.get(k) != current_subset.get(k))

    return {
        "original": original_subset,
        "current": current_subset,
        "changed_fields": changed_fields,
    }


# ---------------------------------------------------------------------------
# M4-I: Validator audit trail viewer (incident_verification_history)
# ---------------------------------------------------------------------------


@router.get("/audit-log")
def get_encoder_audit_log(
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
):
    """Return the current encoder's own action history from incident_verification_history."""
    encoder_id = str(user["user_id"])
    where_clauses = [
        "ivh.target_type = 'OFFICIAL'",
        "ivh.action_by_user_id = CAST(:encoder_id AS uuid)",
    ]
    params: dict[str, Any] = {"encoder_id": encoder_id}
    if date_from:
        where_clauses.append("ivh.action_timestamp >= CAST(:date_from AS timestamptz)")
        params["date_from"] = date_from
    if date_to:
        where_clauses.append("ivh.action_timestamp <= CAST(:date_to AS timestamptz)")
        params["date_to"] = date_to
    where_sql = " AND ".join(where_clauses)

    rows = db.execute(
        text(
            f"""
            SELECT
                ivh.history_id, ivh.target_id,
                ivh.action_label, ivh.previous_status, ivh.new_status,
                ivh.notes, ivh.action_timestamp
            FROM wims.incident_verification_history ivh
            WHERE {where_sql}
            ORDER BY ivh.action_timestamp DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        {**params, "limit": limit, "offset": offset},
    ).fetchall()

    total = (
        db.execute(
            text(f"SELECT COUNT(*) FROM wims.incident_verification_history ivh WHERE {where_sql}"),
            params,
        ).scalar()
        or 0
    )

    return {
        "items": [
            {
                "history_id": r[0],
                "incident_id": r[1],
                "action_label": r[2],
                "previous_status": r[3],
                "new_status": r[4],
                "notes": r[5],
                "action_timestamp": r[6].isoformat() if r[6] else None,
            }
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


def _build_audit_log_query(
    *,
    date_from: str | None,
    date_to: str | None,
    region_id: int | None,
    validator_id: str | None,
    action: str | None,
) -> tuple[str, dict[str, Any]]:
    """Compose a parameterized WHERE clause for audit log queries.

    Returns (where_sql, params). The caller plugs where_sql into a SELECT.
    """
    where_clauses = ["ivh.target_type = 'OFFICIAL'"]
    params: dict[str, Any] = {}
    if date_from:
        where_clauses.append("ivh.action_timestamp >= CAST(:date_from AS timestamptz)")
        params["date_from"] = date_from
    if date_to:
        where_clauses.append("ivh.action_timestamp <= CAST(:date_to AS timestamptz)")
        params["date_to"] = date_to
    if region_id is not None:
        where_clauses.append("fi.region_id = :region_id")
        params["region_id"] = region_id
    if validator_id:
        where_clauses.append("ivh.action_by_user_id = CAST(:validator_id AS uuid)")
        params["validator_id"] = validator_id
    if action:
        where_clauses.append("ivh.action_label = :action")
        params["action"] = action
    return " AND ".join(where_clauses), params


@router.get("/validator/audit-logs")
def get_validator_audit_logs(
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    region_id: Optional[int] = None,
    validator_id: Optional[str] = None,
    action: Optional[str] = None,  # filter by action_label (APPROVED/REJECTED/BULK_APPROVED/etc.)
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
):
    """Paginated audit-log query over wims.incident_verification_history."""
    where_sql, params = _build_audit_log_query(
        date_from=date_from,
        date_to=date_to,
        region_id=region_id,
        validator_id=validator_id,
        action=action,
    )
    list_params = {**params, "limit": limit, "offset": offset}

    rows = db.execute(
        text(
            f"""
            SELECT
                ivh.history_id, ivh.target_id, fi.region_id,
                ivh.action_by_user_id, ivh.previous_status, ivh.new_status,
                ivh.notes, ivh.action_timestamp,
                u.username AS actor_username,
                rr.region_name AS region_display,
                ivh.action_label
            FROM wims.incident_verification_history ivh
            JOIN wims.fire_incidents fi ON fi.incident_id = ivh.target_id
            LEFT JOIN wims.users u ON u.user_id = ivh.action_by_user_id
            LEFT JOIN wims.ref_regions rr ON rr.region_id = fi.region_id
            WHERE {where_sql}
            ORDER BY ivh.action_timestamp DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        list_params,
    ).fetchall()

    total = (
        db.execute(
            text(
                f"""
            SELECT COUNT(*)
            FROM wims.incident_verification_history ivh
            JOIN wims.fire_incidents fi ON fi.incident_id = ivh.target_id
            WHERE {where_sql}
            """
            ),
            params,
        ).scalar()
        or 0
    )

    return {
        "items": [
            {
                "history_id": r[0],
                "incident_id": r[1],
                "region_id": r[2],
                "action_by_user_id": str(r[3]) if r[3] else None,
                "previous_status": r[4],
                "new_status": r[5],
                "notes": r[6],
                "action_timestamp": r[7].isoformat() if r[7] else None,
                "actor_username": r[8],
                "region_display": r[9],
                "action_label": r[10],
            }
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.get("/validator/audit-logs/export")
def export_validator_audit_logs(
    user: Annotated[dict, Depends(get_national_validator)],
    db: Annotated[Session, Depends(get_db_with_rls)],
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    region_id: Optional[int] = None,
    validator_id: Optional[str] = None,
    action: Optional[str] = None,
):
    """Return an audit-log CSV. Honors the same filters as the list endpoint."""
    where_sql, params = _build_audit_log_query(
        date_from=date_from,
        date_to=date_to,
        region_id=region_id,
        validator_id=validator_id,
        action=action,
    )

    rows = db.execute(
        text(
            f"""
            SELECT
                ivh.history_id, ivh.target_id, fi.region_id,
                ivh.action_by_user_id, ivh.previous_status, ivh.new_status,
                ivh.notes, ivh.action_timestamp,
                u.username AS actor_username,
                rr.region_name AS region_display,
                ivh.action_label
            FROM wims.incident_verification_history ivh
            JOIN wims.fire_incidents fi ON fi.incident_id = ivh.target_id
            LEFT JOIN wims.users u ON u.user_id = ivh.action_by_user_id
            LEFT JOIN wims.ref_regions rr ON rr.region_id = fi.region_id
            WHERE {where_sql}
            ORDER BY ivh.action_timestamp DESC
            """
        ),
        params,
    ).fetchall()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "history_id",
            "incident_id",
            "region_id",
            "region_display",
            "action_by_user_id",
            "actor_username",
            "previous_status",
            "new_status",
            "action_label",
            "notes",
            "action_timestamp",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r[0],
                r[1],
                r[2],
                r[9] or "",
                str(r[3]) if r[3] else "",
                r[8] or "",
                r[4],
                r[5],
                r[10] or "",
                (r[6] or "").replace("\n", " "),
                r[7].isoformat() if r[7] else "",
            ]
        )

    export_date = datetime.utcnow().strftime("%Y%m%d")
    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=audit-log-{export_date}.csv",
        },
    )
