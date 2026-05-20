"""Celery tasks for analytics exports."""

from __future__ import annotations

import csv
import json
import logging
import os
import uuid
from typing import Any, Callable

from sqlalchemy import text
from sqlalchemy.orm import Session

from celery_config import celery_app
from database import get_session, set_rls_context
from services.analytics_read_model import (
    get_analyst_export_rows,
    get_export_rows,
    get_incident_export_data,
)

logger = logging.getLogger(__name__)

ALLOWED_EXPORT_COLUMNS = {
    "incident_id",
    "notification_dt",
    "alarm_level",
    "general_category",
    "sub_category",
    "fire_origin",
    "extent_of_damage",
    "structures_affected",
    "households_affected",
    "individuals_affected",
    "vehicles_affected",
    "total_response_time_minutes",
    "total_gas_consumed_liters",
    "extent_total_floor_area_sqm",
    "extent_total_land_area_hectares",
    "civilian_injured",
    "civilian_deaths",
    "firefighter_injured",
    "firefighter_deaths",
    "fire_station_name",
    "region_id",
    "verification_status",
    "estimated_damage_php",
    "municipality_name",
    "province_name",
}

DEFAULT_EXPORT_COLUMNS = [
    "incident_id",
    "notification_dt",
    "region_id",
    "province_name",
    "municipality_name",
    "alarm_level",
    "general_category",
    "sub_category",
    "estimated_damage_php",
    "total_response_time_minutes",
]

EXPORT_DIR = os.environ.get("EXPORT_DIR", "/tmp/wims-exports")
AFOR_TEMPLATE_PATH = os.environ.get(
    "AFOR_TEMPLATE_PATH",
    "/app/AFOR-FORMATTED.xlsx",
)


def _serialize_value(v: Any) -> str:
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return str(v.isoformat())
    return str(v)


def _valid_columns(columns: list[str]) -> list[str]:
    valid_cols = [c for c in columns if c in ALLOWED_EXPORT_COLUMNS]
    return valid_cols or DEFAULT_EXPORT_COLUMNS


def _write_csv(path: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: _serialize_value(row.get(col)) for col in columns})


def _write_xlsx(path: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Incidents"
    worksheet.append(columns)
    for row in rows:
        worksheet.append([_serialize_value(row.get(col)) for col in columns])
    workbook.save(path)


# ─── AFOR Template Writers ────────────────────────────────────────────────────

# AFOR cell address map: field_key → (row, col_letter)
# col letters: A=1, B=2, C=3, D=4, E=5, F=6
AFOR_CELL_MAP: dict[str, tuple[int, int]] = {
    # Section A: Response Details
    "region": (10, 1),  # A10 "(INSERT REGION)"
    "address": (11, 1),  # A11 "(INSERT ADDRESS)"
    "contact": (12, 1),  # A12
    "date_notification": (22, 4),  # D22
    "time_notification": (23, 4),  # D23
    "region_name": (24, 4),  # D24
    "province": (25, 4),  # D25
    "municipality": (26, 4),  # D26
    "address_fire": (27, 4),  # D27
    "landmark": (28, 4),  # D28
    "caller_name": (29, 4),  # D29
    "personnel_received": (30, 4),  # D30
    "engine_dispatched": (31, 4),  # D31
    "time_dispatched": (34, 4),  # D34
    "time_arrived": (37, 4),  # D37
    "response_time_min": (40, 4),  # D40
    "distance_km": (41, 4),  # D41
    "highest_alarm": (42, 4),  # D42
    "time_returned_base": (43, 4),  # D43
    "gas_consumed_liters": (44, 4),  # D44
    # Section B: Nature & Classification
    "general_category": (47, 4),  # D47
    "sub_category": (47, 5),  # E47
    "owner_name": (51, 4),  # D51
    "establishment_name": (51, 5),  # E51
    "general_description": (52, 4),  # D52
    "area_of_origin": (53, 4),  # D53
    "stage_of_fire": (54, 4),  # D54
    "extent_of_damage": (55, 4),  # D55
    "structures_affected": (62, 4),  # D62
    "households_affected": (63, 4),  # D63
    "families_affected": (64, 4),  # D64
    "individuals_affected": (65, 4),  # D65
    "vehicles_affected": (66, 4),  # D66
    # Section C: Assets & Resources
    "bfp_fire_trucks": (70, 4),  # D70
    "bfp_manned_trucks": (71, 4),  # D71
    "non_bfp_trucks": (72, 4),  # D72
    "bfp_ambulance": (73, 4),  # D73
    "non_bfp_ambulance": (74, 4),  # D74
    "bfp_rescue_trucks": (75, 4),  # D75
    "non_bfp_rescue_trucks": (76, 4),  # D76
    "other_vehicles": (77, 4),  # D77
    "scba_used": (79, 4),  # D79
    "rope_used": (80, 4),  # D80
    "ladder_used": (81, 4),  # D81
    "hoseline_used": (82, 4),  # D82
    "hydraulic_tools": (83, 4),  # D83
    "other_tools": (84, 4),  # D84
    "hydrant_location": (85, 4),  # D85
    # Section D: Fire Alarm Level
    "alarm_1st": (89, 4),  # D89  1ST ALARM
    "alarm_2nd": (90, 4),  # D90  2ND ALARM
    "alarm_3rd": (91, 4),  # D91  3RD ALARM
    "alarm_4th": (92, 4),  # D92  4TH ALARM
    "alarm_5th": (93, 4),  # D93  5TH ALARM
    "alarm_tf_alpha": (94, 4),  # D94  TASK FORCE ALPHA
    "alarm_tf_bravo": (95, 4),  # D95  TASK FORCE BRAVO
    "alarm_tf_charlie": (96, 4),  # D96  TASK FORCE CHARLIE
    "alarm_tf_delta": (97, 4),  # D97  TASK FORCE DELTA
    "alarm_general": (98, 4),  # D98  GENERAL ALARM
    "alarm_fuc": (99, 4),  # D99  FIRE UNDER CONTROL
    "alarm_fo": (100, 4),  # D100 FIRE OUT
    # Section D: ICP
    "icp_status": (101, 4),  # D101
    "icp_location": (102, 4),  # D102
    # Section E: Casualties
    "civilian_injured_male": (106, 4),  # D106
    "civilian_injured_female": (106, 5),  # E106
    "firefighter_injured_male": (107, 4),  # D107
    "firefighter_injured_female": (107, 5),  # E107
    "civilian_deaths_male": (109, 4),  # D109
    "civilian_deaths_female": (109, 5),  # E109
    "firefighter_deaths_male": (110, 4),  # D110
    "firefighter_deaths_female": (110, 5),  # E110
    # Section F: Personnel
    "engine_commander": (114, 4),  # D114
    "shift_in_charge": (115, 4),  # D115
    "nozzleman": (116, 4),  # D116
    "lineman": (117, 4),  # D117
    "engine_crew": (118, 4),  # D118
    "driver_dpo": (119, 4),  # D119
    "safety_officer": (120, 4),  # D120
    "investigator": (121, 4),  # D121
    # Prepared by / Noted by
    "prepared_by_rank": (239, 2),  # B239
    "prepared_by_name": (240, 2),  # B240
    "noted_by_rank": (239, 5),  # E239
    "noted_by_name": (240, 5),  # E240
    # Additional numeric fields
    "extent_floor_area": (58, 4),  # D58 (Confined to Room floor area)
    "extent_land_area": (59, 4),  # D59
    "water_tankers": (85, 4),  # D85
    "responder_type": (19, 4),  # D19 (Type of Responder)
    "estimated_damage": (42, 4),  # D42 alternate
}


def _write_afor_excel(path: str, data: dict[str, Any]) -> None:
    """Fill the AFOR XLSX template with incident data and save as .xlsx."""
    import openpyxl

    template = AFOR_TEMPLATE_PATH if os.path.isfile(AFOR_TEMPLATE_PATH) else None
    if template:
        wb = openpyxl.load_workbook(template)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AFOR"
        for r in range(1, 250):
            for c in range(1, 7):
                ws.cell(r, c).value = None

    ws = wb["AFOR"]

    for field, (row, col) in AFOR_CELL_MAP.items():
        value = data.get(field, "")
        if value is None:
            value = ""
        ws.cell(row, col, str(value))

    wb.save(path)


def _write_afor_pdf(path: str, data: dict[str, Any]) -> None:
    """
    Produce a PDF that visually renders the AFOR form layout.
    Uses openpyxl to fill the template, saves as xlsx, then converts
    to PDF using a two-step approach: html page → PDF via reportlab.
    """
    xlsx_path = path.replace(".pdf", ".xlsx")
    _write_afor_excel(xlsx_path, data)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    styles = getSampleStyleSheet()
    maroon = colors.HexColor("#7f1d1d")
    light_gray = colors.HexColor("#f3f4f6")

    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(letter),
        title=f"AFOR – Incident {data.get('incident_id', 'N/A')}",
    )

    def section_header(text: str) -> Paragraph:
        return Paragraph(
            f'<font color="white"><b>{text}</b></font>',
            ParagraphStyle(
                "SectionHeader",
                parent=styles["Normal"],
                backColor=maroon,
                textColor=colors.white,
                fontName="Helvetica-Bold",
                fontSize=9,
                leading=14,
                leftIndent=4,
                spaceAfter=4,
            ),
        )

    def label(text: str) -> Paragraph:
        return Paragraph(
            f'<font name="Helvetica-Bold" size="8">{text}</font>',
            styles["Normal"],
        )

    def value(text: str) -> Paragraph:
        return Paragraph(str(text) if text else "—", styles["Normal"])

    def kv(label_txt: str, val_txt: str) -> Table:
        return Table(
            [[label(label_txt), value(val_txt)]],
            colWidths=[2.2 * inch, 4.0 * inch],
            style=TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (0, 0), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                ]
            ),
        )

    story = []
    story.append(Paragraph("AFTER FIRE OPERATIONS REPORT (AFOR)", styles["Title"]))
    story.append(Spacer(1, 8))

    incident_id = data.get("incident_id", "")
    story.append(
        Paragraph(
            f"<b>Incident ID:</b> {incident_id}  |  <b>Reference:</b> {data.get('reference_number', 'N/A')}  |  <b>Date:</b> {data.get('date_notification', 'N/A')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 10))

    # ── Section A: Response Details ──────────────────────────────────────────────
    story.append(section_header("A. RESPONSE DETAILS"))
    story.append(Spacer(1, 4))
    a_data = [
        ["Region", data.get("region_name", ""), "Province", data.get("province", "")],
        [
            "Municipality",
            data.get("municipality", ""),
            "Fire Origin / Address",
            data.get("address_fire", ""),
        ],
        [
            "Date Notified",
            data.get("date_notification", ""),
            "Time Notified",
            data.get("time_notification", ""),
        ],
        [
            "Engine Dispatched",
            data.get("engine_dispatched", ""),
            "Time Dispatched",
            data.get("time_dispatched", ""),
        ],
        [
            "Time Arrived",
            data.get("time_arrived", ""),
            "Response Time (min)",
            data.get("response_time_min", ""),
        ],
        [
            "Distance (km)",
            data.get("distance_km", ""),
            "Highest Alarm",
            data.get("highest_alarm", ""),
        ],
        [
            "Gas Consumed (L)",
            data.get("gas_consumed_liters", ""),
            "Responder Type",
            data.get("responder_type", ""),
        ],
    ]
    a_table = Table(a_data, colWidths=[1.6 * inch, 2.2 * inch, 1.6 * inch, 2.2 * inch])
    a_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), light_gray),
                ("BACKGROUND", (2, 0), (2, -1), light_gray),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(a_table)
    story.append(Spacer(1, 10))

    # ── Section B: Nature & Classification ───────────────────────────────────────
    story.append(section_header("B. NATURE AND CLASSIFICATION OF INVOLVED"))
    story.append(Spacer(1, 4))
    b_data = [
        [
            "Classification",
            data.get("general_category", ""),
            "Sub-Category",
            data.get("sub_category", ""),
        ],
        [
            "Area of Origin",
            data.get("area_of_origin", ""),
            "Stage of Fire",
            data.get("stage_of_fire", ""),
        ],
        [
            "Extent of Damage",
            data.get("extent_of_damage", ""),
            "Est. Damage (PHP)",
            data.get("estimated_damage", ""),
        ],
        [
            "Structures Affected",
            data.get("structures_affected", ""),
            "Households Affected",
            data.get("households_affected", ""),
        ],
        [
            "Families Affected",
            data.get("families_affected", ""),
            "Individuals Affected",
            data.get("individuals_affected", ""),
        ],
        ["Vehicles Affected", data.get("vehicles_affected", ""), "", ""],
        [
            "Floor Area (sqm)",
            data.get("extent_floor_area", ""),
            "Land Area (ha)",
            data.get("extent_land_area", ""),
        ],
    ]
    b_table = Table(b_data, colWidths=[1.6 * inch, 2.2 * inch, 1.6 * inch, 2.2 * inch])
    b_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), light_gray),
                ("BACKGROUND", (2, 0), (2, -1), light_gray),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(b_table)
    story.append(Spacer(1, 10))

    # ── Section C: Assets & Resources ────────────────────────────────────────────
    story.append(section_header("C. ASSETS AND RESOURCES"))
    story.append(Spacer(1, 4))
    c_data = [
        [
            "BFP Fire Trucks",
            data.get("bfp_fire_trucks", ""),
            "BFP Manned Trucks",
            data.get("bfp_manned_trucks", ""),
        ],
        [
            "Non-BFP Fire Trucks",
            data.get("non_bfp_trucks", ""),
            "BFP Ambulance",
            data.get("bfp_ambulance", ""),
        ],
        [
            "Non-BFP Ambulance",
            data.get("non_bfp_ambulance", ""),
            "BFP Rescue Trucks",
            data.get("bfp_rescue_trucks", ""),
        ],
        [
            "Non-BFP Rescue Trucks",
            data.get("non_bfp_rescue_trucks", ""),
            "Other Vehicles",
            data.get("other_vehicles", ""),
        ],
        ["SCBA Used", data.get("scba_used", ""), "Water Tankers", data.get("water_tankers", "")],
    ]
    c_table = Table(c_data, colWidths=[1.6 * inch, 2.2 * inch, 1.6 * inch, 2.2 * inch])
    c_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), light_gray),
                ("BACKGROUND", (2, 0), (2, -1), light_gray),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(c_table)
    story.append(Spacer(1, 10))

    # ── Section D: Fire Alarm Level ─────────────────────────────────────────────
    story.append(section_header("D. FIRE ALARM LEVEL"))
    story.append(Spacer(1, 4))
    d_data = [
        ["1ST ALARM", data.get("alarm_1st", "")],
        ["2ND ALARM", data.get("alarm_2nd", "")],
        ["3RD ALARM", data.get("alarm_3rd", "")],
        ["4TH ALARM", data.get("alarm_4th", "")],
        ["5TH ALARM", data.get("alarm_5th", "")],
        ["TASK FORCE ALPHA", data.get("alarm_tf_alpha", "")],
        ["TASK FORCE BRAVO", data.get("alarm_tf_bravo", "")],
        ["TASK FORCE CHARLIE", data.get("alarm_tf_charlie", "")],
        ["TASK FORCE DELTA", data.get("alarm_tf_delta", "")],
        ["GENERAL ALARM", data.get("alarm_general", "")],
        ["FIRE UNDER CONTROL", data.get("alarm_fuc", "")],
        ["FIRE OUT", data.get("alarm_fo", "")],
    ]
    d_table = Table(d_data, colWidths=[2.0 * inch, 4.5 * inch])
    d_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), light_gray),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(d_table)
    story.append(Spacer(1, 10))

    # ── Section E: Casualties ─────────────────────────────────────────────────────
    story.append(section_header("E. PROFILE OF CASUALTIES"))
    story.append(Spacer(1, 4))
    e_data = [
        ["", "Male", "Female"],
        [
            "Injured Civilian",
            data.get("civilian_injured_male", ""),
            data.get("civilian_injured_female", ""),
        ],
        [
            "Injured BFP Firefighter",
            data.get("firefighter_injured_male", ""),
            data.get("firefighter_injured_female", ""),
        ],
        [
            "Civilian Fatality",
            data.get("civilian_deaths_male", ""),
            data.get("civilian_deaths_female", ""),
        ],
        [
            "BFP Firefighter Fatality",
            data.get("firefighter_deaths_male", ""),
            data.get("firefighter_deaths_female", ""),
        ],
    ]
    e_table = Table(e_data, colWidths=[2.4 * inch, 2.0 * inch, 2.0 * inch])
    e_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), maroon),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (0, -1), light_gray),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(e_table)

    doc.build(story)


def _write_afor_csv(path: str, data: dict[str, Any]) -> None:
    """
    Write incident data as a section-based CSV matching AFOR layout.
    Each section is a block of labelled key-value rows.
    """
    sections = [
        (
            "A. RESPONSE DETAILS",
            [
                ("Region", data.get("region_name", "")),
                ("Province", data.get("province", "")),
                ("Municipality", data.get("municipality", "")),
                ("Date Notification Received", data.get("date_notification", "")),
                ("Time Notification Received", data.get("time_notification", "")),
                ("Engine Dispatched", data.get("engine_dispatched", "")),
                ("Time Dispatched", data.get("time_dispatched", "")),
                ("Time Arrived at Scene", data.get("time_arrived", "")),
                ("Response Time (min)", data.get("response_time_min", "")),
                ("Distance (km)", data.get("distance_km", "")),
                ("Highest Alarm Level", data.get("highest_alarm", "")),
                ("Time Returned to Base", data.get("time_returned_base", "")),
                ("Total Gas Consumed (L)", data.get("gas_consumed_liters", "")),
                ("Responder Type", data.get("responder_type", "")),
            ],
        ),
        (
            "B. NATURE AND CLASSIFICATION OF INVOLVED",
            [
                ("General Category", data.get("general_category", "")),
                ("Sub-Category", data.get("sub_category", "")),
                ("Area of Origin", data.get("area_of_origin", "")),
                ("Stage of Fire", data.get("stage_of_fire", "")),
                ("Extent of Damage", data.get("extent_of_damage", "")),
                ("Est. Damage (PHP)", data.get("estimated_damage", "")),
                ("Structures Affected", data.get("structures_affected", "")),
                ("Households Affected", data.get("households_affected", "")),
                ("Families Affected", data.get("families_affected", "")),
                ("Individuals Affected", data.get("individuals_affected", "")),
                ("Vehicles Affected", data.get("vehicles_affected", "")),
                ("Floor Area (sqm)", data.get("extent_floor_area", "")),
                ("Land Area (ha)", data.get("extent_land_area", "")),
            ],
        ),
        (
            "C. ASSETS AND RESOURCES",
            [
                ("BFP Fire Trucks", data.get("bfp_fire_trucks", "")),
                ("BFP Manned Fire Trucks", data.get("bfp_manned_trucks", "")),
                ("Non-BFP Fire Trucks", data.get("non_bfp_trucks", "")),
                ("BFP Ambulance", data.get("bfp_ambulance", "")),
                ("Non-BFP Ambulance", data.get("non_bfp_ambulance", "")),
                ("BFP Rescue Trucks", data.get("bfp_rescue_trucks", "")),
                ("Non-BFP Rescue Trucks", data.get("non_bfp_rescue_trucks", "")),
                ("Other Vehicles", data.get("other_vehicles", "")),
                ("SCBA Used", data.get("scba_used", "")),
                ("Water Tankers", data.get("water_tankers", "")),
            ],
        ),
        (
            "D. FIRE ALARM LEVEL",
            [
                ("1ST ALARM", data.get("alarm_1st", "")),
                ("2ND ALARM", data.get("alarm_2nd", "")),
                ("3RD ALARM", data.get("alarm_3rd", "")),
                ("4TH ALARM", data.get("alarm_4th", "")),
                ("5TH ALARM", data.get("alarm_5th", "")),
                ("TASK FORCE ALPHA", data.get("alarm_tf_alpha", "")),
                ("TASK FORCE BRAVO", data.get("alarm_tf_bravo", "")),
                ("TASK FORCE CHARLIE", data.get("alarm_tf_charlie", "")),
                ("TASK FORCE DELTA", data.get("alarm_tf_delta", "")),
                ("GENERAL ALARM", data.get("alarm_general", "")),
                ("FIRE UNDER CONTROL", data.get("alarm_fuc", "")),
                ("FIRE OUT", data.get("alarm_fo", "")),
                ("ICP Status", data.get("icp_status", "")),
                ("ICP Location", data.get("icp_location", "")),
            ],
        ),
        (
            "E. PROFILE OF CASUALTIES",
            [
                ("Civilian Injured (Male)", data.get("civilian_injured_male", "")),
                ("Civilian Injured (Female)", data.get("civilian_injured_female", "")),
                ("Firefighter Injured (Male)", data.get("firefighter_injured_male", "")),
                ("Firefighter Injured (Female)", data.get("firefighter_injured_female", "")),
                ("Civilian Fatality (Male)", data.get("civilian_deaths_male", "")),
                ("Civilian Fatality (Female)", data.get("civilian_deaths_female", "")),
                ("Firefighter Fatality (Male)", data.get("firefighter_deaths_male", "")),
                ("Firefighter Fatality (Female)", data.get("firefighter_deaths_female", "")),
            ],
        ),
        (
            "F. PERSONNEL ON DUTY",
            [
                ("Engine Commander", data.get("engine_commander", "")),
                ("Shift-in-Charge", data.get("shift_in_charge", "")),
                ("Nozzleman", data.get("nozzleman", "")),
                ("Lineman", data.get("lineman", "")),
                ("Engine Crew", data.get("engine_crew", "")),
                ("Driver / DPO", data.get("driver_dpo", "")),
                ("Safety Officer", data.get("safety_officer", "")),
                ("Fire Investigator", data.get("investigator", "")),
            ],
        ),
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([f"AFOR Export – Incident {data.get('incident_id', 'N/A')}"])
        writer.writerow([f"Reference: {data.get('reference_number', 'N/A')}"])
        writer.writerow([])

        for section_name, rows in sections:
            writer.writerow([section_name])
            writer.writerow(["Field", "Value"])
            for field, val in rows:
                writer.writerow([field, val])
            writer.writerow([])


def _insert_export_log(
    db: Any,
    *,
    user_id: str,
    export_format: str,
    export_type: str,
    filters: dict[str, Any],
    columns: list[str],
    task_id: str | None,
    path: str,
    content_type: str,
    row_count: int,
) -> None:
    db.execute(
        text("""
            INSERT INTO wims.analytics_export_log
                (user_id, format, filters_json, columns_json, row_count,
                 task_id, file_path, file_name, content_type, export_type)
            VALUES
                (:user_id, :format, CAST(:filters_json AS jsonb), CAST(:columns_json AS jsonb),
                 :row_count, :task_id, :file_path, :file_name, :content_type, :export_type)
        """),
        {
            "user_id": user_id,
            "format": export_format,
            "export_type": export_type,
            "filters_json": json.dumps(filters or {}),
            "columns_json": json.dumps(columns),
            "row_count": row_count,
            "task_id": task_id,
            "file_path": path,
            "file_name": os.path.basename(path),
            "content_type": content_type,
        },
    )
    db.commit()


def _export(
    *,
    task_id: str | None,
    user_id: str,
    filters: dict[str, Any],
    columns: list[str],
    export_format: str,
    extension: str,
    content_type: str,
    writer: Callable[[str, list[dict[str, Any]]], None],
    incident_ids: list[int] | None = None,
    export_type: str = "analytics",
    data_provider: Callable[[Session, int], dict[str, Any]] | None = None,
) -> str:
    valid_cols = _valid_columns(columns)
    logger.info(
        "%s export started: user_id=%s, filters=%s, columns=%s",
        export_format.upper(),
        user_id,
        filters,
        valid_cols,
    )

    db = get_session()
    try:
        set_rls_context(db, uuid.UUID(user_id))
        if export_type == "analyst":
            rows = get_analyst_export_rows(db, filters or {}, valid_cols, incident_ids)
        else:
            rows = get_export_rows(db, filters or {}, valid_cols)
        os.makedirs(EXPORT_DIR, exist_ok=True)
        path = os.path.join(EXPORT_DIR, f"analytics_export_{uuid.uuid4().hex[:12]}.{extension}")
        writer(path, rows, valid_cols)
        _insert_export_log(
            db,
            user_id=user_id,
            export_format=export_format,
            export_type=export_type,
            filters=filters or {},
            columns=valid_cols,
            task_id=task_id,
            path=path,
            content_type=content_type,
            row_count=len(rows),
        )
    finally:
        db.close()

    logger.info("%s export complete: %d rows -> %s", export_format.upper(), len(rows), path)
    return path


def _export_single_incident(
    *,
    task_id: str | None,
    user_id: str,
    incident_id: int,
    export_format: str,
    extension: str,
    content_type: str,
    writer: Callable[[str, dict[str, Any]], None],
) -> str:
    """Export a single incident using the AFOR template layout."""
    logger.info(
        "AFOR %s export started: user_id=%s, incident_id=%s",
        export_format.upper(),
        user_id,
        incident_id,
    )

    db = get_session()
    try:
        set_rls_context(db, uuid.UUID(user_id))
        data = get_incident_export_data(db, incident_id)
        os.makedirs(EXPORT_DIR, exist_ok=True)
        path = os.path.join(EXPORT_DIR, f"afor_{incident_id}_{uuid.uuid4().hex[:12]}.{extension}")
        writer(path, data)
        _insert_export_log(
            db,
            user_id=user_id,
            export_format=export_format,
            export_type="afor",
            filters={"incident_id": incident_id},
            columns=[],
            task_id=task_id,
            path=path,
            content_type=content_type,
            row_count=1,
        )
    finally:
        db.close()

    logger.info("AFOR %s export complete: %s", export_format.upper(), path)
    return path


# ─── Bulk Export Tasks (unchanged — for dashboard queue) ────────────────────────


def _write_csv_bulk(path: str, rows: list[dict[str, Any]], _columns: list[str]) -> None:
    _write_csv(path, rows, _columns)


def _write_xlsx_bulk(path: str, rows: list[dict[str, Any]], _columns: list[str]) -> None:
    _write_xlsx(path, rows, _columns)


def _write_pdf_bulk(path: str, rows: list[dict[str, Any]], _columns: list[str]) -> None:
    _write_pdf(path, rows, _columns)


@celery_app.task(bind=True, name="tasks.exports.export_incidents_csv")
def export_incidents_csv_task(
    self, user_id: str, filters: dict[str, Any], columns: list[str]
) -> str:
    """Export verified, non-archived incidents to a real CSV file."""
    return _export(
        task_id=getattr(self.request, "id", None),
        user_id=user_id,
        filters=filters,
        columns=columns,
        export_format="csv",
        extension="csv",
        content_type="text/csv",
        writer=_write_csv_bulk,
    )


@celery_app.task(bind=True, name="tasks.exports.export_incidents_pdf")
def export_incidents_pdf_task(
    self, user_id: str, filters: dict[str, Any], columns: list[str]
) -> str:
    """Export verified, non-archived incidents to a real PDF file."""
    return _export(
        task_id=getattr(self.request, "id", None),
        user_id=user_id,
        filters=filters,
        columns=columns,
        export_format="pdf",
        extension="pdf",
        content_type="application/pdf",
        writer=_write_pdf_bulk,
    )


@celery_app.task(bind=True, name="tasks.exports.export_incidents_excel")
def export_incidents_excel_task(
    self, user_id: str, filters: dict[str, Any], columns: list[str]
) -> str:
    """Export verified, non-archived incidents to a real XLSX file."""
    return _export(
        task_id=getattr(self.request, "id", None),
        user_id=user_id,
        filters=filters,
        columns=columns,
        export_format="excel",
        extension="xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        writer=_write_xlsx_bulk,
    )


@celery_app.task(bind=True, name="tasks.exports.export_analyst_incidents")
def export_analyst_incidents_task(
    self,
    user_id: str,
    filters: dict[str, Any],
    columns: list[str],
    incident_ids: list[int] | None = None,
    format: str = "csv",
    export_mode: str = "bulk",
) -> str:
    """
    Export analyst incident selections.

    export_mode controls output layout:
      'bulk'    — flat tabular (original behaviour, matches queue page CSV)
      'afor'   — AFOR section-based layout (for single-incident detail export)
    """
    incident_id = filters.get("incident_id") if isinstance(filters, dict) else None

    if export_mode == "afor" and incident_id:
        normalized_format = (format or "csv").lower()
        writers = {
            "csv": ("csv", "text/csv", _write_afor_csv),
            "pdf": ("pdf", "application/pdf", _write_afor_pdf),
            "excel": (
                "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                _write_afor_excel,
            ),
        }
        if normalized_format not in writers:
            raise ValueError(f"Unsupported export format: {format}")
        ext, ctype, writer = writers[normalized_format]
        return _export_single_incident(
            task_id=getattr(self.request, "id", None),
            user_id=user_id,
            incident_id=int(incident_id),
            export_format=normalized_format,
            extension=ext,
            content_type=ctype,
            writer=writer,
        )

    normalized_format = (format or "csv").lower()
    writers = {
        "csv": ("csv", "text/csv", _write_csv),
        "pdf": ("pdf", "application/pdf", _write_pdf),
        "excel": (
            "xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            _write_xlsx,
        ),
    }
    if normalized_format not in writers:
        raise ValueError(f"Unsupported export format: {format}")

    extension, content_type, writer = writers[normalized_format]
    return _export(
        task_id=getattr(self.request, "id", None),
        user_id=user_id,
        filters=filters,
        columns=columns,
        export_format=normalized_format,
        extension=extension,
        content_type=content_type,
        writer=writer,
        incident_ids=sorted(set(incident_ids or [])) or None,
        export_type="analyst",
    )


def _write_pdf(path: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path, pagesize=landscape(letter), title="WIMS-BFP AFOR Export")
    story = [
        Paragraph("WIMS-BFP AFOR Incident Export", styles["Title"]),
        Paragraph(f"Rows: {len(rows)}", styles["Normal"]),
        Spacer(1, 12),
    ]
    table_data = [columns]
    table_data.extend([[_serialize_value(row.get(col)) for col in columns] for row in rows])
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7f1d1d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    doc.build(story)

